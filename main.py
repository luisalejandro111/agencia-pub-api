from fastapi import FastAPI, Request, Depends, WebSocket, WebSocketDisconnect, HTTPException
from starlette.responses import RedirectResponse
from fastapi import HTTPException, status
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, cast, Date
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import engine, AsyncSessionLocal, get_db
from app import models, auth, crud
from datetime import datetime, date, timedelta
from itsdangerous import URLSafeSerializer
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import pandas as pd
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import selectinload
from typing import Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta, date
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from datetime import datetime, timedelta, date
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from fastapi import status
from sqlalchemy import func
from sqlalchemy import select, delete 
from fastapi import Request, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models import Cliente
from app.models import Presupuesto, TipoTrabajo, Cliente, ItemPresupuestoCliente, ItemPresupuestoInterno
from sqlalchemy import select, func
import uuid
from app.models import Presupuesto, TipoTrabajo, Cliente, ItemPresupuesto
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse
from weasyprint import HTML, CSS
import io
import base64
import os
from fastapi.responses import StreamingResponse
from fastapi import Request, Depends, HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, datetime
from typing import  Optional
from urllib.parse import quote 
import os
import uuid
from fastapi import UploadFile, File
from mimetypes import guess_type
from starlette.datastructures import UploadFile
from app.models import Base
import os
from dotenv import load_dotenv
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from werkzeug.utils import secure_filename 
import re 
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.exceptions import HTTPException
from urllib.parse import unquote
from app.models import ArchivoTrabajo
from typing import List
from fastapi import File, UploadFile
from sqlalchemy import Column, Integer, String, Float, DateTime, DECIMAL, ForeignKey, Numeric
from sqlalchemy import select, func, and_, or_, text, case
from sqlalchemy.orm import selectinload, joinedload
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import FastAPI, Request, Response, Depends, Form, UploadFile, File, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from app.models import (
    Trabajo, 
    Asignacion, 
    Empleado, 
    Rol, 
    MaterialUsado, 
    Cliente, 
    Presupuesto,
    ArchivoTrabajo,
    Usuario,
    Asistencia, 
  
)
from app.models import  ServicioExterno
from datetime import datetime
from app.models import MaterialInventario
from app.models import CategoriaInventario  
from app.models import Proveedor
from app.models import EntradaInventario
from app.models import MovimientoInventario
from sqlalchemy.orm import selectinload
from app.models import ActivoFijo, CategoriaActivoFijo, MovimientoActivoFijo
from app.models import DeduccionEmpleado, Empleado, Trabajo, Usuario, SalidaMaterial, PagoSemanal, IngresoDiario, CategoriaGasto, SubcategoriaGasto
from app.forecasting import generar_alerta_stock
import os
import time
from datetime import date, datetime 
from app.models import Trabajo, Empleado, Prestamo, GastoDiario, CierreCaja 
from decimal import Decimal
from sqlalchemy import Date as SQLDate
from fastapi.staticfiles import StaticFiles
from app.models import Cliente, Trabajo, Usuario
from app.routes.usuarios import router as usuarios_router
from starlette.middleware.sessions import SessionMiddleware
from app.models import Usuario
from app.routes import asistencia
from app.routers import deudas
from app.api import router as api_router
from app.services.currency_service import CurrencyService
from app.routers import tasas_cambio
from app.routes.auth import router as auth_router
from app.middleware import RateLimitMiddleware
from app.routes.install import router as install_router
from app.auth import get_current_user_from_session
from dotenv import load_dotenv


# Configurar zona horaria de Venezuela
os.environ['TZ'] = 'America/Caracas'
time.tzset()


app = FastAPI(title="Agencia de Publicidad API")


ENV = os.getenv("ENVIRONMENT", "production")
RATE_LIMIT = 600 if ENV == "test" else 60

print(f"🔍 ENVIRONMENT={ENV}, RATE_LIMIT={RATE_LIMIT}") 

app.add_middleware(RateLimitMiddleware, calls_per_minute=RATE_LIMIT)

# main.py - Al inicio
from app.logger import logger
import time

# ============================================
# GENERAR NÚMERO DE PRESUPUESTO
# ============================================
async def generar_numero_presupuesto(db: AsyncSession) -> str:
    """Genera número de presupuesto secuencial: PRE01, PRE02, ..."""
    from sqlalchemy import func, select
    import re
    
    # Buscar el último número de presupuesto
    result = await db.execute(
        select(func.max(Presupuesto.numero_presupuesto))
    )
    ultimo = result.scalar()
    
    if ultimo:
        # Extraer el número (ej: PRE05 -> 5)
        numeros = re.findall(r'\d+', ultimo)
        if numeros:
            siguiente = int(numeros[-1]) + 1
        else:
            siguiente = 1
    else:
        siguiente = 1
    
    # Formato PRE01, PRE02, PRE03, ...
    return f"PRE{str(siguiente).zfill(2)}"

# Middleware para loggear todas las peticiones
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start_time = time.time()
    response = await call_next(request)
    process_time = time.time() - start_time
    
    logger.info(
        f"{request.method} {request.url.path} - "
        f"Status: {response.status_code} - "
        f"Tiempo: {process_time:.3f}s"
    )
    return response

app.include_router(usuarios_router)

app.include_router(asistencia.router)

SECRET_KEY = os.getenv("SECRET_KEY", "SECRET_KEY_CAMBIAR_EN_PRODUCCION")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

app.include_router(tasas_cambio.router)

app.include_router(auth_router)

app.include_router(install_router)

# Incluir routers
app.include_router(api_router, prefix="/api/v1")
app.include_router(deudas.router, prefix="/api/v1")

# Ruta para la plantilla
@app.get("/cuentas-por-cobrar")
async def cuentas_por_cobrar(request: Request):
    return templates.TemplateResponse("cuentas_por_cobrar.html", {"request": request})



BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")

app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="app/templates")

def parse_decimal(value, default=0.0):
    """Convierte un valor a float de forma segura"""
    if value is None or value == "" or value == "None":
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default



# En tu main.py, cerca de donde defines la app
def format_decimal(value, decimals=2):
    if value is None:
        return "0.00"
    try:
        return f"{float(value):.{decimals}f}"
    except (ValueError, TypeError):
        return "0.00"

# Registrar el filtro en Jinja2
templates.env.filters["format_decimal"] = format_decimal


@app.get("/uploads/trabajos/{trabajo_id}/{filename}")
async def get_archivo_trabajo(trabajo_id: int, filename: str, db: AsyncSession = Depends(get_db)):
    print(f"🔍 Solicitud de descarga: trabajo_id={trabajo_id}, filename={filename}")
    
    filename_decoded = unquote(filename)
    print(f"🔍 Filename decodificado: {filename_decoded}")
    
    # Verificar en BD
    archivo_query = await db.execute(
        select(ArchivoTrabajo).where(
            ArchivoTrabajo.trabajo_id == trabajo_id,
            ArchivoTrabajo.nombre_guardado == filename_decoded
        )
    )
    archivo = archivo_query.scalar_one_or_none()
    if not archivo:
        print(f"❌ Archivo no encontrado en BD: trabajo_id={trabajo_id}, nombre_guardado={filename_decoded}")
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    filepath = archivo.ruta_completa
    print(f"🔍 Ruta de BD: {filepath}")
    print(f"🔍 Existe en disco: {os.path.exists(filepath)}")
    
    if not os.path.exists(filepath):
        print(f"❌ Archivo no encontrado en disco: {filepath}")
        raise HTTPException(status_code=404, detail="Archivo no encontrado en disco")
    
    return FileResponse(filepath, filename=archivo.nombre_original)

# Cargar variables de entorno
load_dotenv()

# Configuración de base de datos
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL no está definida en el archivo .env")

# Para operaciones asíncronas (usa asyncpg)
ASYNC_DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://")

# Motores de base de datos

async_engine = create_async_engine(ASYNC_DATABASE_URL, echo=True)
AsyncSessionLocal = sessionmaker(bind=async_engine, class_=AsyncSession, expire_on_commit=False)

def get_logo_base64():
    """Convierte el logo a base64 para embeberlo en el PDF"""
    import os
    current_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(current_dir, "static", "images", "logo.png")
    
    print(f"🔍 Buscando logo en: {logo_path}")
    
    if not os.path.exists(logo_path):
        print("❌ Logo no encontrado")
        return None
    
    try:
        with open(logo_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            print("✅ Logo convertido a base64 exitosamente")
            return encoded_string
    except Exception as e:
        print(f"❌ Error al leer el logo: {e}")
        return None


async def calcular_comisiones_trabajo(trabajo, sync_db=None, async_db=None):
    """Calcula comisiones por rol y las divide entre empleados asignados a ese rol."""
    from sqlalchemy import select
    from app import models
    
    comisiones_detalle = []
    total_comisiones = 0.0
    monto_total = float(trabajo.monto_total)
    
    if async_db:
        # Obtener todas las asignaciones del trabajo
        asignaciones = (await async_db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo.id)
        )).scalars().all()
        
        # Agrupar asignaciones por rol
        from collections import defaultdict
        asignaciones_por_rol = defaultdict(list)
        for a in asignaciones:
            asignaciones_por_rol[a.rol_id].append(a)
        
        # Calcular comisión por rol
        for rol_id, asignaciones_rol in asignaciones_por_rol.items():
            if not asignaciones_rol:
                continue
                
            # Tomar la configuración de la primera asignación del rol
            primera_asignacion = asignaciones_rol[0]
            tipo_comision = primera_asignacion.tipo_comision
            valor_comision = float(primera_asignacion.valor_comision or 0)
            
            if valor_comision <= 0:
                continue
            
            # Calcular comisión TOTAL para este rol
            comision_rol_total = 0.0
            if tipo_comision == "porcentaje":
                comision_rol_total = monto_total * (valor_comision / 100)
            elif tipo_comision == "metro":
                metros = float(getattr(trabajo, 'metros_cuadrados', 0) or 0)
                comision_rol_total = metros * valor_comision
            elif tipo_comision == "unidad":
                unidades = int(getattr(trabajo, 'unidades', 0) or 0)
                comision_rol_total = unidades * valor_comision
            
            if comision_rol_total <= 0:
                continue
            
            # Dividir entre el número de empleados en este rol
            num_empleados = len(asignaciones_rol)
            comision_por_empleado = comision_rol_total / num_empleados
            
            # Asignar a cada empleado
            for asignacion in asignaciones_rol:
                emp_result = await async_db.execute(
                    select(models.Empleado.nombre_completo).where(
                        models.Empleado.id == asignacion.empleado_id
                    )
                )
                emp_nombre = emp_result.scalar() or f"Empleado ID {asignacion.empleado_id}"
                
                comisiones_detalle.append({
                    "nombre": f"{emp_nombre} - {trabajo.nombre_trabajo}",
                    "comision": round(comision_por_empleado, 2),
                    "rol_id": rol_id
                })
            
            total_comisiones += comision_rol_total
    
    elif sync_db:
        # Versión sincrónica (para reportes automáticos)
        asignaciones = sync_db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo.id)
        ).scalars().all()
        
        from collections import defaultdict
        asignaciones_por_rol = defaultdict(list)
        for a in asignaciones:
            asignaciones_por_rol[a.rol_id].append(a)
        
        for rol_id, asignaciones_rol in asignaciones_por_rol.items():
            if not asignaciones_rol:
                continue
                
            primera_asignacion = asignaciones_rol[0]
            tipo_comision = primera_asignacion.tipo_comision
            valor_comision = float(primera_asignacion.valor_comision or 0)
            
            if valor_comision <= 0:
                continue
            
            comision_rol_total = 0.0
            if tipo_comision == "porcentaje":
                comision_rol_total = monto_total * (valor_comision / 100)
            elif tipo_comision == "metro":
                metros = float(getattr(trabajo, 'metros_cuadrados', 0) or 0)
                comision_rol_total = metros * valor_comision
            elif tipo_comision == "unidad":
                unidades = int(getattr(trabajo, 'unidades', 0) or 0)
                comision_rol_total = unidades * valor_comision
            
            if comision_rol_total <= 0:
                continue
            
            num_empleados = len(asignaciones_rol)
            comision_por_empleado = comision_rol_total / num_empleados
            
            for asignacion in asignaciones_rol:
                emp_nombre = sync_db.execute(
                    select(models.Empleado.nombre_completo).where(
                        models.Empleado.id == asignacion.empleado_id
                    )
                ).scalar() or f"Empleado ID {asignacion.empleado_id}"
                
                comisiones_detalle.append({
                    "nombre": f"{emp_nombre} - {trabajo.nombre}",
                    "comision": round(comision_por_empleado, 2),
                    "rol_id": rol_id
                })
            
            total_comisiones += comision_rol_total
    
    return {
        "detalle": comisiones_detalle,
        "total_comisiones": round(total_comisiones, 2)
    }

def generar_html_comisiones(comisiones):
    if not comisiones:
        return '<p class="text-gray-500">No hay comisiones registradas.</p>'
    html = ""
    for c in comisiones:
        html += f'<div class="flex justify-between py-1 border-b"><span>{c["nombre"]}</span><span class="font-medium">${c["comision"]:.2f}</span></div>'
    return html

def generar_html_gastos(gastos):
    if not gastos:
        return '<p class="text-gray-500">No hay gastos registrados.</p>'
    html = ""
    for g in gastos:
        categoria = f"({g['categoria']})" if g["categoria"] else ""
        html += f'<div class="flex justify-between py-1 border-b"><span>{g["descripcion"]} {categoria}</span><span class="font-medium">${g["monto"]:.2f}</span></div>'
    return html

def generar_html_prestamos(prestamos):
    if not prestamos:
        return '<p class="text-gray-500">No hay préstamos otorgados.</p>'
    html = ""
    for p in prestamos:
        html += f'<div class="flex justify-between py-1 border-b"><span>{p["empleado"]} - {p["motivo"]}</span><span class="font-medium">${p["monto"]:.2f}</span></div>'
    return html


@app.on_event("startup")
async def startup():
    async with engine.begin() as conn:
        await conn.run_sync(models.Base.metadata.create_all)
 

# === AUTENTICACIÓN ===

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {
        "request": request,
        "error": None,
        "now": datetime.now()
    })
@app.post("/login")
async def login_submit(request: Request, db: AsyncSession = Depends(get_db)):
    form = await request.form()
    username_or_email = form.get("username_or_email")  # 🔥 Cambio de "email" a "username_or_email"
    password = form.get("password")

    print(f"🔍 Intentando login con: {username_or_email}")

    # Validar que los campos no estén vacíos
    if not username_or_email or not password:
        print("❌ Usuario/Email o contraseña vacíos")
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Usuario/Email y contraseña son requeridos",
            "now": datetime.now()
        })

    # 🔥 Autenticar usando la nueva función que acepta username o email
    user = await auth.authenticate_user(db, username_or_email, password)
    
    if not user:
        print(f"❌ Usuario no autenticado para: {username_or_email}")
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Usuario/Email o contraseña incorrectos",
            "now": datetime.now()
        })

    print(f"✅ Usuario autenticado: {user.nombre} (Username: {user.username}) - Rol: {user.rol}")

    # 🔥 Usar username para la sesión (siempre existe, a diferencia del email que es opcional)
    s = URLSafeSerializer(SECRET_KEY, salt="session")
    session_id = s.dumps(user.username)  # 🔥 Cambio: usar username en lugar de email

    response = RedirectResponse(url="/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    response.set_cookie(
        key="session_id",
        value=session_id,
        httponly=True,
        secure=False,  # En producción con HTTPS poner True
        samesite="lax"
    )
    print(f"🍪 Cookie session_id creada para usuario: {user.username}")
    print(f"➡️ Redirigiendo a /dashboard")
    return response

@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie("session_id")
    return response

# === DASHBOARD ===
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    from datetime import datetime, timedelta
    
    # Calcular fechas de la semana actual (lunes a domingo)
    hoy = datetime.now().date()
    dia_semana = hoy.weekday()  # Lunes=0, Domingo=6
    inicio_semana = hoy - timedelta(days=dia_semana)
    fin_semana = inicio_semana + timedelta(days=6)
    
    # Inicializar variables
    ingreso_bruto = 0.0
    caja_semanal = 0.0
    empleados_activos = 0
    trabajos_pendientes = 0
    ultimos_trabajos = []

    clientes_pendientes = []
    monto_pendiente_usd = 0.0
    
    try:
        # === INGRESO BRUTO (trabajos completados esta semana) ===
        trabajos_completados = await db.execute(
            select(Trabajo)
            .where(Trabajo.estado == 'completado')
            .where(Trabajo.fecha_inicio >= inicio_semana)
            .where(Trabajo.fecha_inicio <= fin_semana)
        )
        ingreso_bruto = sum(float(t.monto_total or 0) for t in trabajos_completados.scalars().all())
        
        # === CAJA SEMANAL (pagos recibidos esta semana) ===
        # Esto dependerá de cómo almacenes los pagos en tu sistema
        # Por ahora, usamos el mismo valor que ingreso_bruto como ejemplo
        caja_semanal = ingreso_bruto
        
        # === EMPLEADOS ACTIVOS ===
        empleados_result = await db.execute(
            select(Empleado).where(Empleado.activo == True)
        )
        empleados_activos = len(empleados_result.scalars().all())
        
        # === TRABAJOS PENDIENTES ===
        pendientes_result = await db.execute(
            select(Trabajo)
            .where(Trabajo.estado.in_(['pendiente', 'en_ejecucion']))
        )
        trabajos_pendientes = len(pendientes_result.scalars().all())
        
        # === ÚLTIMOS TRABAJOS FINALIZADOS ===
        ultimos_trabajos_result = await db.execute(
            select(Trabajo)
            .where(Trabajo.estado.in_(['finalizado', 'entregado', 'completado']))
            .order_by(Trabajo.fecha_inicio.desc())
            .limit(5)
        )
        ultimos_trabajos = ultimos_trabajos_result.scalars().all()
        
    except Exception as e:
        print(f"Error al cargar datos del dashboard: {e}")
        # Continuar con valores por defecto si hay error
    
    # Mensaje HTML (vacío por defecto, o desde parámetros URL)
    mensaje_html = ""
    mensaje_param = request.query_params.get("mensaje")
    error_param = request.query_params.get("error")
    
    if mensaje_param:
        mensaje_html = f'<div class="mb-6 p-4 bg-green-50 text-green-700 rounded-lg border border-green-200 flex items-center gap-2"><i class="fas fa-check-circle"></i> {mensaje_param}</div>'
    elif error_param:
        mensaje_html = f'<div class="mb-6 p-4 bg-red-50 text-red-700 rounded-lg border border-red-200 flex items-center gap-2"><i class="fas fa-exclamation-circle"></i> {error_param}</div>'
    
    return templates.TemplateResponse("dashboard/dashboard.html", {
        "request": request,
        "user": user,
        "mensaje_html": mensaje_html,
        "ingreso_bruto": ingreso_bruto,
        "caja_semanal": caja_semanal,
        "empleados_activos": empleados_activos,
        "trabajos_pendientes": trabajos_pendientes,
        "ultimos_trabajos": ultimos_trabajos,
        "inicio_semana": inicio_semana,
        "fin_semana": fin_semana,
        "clientes_pendientes": clientes_pendientes, 
        "monto_pendiente_usd": monto_pendiente_usd
    })



# === FUNCIÓN AUXILIAR: Generar datos del reporte diario ===
@app.get("/caja/diario", response_class=HTMLResponse)
async def reporte_diario(
    request: Request,
    fecha: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")

    try:
        if not fecha:
            from datetime import date as dt_date
            fecha = dt_date.today().isoformat()
    
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()

        # === GASTOS ===
        gastos = (await db.execute(
            select(models.GastoDiario).where(models.GastoDiario.fecha == fecha_obj)
        )).scalars().all()
        gastos_detalle = []
        gasto_total = 0.0
        for gasto in gastos:
            monto = float(gasto.monto)
            gasto_total += monto
            gastos_detalle.append({
                "descripcion": gasto.descripcion or "Sin descripción",
                "categoria": gasto.categoria or "General",
                "monto": round(monto, 2)
            })

        # === PRÉSTAMOS ===
        prestamos = (await db.execute(
            select(
                models.Prestamo.id,
                models.Prestamo.monto,
                models.Prestamo.descripcion,
                models.Empleado.nombre_completo
            )
            .select_from(models.Prestamo)
            .join(models.Empleado, models.Prestamo.empleado_id == models.Empleado.id)
            .where(models.Prestamo.fecha == fecha_obj)
        )).fetchall()
        prestamos_detalle = []
        prestamo_total = 0.0
        for prestamo in prestamos:
            monto = float(prestamo.monto)
            prestamo_total += monto
            prestamos_detalle.append({
                "empleado": prestamo.nombre_completo,
                "monto": round(monto, 2),
                "descripcion": prestamo.descripcion or "-"
            })

        # === TRABAJOS FINALIZADOS ===
        trabajos = (await db.execute(
            select(models.Trabajo).where(
                models.Trabajo.estado == "finalizado",
                models.Trabajo.fecha_inicio == fecha_obj
            )
        )).scalars().all()
        ingreso_bruto = 0.0
        comisiones_detalle = []
        total_comisiones = 0.0
        trabajos_info = []
        
        for trabajo in trabajos:
            monto_total = float(trabajo.monto_total)
            ingreso_bruto += monto_total
            trabajos_info.append({
                "nombre": trabajo.nombre,
                "monto_total": monto_total,
                "metros_cuadrados": float(trabajo.metros_cuadrados) if trabajo.metros_cuadrados else None,
                "unidades": trabajo.unidades if trabajo.unidades else None
            })
            
            # Calcular comisiones
            calc = await calcular_comisiones_trabajo(trabajo, async_db=db)
            total_comisiones += calc["total_comisiones"]
            comisiones_detalle.extend(calc["detalle"])

        # === RESULTADOS ===
        ingreso_neto = ingreso_bruto - total_comisiones
        caja_diaria = ingreso_neto - gasto_total

        return templates.TemplateResponse("reportes/diario.html", {
            "request": request,
            "datos": {
                "fecha": fecha_obj.isoformat(),
                "ingreso_bruto": round(ingreso_bruto, 2),
                "total_comisiones": round(total_comisiones, 2),
                "ingreso_neto": round(ingreso_neto, 2),
                "gasto_total": round(gasto_total, 2),
                "prestamo_total": round(prestamo_total, 2),
                "caja_diaria": round(caja_diaria, 2),
                "trabajos": trabajos_info,
                "comisiones_detalle": comisiones_detalle,  # ← ¡Incluido!
                "gastos_detalle": gastos_detalle,          # ← ¡Incluido!
                "prestamos_detalle": prestamos_detalle     # ← ¡Incluido!
            }
        })
        
    except Exception as e:
        print(f"ERROR EN REPORTE DIARIO: {str(e)}")
        return templates.TemplateResponse("reportes/diario.html", {
            "request": request,
            "datos": {
                "fecha": fecha_obj.isoformat() if 'fecha_obj' in locals() else "",
                "ingreso_bruto": 0.0,
                "total_comisiones": 0.0,
                "ingreso_neto": 0.0,
                "gasto_total": 0.0,
                "prestamo_total": 0.0,
                "caja_diaria": 0.0,
                "trabajos": [],
                "comisiones_detalle": [],
                "gastos_detalle": [],
                "prestamos_detalle": []
            }
        })

        # === FUNCIONES AUXILIARES PARA HTML ===
        def generar_html_comisiones(comisiones):
            if not comisiones:
                return '<p class="text-gray-500">No hay comisiones registradas.</p>'
            html = ""
            html += f'<div class="flex justify-between py-1 border-b"><span>{c["nombre"]}</span><span class="font-medium">${c["comision"]:.2f}</span></div>'
            return html

        def generar_html_gastos(gastos):
            if not gastos:
                return '<p class="text-gray-500">No hay gastos registrados.</p>'
            html = ""
            for g in gastos:
                categoria = f"({g['categoria']})" if g["categoria"] else ""
                html += f'<div class="flex justify-between py-1 border-b"><span>{g["descripcion"]} {categoria}</span><span class="font-medium">${g["monto"]:.2f}</span></div>'
            return html

        def generar_html_prestamos(prestamos):
            if not prestamos:
                return '<p class="text-gray-500">No hay préstamos otorgados.</p>'
            html = ""
            for p in prestamos:
                html += f'<div class="flex justify-between py-1 border-b"><span>{p["empleado"]} - {p["motivo"]}</span><span class="font-medium">${p["monto"]:.2f}</span></div>'
            return html

        # === HTML ===
        html_comisiones = generar_html_comisiones(reporte['detalle']['comisiones'])
        html_gastos = generar_html_gastos(reporte['detalle']['gastos'])
        html_prestamos = generar_html_prestamos(reporte['detalle']['prestamos'])

        html_content = f"""
        <html>
            <head>
                <title>Reporte Diario de Caja - Agencia Publicidad</title>
                <script src="https://cdn.tailwindcss.com"></script>
            </head>
            <body class="bg-gray-50 min-h-screen">
                <div class="max-w-4xl mx-auto p-6">
                    <a href="/dashboard" class="inline-block mb-6 text-indigo-600">&larr; Volver al dashboard</a>
                    
                    <div class="bg-white rounded-xl shadow-md p-6 mb-6">
                        <h1 class="text-2xl font-bold text-gray-800">💰 Reporte Diario de Caja</h1>
                        <p class="text-gray-600">Fecha: {reporte['fecha']}</p>
                    </div>


                    <!-- Resumen detallado -->
                    <div class="grid grid-cols-1 md:grid-cols-4 gap-4 mb-6">
                        <div class="bg-gray-50 rounded-lg p-4 text-center">
                            <p class="text-sm text-gray-700">Ingresos Brutos</p>
                            <p class="text-lg font-bold text-gray-800">${reporte['ingreso_bruto']:.2f}</p>
                        </div>
                        <div class="bg-rose-50 rounded-lg p-4 text-center">
                            <p class="text-sm text-rose-700">Comisiones</p>
                            <p class="text-lg font-bold text-rose-800">-${reporte['total_comisiones']:.2f}</p>
                        </div>
                        <div class="bg-green-50 rounded-lg p-4 text-center">
                            <p class="text-sm text-green-700">Ingresos Netos</p>
                            <p class="text-lg font-bold text-green-800">${reporte['ingresos']:.2f}</p>
                        </div>
                        <div class="bg-blue-50 rounded-lg p-4 text-center">
                            <p class="text-sm text-blue-700">Saldo Final</p>
                            <p class="text-lg font-bold {'text-green-800' if reporte['saldo'] >= 0 else 'text-red-800'}">
                                ${reporte['saldo']:.2f}
                            </p>
                        </div>
                    </div>

                    <!-- Detalles -->
                    <div class="bg-white rounded-xl shadow-md p-6 mb-6">
                        <h2 class="text-lg font-semibold text-gray-800 mb-3">📤 Comisiones Pagadas</h2>
                        {html_comisiones}
                    </div>

                    <div class="bg-white rounded-xl shadow-md p-6 mb-6">
                        <h2 class="text-lg font-semibold text-gray-800 mb-3">📤 Gastos</h2>
                        {html_gastos}
                    </div>

                    <div class="bg-white rounded-xl shadow-md p-6">
                        <h2 class="text-lg font-semibold text-gray-800 mb-3">💳 Préstamos</h2>
                        {html_prestamos}
                    </div>
                    <!-- Botones de exportación al final, centrados -->
<div class="mt-8 flex justify-center gap-4">
    <a href="/caja/diario/export/pdf?fecha={reporte['fecha']}" 
       class="px-5 py-2.5 bg-red-600 text-white rounded-lg hover:bg-red-700 shadow-md transition text-sm font-medium">
        📄 Exportar PDF
    </a>
    <a href="/caja/diario/export/excel?fecha={reporte['fecha']}" 
       class="px-5 py-2.5 bg-green-600 text-white rounded-lg hover:bg-green-700 shadow-md transition text-sm font-medium">
        📊 Exportar Excel
    </a>
</div>
                </div>
            </body>
        </html>
        """
        return HTMLResponse(html_content)

    except Exception as e:
        print("ERROR EN REPORTE DIARIO:", str(e))
        return HTMLResponse(f"""
        <html>
            <head><title>Error</title>
            <script src="https://cdn.tailwindcss.com"></script>
            </head>
            <body class="bg-gray-50 p-6">
                <div class="max-w-2xl mx-auto">
                    <h1 class="text-2xl text-red-700">Error al generar el reporte</h1>
                    <p class="mt-2">Detalles: {str(e)}</p>
                    <a href="/dashboard" class="mt-4 inline-block text-indigo-600">← Volver</a>
                </div>
            </body>
        </html>
        """)


@app.get("/caja/diario/exportar/pdf")
async def exportar_pdf(
    fecha: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    if not fecha:
        from datetime import date as dt_date
        fecha = dt_date.today().isoformat()
    
    from datetime import datetime
    fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    
    # Obtener los mismos datos que en Excel
    # ... (mismo código para obtener datos) ...
    
    gastos = (await db.execute(
        select(models.GastoDiario).where(models.GastoDiario.fecha == fecha_obj)
    )).scalars().all()
    
    prestamos = (await db.execute(
        select(
            models.Prestamo.id,
            models.Prestamo.monto,
            models.Prestamo.descripcion,
            models.Empleado.nombre_completo
        )
        .select_from(models.Prestamo)
        .join(models.Empleado, models.Prestamo.empleado_id == models.Empleado.id)
        .where(models.Prestamo.fecha == fecha_obj)
    )).fetchall()
    
    trabajos = (await db.execute(
        select(models.Trabajo).where(
            models.Trabajo.estado == "finalizado",
            models.Trabajo.fecha_inicio == fecha_obj
        )
    )).scalars().all()
    
    comisiones_detalle = []
    for trabajo in trabajos:
        calc = await calcular_comisiones_trabajo(trabajo, async_db=db)
        comisiones_detalle.extend(calc["detalle"])
    
    ingreso_bruto = sum(float(t.monto_total) for t in trabajos)
    total_comisiones = sum(c["comision"] for c in comisiones_detalle)
    gasto_total = sum(float(g.monto) for g in gastos)
    prestamo_total = sum(float(p.monto) for p in prestamos)
    caja_diaria = ingreso_bruto - total_comisiones - gasto_total
    
    # Crear PDF
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    elements.append(Paragraph(f"REPORTE DIARIO - {fecha}", title_style))
    
    # Resumen
    resumen_data = [
        ['Ingreso Bruto', 'Comisiones','Préstamos','Gastos','Caja Diaria'],
        [f"${ingreso_bruto:.2f}", f"${total_comisiones:.2f}", f"${prestamo_total:.2f}", 
         f"${gasto_total:.2f}", f"${caja_diaria:.2f}"]
    ]
    
    resumen_table = Table(resumen_data, colWidths=[80, 80, 80, 80, 80])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))
    
    # Trabajos
    if trabajos:
        elements.append(Paragraph("TRABAJOS FINALIZADOS", styles['Heading2']))
        trabajos_data = [['Nombre', 'Monto', 'm²', 'Unidades', 'Estado']]
        for t in trabajos:
            trabajos_data.append([
                t.nombre,
                f"${float(t.monto_total):.2f}",
                f"{float(t.metros_cuadrados):.2f}" if t.metros_cuadrados else "",
                str(t.unidades) if t.unidades else "",
                t.estado
            ])
        
        trabajos_table = Table(trabajos_data, colWidths=[120, 60, 40, 40, 60])
        trabajos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(trabajos_table)
        elements.append(Spacer(1, 20))
    
    # Comisiones
    if comisiones_detalle:
        elements.append(Paragraph("COMISIONES", styles['Heading2']))
        comisiones_data = [['Empleado - Trabajo', 'Monto']]
        for c in comisiones_detalle:
            comisiones_data.append([c["nombre"], f"${c['comision']:.2f}"])
        
        comisiones_table = Table(comisiones_data, colWidths=[200, 60])
        comisiones_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(comisiones_table)
        elements.append(Spacer(1, 20))
    
    # Gastos
    if gastos:
        elements.append(Paragraph("GASTOS", styles['Heading2']))
        gastos_data = [['Descripción', 'Categoría', 'Monto']]
        for g in gastos:
            gastos_data.append([
                g.descripcion or "Sin descripción",
                g.categoria or "General",
                f"${float(g.monto):.2f}"
            ])
        
        gastos_table = Table(gastos_data, colWidths=[150, 80, 60])
        gastos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(gastos_table)
        elements.append(Spacer(1, 20))
    
    # Préstamos
    if prestamos:
        elements.append(Paragraph("PRÉSTAMOS", styles['Heading2']))
        prestamos_data = [['Empleado', 'Motivo', 'Monto']]
        for p in prestamos:
            prestamos_data.append([
                p.nombre_completo,
                p.descripcion or "-",
                f"${float(p.monto):.2f}"
            ])
        
        prestamos_table = Table(prestamos_data, colWidths=[120, 120, 60])
        prestamos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(prestamos_table)
    
    # Generar PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_diario_{fecha}.pdf"}
    )


@app.get("/caja/diario/exportar/excel")
async def exportar_excel(
    fecha: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    # Validar y obtener fecha
    if not fecha:
        from datetime import date as dt_date
        fecha = dt_date.today().isoformat()
    
    from datetime import datetime
    fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    
    # Obtener datos (reutiliza la lógica de reporte_diario)
    # ... (mismo código para obtener gastos, prestamos, trabajos, comisiones) ...
    
    # === GASTOS ===
    gastos = (await db.execute(
        select(models.GastoDiario).where(models.GastoDiario.fecha == fecha_obj)
    )).scalars().all()
    
    # === PRÉSTAMOS ===
    prestamos = (await db.execute(
        select(
            models.Prestamo.id,
            models.Prestamo.monto,
            models.Prestamo.descripcion,
            models.Empleado.nombre_completo
        )
        .select_from(models.Prestamo)
        .join(models.Empleado, models.Prestamo.empleado_id == models.Empleado.id)
        .where(models.Prestamo.fecha == fecha_obj)
    )).fetchall()
    
    # === TRABAJOS ===
    trabajos = (await db.execute(
        select(models.Trabajo).where(
            models.Trabajo.estado == "finalizado",
            models.Trabajo.fecha_inicio == fecha_obj
        )
    )).scalars().all()
    
    # === COMISIONES ===
    comisiones_detalle = []
    for trabajo in trabajos:
        calc = await calcular_comisiones_trabajo(trabajo, async_db=db)
        comisiones_detalle.extend(calc["detalle"])
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {fecha}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # === RESUMEN ===
    ws.merge_cells('A1:E1')
    ws['A1'] = f"REPORTE DIARIO - {fecha}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    resumen_data = [
        ["Ingreso Bruto", "Comisiones", "Préstamos", "Gastos", "Caja Diaria"],
        [0, 0, 0, 0, 0]
    ]
    
    # Calcular totales
    ingreso_bruto = sum(float(t.monto_total) for t in trabajos)
    total_comisiones = sum(c["comision"] for c in comisiones_detalle)
    gasto_total = sum(float(g.monto) for g in gastos)
    prestamo_total = sum(float(p.monto) for p in prestamos)
    caja_diaria = ingreso_bruto - total_comisiones - gasto_total
    
    resumen_data[1] = [
        round(ingreso_bruto, 2),
        round(total_comisiones, 2),
        round(prestamo_total, 2),
        round(gasto_total, 2),
        round(caja_diaria, 2)
    ]
    
    for col, (header, value) in enumerate(zip(resumen_data[0], resumen_data[1]), 1):
        ws.cell(row=3, column=col, value=header).font = header_font
        ws.cell(row=3, column=col).fill = header_fill
        ws.cell(row=4, column=col, value=value).fill = total_fill
    
    # === TRABAJOS ===
    ws.merge_cells(f'A6:E6')
    ws['A6'] = "TRABAJOS FINALIZADOS"
    ws['A6'].font = header_font
    ws['A6'].fill = header_fill
    ws['A6'].alignment = Alignment(horizontal="center")
    
    ws.cell(row=7, column=1, value="Nombre").font = header_font
    ws.cell(row=7, column=2, value="Monto").font = header_font
    ws.cell(row=7, column=3, value="m²").font = header_font
    ws.cell(row=7, column=4, value="Unidades").font = header_font
    ws.cell(row=7, column=5, value="Estado").font = header_font
    
    for row, trabajo in enumerate(trabajos, 8):
        ws.cell(row=row, column=1, value=trabajo.nombre)
        ws.cell(row=row, column=2, value=float(trabajo.monto_total))
        ws.cell(row=row, column=3, value=float(trabajo.metros_cuadrados) if trabajo.metros_cuadrados else "")
        ws.cell(row=row, column=4, value=trabajo.unidades if trabajo.unidades else "")
        ws.cell(row=row, column=5, value=trabajo.estado)
    
    # === COMISIONES ===
    ultimo_trabajo = 8 + len(trabajos)
    ws.merge_cells(f'A{ultimo_trabajo}:B{ultimo_trabajo}')
    ws[f'A{ultimo_trabajo}'] = "COMISIONES"
    ws[f'A{ultimo_trabajo}'].font = header_font
    ws[f'A{ultimo_trabajo}'].fill = header_fill
    ws[f'A{ultimo_trabajo}'].alignment = Alignment(horizontal="center")
    
    ws.cell(row=ultimo_trabajo+1, column=1, value="Empleado - Trabajo").font = header_font
    ws.cell(row=ultimo_trabajo+1, column=2, value="Monto").font = header_font
    
    for i, comision in enumerate(comisiones_detalle):
        ws.cell(row=ultimo_trabajo+2+i, column=1, value=comision["nombre"])
        ws.cell(row=ultimo_trabajo+2+i, column=2, value=comision["comision"])
    
    # === GASTOS ===
    ultimo_comision = ultimo_trabajo+2+len(comisiones_detalle)
    ws.merge_cells(f'A{ultimo_comision}:C{ultimo_comision}')
    ws[f'A{ultimo_comision}'] = "GASTOS"
    ws[f'A{ultimo_comision}'].font = header_font
    ws[f'A{ultimo_comision}'].fill = header_fill
    ws[f'A{ultimo_comision}'].alignment = Alignment(horizontal="center")
    
    ws.cell(row=ultimo_comision+1, column=1, value="Descripción").font = header_font
    ws.cell(row=ultimo_comision+1, column=2, value="Categoría").font = header_font
    ws.cell(row=ultimo_comision+1, column=3, value="Monto").font = header_font
    
    for i, gasto in enumerate(gastos):
        ws.cell(row=ultimo_comision+2+i, column=1, value=gasto.descripcion or "Sin descripción")
        ws.cell(row=ultimo_comision+2+i, column=2, value=gasto.categoria or "General")
        ws.cell(row=ultimo_comision+2+i, column=3, value=float(gasto.monto))
    
    # === PRÉSTAMOS ===
    ultimo_gasto = ultimo_comision+2+len(gastos)
    ws.merge_cells(f'A{ultimo_gasto}:C{ultimo_gasto}')
    ws[f'A{ultimo_gasto}'] = "PRÉSTAMOS"
    ws[f'A{ultimo_gasto}'].font = header_font
    ws[f'A{ultimo_gasto}'].fill = header_fill
    ws[f'A{ultimo_gasto}'].alignment = Alignment(horizontal="center")
    
    ws.cell(row=ultimo_gasto+1, column=1, value="Empleado").font = header_font
    ws.cell(row=ultimo_gasto+1, column=2, value="Motivo").font = header_font
    ws.cell(row=ultimo_gasto+1, column=3, value="Monto").font = header_font
    
    for i, prestamo in enumerate(prestamos):
        ws.cell(row=ultimo_gasto+2+i, column=1, value=prestamo.nombre_completo)
        ws.cell(row=ultimo_gasto+2+i, column=2, value=prestamo.descripcion or "-")
        ws.cell(row=ultimo_gasto+2+i, column=3, value=float(prestamo.monto))
    
   
   # Ajustar ancho de columnas de forma segura
    columnas = ['A', 'B', 'C', 'D', 'E']
    for col in columnas:
        ws.column_dimensions[col].width = 25
    
    # Guardar en memoria
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_diario_{fecha}.xlsx"}
    )
@app.get("/reportes/semanal", response_class=HTMLResponse)
async def reporte_semanal(
    request: Request,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    from datetime import datetime, timedelta, date
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    from collections import defaultdict
    import json
    
    # Fechas por defecto: semana actual
    if not fecha_inicio or not fecha_fin:
        hoy = datetime.today().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        fecha_inicio = inicio_semana.isoformat()
        fecha_fin = fin_semana.isoformat()
    
    fecha_ini = date.fromisoformat(fecha_inicio)
    fecha_fin_obj = date.fromisoformat(fecha_fin)
    
    print(f"📅 Consultando del {fecha_ini} al {fecha_fin_obj}")
    
    try:
        # ============================================
        # 1. INGRESOS DESDE CIERRES DE CAJA
        # ============================================
        cierres = (await db.execute(
            select(models.CierreCaja).where(
                models.CierreCaja.fecha >= fecha_ini,
                models.CierreCaja.fecha <= fecha_fin_obj
            )
        )).scalars().all()
        
        ingreso_total_usd = 0.0
        tasa_promedio = 0.0
        diferencia_promedio = 0.0
        suma_tasas = 0.0
        suma_diferencias = 0.0
        cantidad_cierres = 0
        ingresos_detalle = []
        
        for cierre in cierres:
            total_ingreso = float(cierre.total_ingresos_usd) if cierre.total_ingresos_usd is not None else 0
            ingreso_total_usd += total_ingreso
            
            tasa = float(cierre.tasa_cambio) if cierre.tasa_cambio is not None else 0
            diferencia = float(cierre.diferencia_usd) if cierre.diferencia_usd is not None else 0
            
            if tasa > 0:
                suma_tasas += tasa
                suma_diferencias += diferencia
                cantidad_cierres += 1
            
            ingresos_detalle.append({
                "fecha": cierre.fecha.isoformat(),
                "descripcion": f"Cierre #{cierre.id}",
                "total_ingreso": round(total_ingreso, 2),
                "tasa": round(tasa, 2),
                "diferencia": round(diferencia, 2)
            })
        
        tasa_promedio = suma_tasas / cantidad_cierres if cantidad_cierres > 0 else 0
        diferencia_promedio = suma_diferencias / cantidad_cierres if cantidad_cierres > 0 else 0
        
        # ============================================
        # 2. GASTOS OPERATIVOS
        # ============================================
        gastos = (await db.execute(
            select(models.GastoDiario).where(
                models.GastoDiario.fecha >= fecha_ini,
                models.GastoDiario.fecha <= fecha_fin_obj
            )
        )).scalars().all()
        
        gasto_total_usd = 0.0
        gastos_detalle = []
        for g in gastos:
            monto = float(g.monto) if g.monto is not None else 0
            monto_usd = monto / 36.50
            gasto_total_usd += monto_usd
            gastos_detalle.append({
                "fecha": g.fecha.isoformat(),
                "descripcion": g.descripcion or "Gasto",
                "monto_usd": round(monto_usd, 2)
            })
        
        # ============================================
        # 3. PAGOS A EMPLEADOS
        # ============================================
        pagos_empleados = (await db.execute(
            select(models.PagoSemanal)
            .where(
                models.PagoSemanal.fecha_pago >= fecha_ini,
                models.PagoSemanal.fecha_pago <= fecha_fin_obj
            )
            .options(selectinload(models.PagoSemanal.empleado))
        )).scalars().all()
        
        pago_empleados_usd = 0.0
        empleados_detalle = []
        
        for pago in pagos_empleados:
            total_neto = float(pago.total_neto) if pago.total_neto is not None else 0
            pago_empleados_usd += total_neto
            
            nombre_empleado = pago.empleado.nombre_completo if pago.empleado else f"Empleado ID {pago.empleado_id}"
            
            empleados_detalle.append({
                "nombre": nombre_empleado,
                "sueldo_fijo": float(pago.sueldo_fijo) if pago.sueldo_fijo is not None else 0,
                "comisiones": float(pago.total_comisiones) if pago.total_comisiones is not None else 0,
                "prestamos": float(pago.total_prestamos) if pago.total_prestamos is not None else 0,
                "deducciones": float(pago.total_deducciones) if pago.total_deducciones is not None else 0,
                "pago_neto": total_neto
            })
        
        # ============================================
        # 4. COMPRAS DE MATERIALES (ENTRADAS AL INVENTARIO)
        # ============================================
        compras_detalle = []
        total_compras_usd = 0.0
        
        try:
            entradas = (await db.execute(
                select(models.EntradaInventario)
                .where(
                    models.EntradaInventario.fecha_entrada >= fecha_ini,
                    models.EntradaInventario.fecha_entrada <= fecha_fin_obj
                )
                .options(selectinload(models.EntradaInventario.material))
            )).scalars().all()
            
            print(f"📦 Compras de material encontradas: {len(entradas)}")
            
            for entrada in entradas:
                if entrada.material:
                    precio = float(entrada.precio_unitario) if entrada.precio_unitario is not None else 0
                    cantidad = float(entrada.cantidad) if entrada.cantidad is not None else 0
                    monto_usd = precio * cantidad
                    total_compras_usd += monto_usd
                    compras_detalle.append({
                        "nombre": entrada.material.nombre,
                        "cantidad": cantidad,
                        "unidad": entrada.material.unidad_medida,
                        "precio_unitario": precio,
                        "costo_total": round(monto_usd, 2),
                        "fecha": entrada.fecha_entrada.isoformat(),
                        "proveedor": entrada.proveedor_nombre or "N/A"
                    })
        except Exception as e:
            print(f"Error en compras de materiales: {e}")
        
        # ============================================
        # 5. MATERIALES UTILIZADOS (SALIDAS)
        # ============================================
        materiales_detalle = []
        total_materiales_usd = 0.0
        
        try:
            salidas = (await db.execute(
                select(models.SalidaMaterial)
                .where(
                    models.SalidaMaterial.fecha_salida >= fecha_ini,
                    models.SalidaMaterial.fecha_salida <= fecha_fin_obj
                )
            )).scalars().all()
            
            print(f"📤 Salidas de material encontradas: {len(salidas)}")
            
            for s in salidas:
                material = await db.execute(
                    select(models.MaterialInventario).where(models.MaterialInventario.id == s.material_id)
                )
                material_obj = material.scalar_one_or_none()
                
                if material_obj:
                    precio = float(material_obj.precio_compra) if material_obj.precio_compra is not None else 0
                    cantidad = float(s.cantidad) if s.cantidad is not None else 0
                    monto_usd = precio * cantidad
                    total_materiales_usd += monto_usd
                    materiales_detalle.append({
                        "nombre": material_obj.nombre,
                        "cantidad": cantidad,
                        "unidad": material_obj.unidad_medida,
                        "costo_total": round(monto_usd, 2),
                        "motivo": s.motivo or "Uso en trabajo",
                        "fecha": s.fecha_salida.isoformat()
                    })
        except Exception as e:
            print(f"Error en salidas de materiales: {e}")
        
        # ============================================
        # 6. ACTIVOS FIJOS (Depreciación)
        # ============================================
        activos_detalle = []
        total_depreciacion_usd = 0.0
        
        try:
            activos = (await db.execute(
                select(models.ActivoFijo).where(models.ActivoFijo.estado == "activo")
            )).scalars().all()
            
            for a in activos:
                vida_util_anios = 5
                if a.categoria_id:
                    cat_result = await db.execute(
                        select(models.CategoriaActivoFijo.vida_util_anios)
                        .where(models.CategoriaActivoFijo.id == a.categoria_id)
                    )
                    vida = cat_result.scalar()
                    if vida:
                        vida_util_anios = float(vida)
                
                costo_inicial = float(a.costo_inicial) if a.costo_inicial is not None else 0
                valor_residual = float(a.valor_residual) if a.valor_residual is not None else 0
                
                if vida_util_anios > 0 and costo_inicial > 0:
                    depreciacion_anual = (costo_inicial - valor_residual) / vida_util_anios
                    depreciacion_semanal = depreciacion_anual / 52
                    total_depreciacion_usd += depreciacion_semanal
                    
                    activos_detalle.append({
                        "nombre": a.nombre,
                        "valor_original": round(costo_inicial, 2),
                        "depreciacion_semanal": round(depreciacion_semanal, 2),
                        "valor_actual": round(costo_inicial - depreciacion_semanal, 2)
                    })
        except Exception as e:
            print(f"Error en activos fijos: {e}")
        
        # ============================================
        # 7. CÁLCULOS FINALES
        # ============================================
        total_egresos_usd = gasto_total_usd + pago_empleados_usd + total_materiales_usd + total_depreciacion_usd
        saldo_final_usd = ingreso_total_usd - total_egresos_usd
        margen_utilidad = (saldo_final_usd / ingreso_total_usd * 100) if ingreso_total_usd > 0 else 0
        
        print(f"📊 RESUMEN FINAL:")
        print(f"   Ingresos: ${ingreso_total_usd:.2f}")
        print(f"   Gastos: ${gasto_total_usd:.2f}")
        print(f"   Empleados: ${pago_empleados_usd:.2f}")
        print(f"   Compras Materiales: ${total_compras_usd:.2f}")
        print(f"   Materiales Usados: ${total_materiales_usd:.2f}")
        print(f"   Depreciación: ${total_depreciacion_usd:.2f}")
        print(f"   Saldo final: ${saldo_final_usd:.2f}")
        
        # ============================================
        # 8. DATOS PARA EL TEMPLATE
        # ============================================
        datos = {
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "resumen": {
                "ingreso_total_usd": round(ingreso_total_usd, 2),
                "gasto_total_usd": round(gasto_total_usd, 2),
                "pago_empleados_usd": round(pago_empleados_usd, 2),
                "total_compras_usd": round(total_compras_usd, 2),
                "total_materiales_usd": round(total_materiales_usd, 2),
                "total_depreciacion_usd": round(total_depreciacion_usd, 2),
                "saldo_final_usd": round(saldo_final_usd, 2),
                "margen_utilidad": round(margen_utilidad, 2),
                "tasa_promedio": round(tasa_promedio, 2),
                "diferencia_promedio": round(diferencia_promedio, 2)
            },
            "ingresos_detalle": ingresos_detalle,
            "gastos_detalle": gastos_detalle,
            "empleados_detalle": empleados_detalle,
            "compras_detalle": compras_detalle,
            "materiales_detalle": materiales_detalle,
            "activos_detalle": activos_detalle
        }
        
        return templates.TemplateResponse("reportes/semanal.html", {
            "request": request,
            "datos": datos,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        })
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        
        datos_vacios = {
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "resumen": {
                "ingreso_total_usd": 0, "gasto_total_usd": 0, "pago_empleados_usd": 0,
                "total_compras_usd": 0, "total_materiales_usd": 0, "total_depreciacion_usd": 0,
                "saldo_final_usd": 0, "margen_utilidad": 0, "tasa_promedio": 0, "diferencia_promedio": 0
            },
            "ingresos_detalle": [], "gastos_detalle": [], 
            "empleados_detalle": [], "compras_detalle": [],
            "materiales_detalle": [], "activos_detalle": []
        }
        
        return templates.TemplateResponse("reportes/semanal.html", {
            "request": request,
            "datos": datos_vacios,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        })
    
async def generar_datos_reporte_completo(fecha_inicio, fecha_fin, db):
    """Función auxiliar para generar datos COMPLETOS del reporte semanal"""
    from datetime import datetime, timedelta, date
    from sqlalchemy import select
    from collections import defaultdict
    
    if not fecha_inicio or not fecha_fin:
        hoy = datetime.today().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        fecha_inicio = inicio_semana.isoformat()
        fecha_fin = fin_semana.isoformat()
    
    fecha_ini = date.fromisoformat(fecha_inicio)
    fecha_fin_obj = date.fromisoformat(fecha_fin)
    
    # Trabajos
    trabajos = (await db.execute(
        select(models.Trabajo).where(
            models.Trabajo.estado == "finalizado",
            models.Trabajo.fecha_inicio >= fecha_ini,
            models.Trabajo.fecha_inicio <= fecha_fin_obj
        ).order_by(models.Trabajo.fecha_inicio)
    )).scalars().all()
    
    ingreso_bruto = sum(float(t.monto_total or 0) for t in trabajos)
    
    # Trabajos con empleados
    trabajos_detalle = []
    for trabajo in trabajos:
        asignaciones = (await db.execute(
            select(models.Asignacion)
            .where(models.Asignacion.trabajo_id == trabajo.id)
            .options(
                selectinload(models.Asignacion.empleado),
                selectinload(models.Asignacion.rol)
            )
        )).scalars().all()
        
        empleados_trabajo = []
        for asign in asignaciones:
            if asign.empleado:
                nombre_rol = "Sin rol"
                if asign.rol and asign.rol.nombre:
                    nombre_rol = asign.rol.nombre
                elif asign.rol_id:
                    rol_result = await db.execute(
                        select(models.Rol.nombre).where(models.Rol.id == asign.rol_id)
                    )
                    nombre_rol = rol_result.scalar() or "Sin rol"
                
                empleados_trabajo.append({
                    "nombre": asign.empleado.nombre_completo,
                    "rol": nombre_rol
                })
        
        trabajos_detalle.append({
            "id": trabajo.id,
            "nombre_trabajo": trabajo.nombre or "Sin nombre",
            "monto_total": float(trabajo.monto_total or 0),
            "fecha_inicio": trabajo.fecha_inicio.isoformat(),
            "metros_cuadrados": float(trabajo.metros_cuadrados or 0),
            "unidades": int(trabajo.unidades or 0),
            "empleados": empleados_trabajo
        })
    
    # Calcular comisiones
    comisiones_por_empleado = {}
    total_comisiones_reporte = 0.0
    
    for trabajo in trabajos:
        asignaciones = (await db.execute(
            select(models.Asignacion)
            .where(models.Asignacion.trabajo_id == trabajo.id)
            .options(selectinload(models.Asignacion.empleado))
        )).scalars().all()
        
        asignaciones_por_rol = defaultdict(list)
        for asign in asignaciones:
            asignaciones_por_rol[asign.rol_id].append(asign)
        
        for rol_id, asignaciones_rol in asignaciones_por_rol.items():
            if not asignaciones_rol:
                continue
            
            primera_asign = asignaciones_rol[0]
            tipo_comision = primera_asign.tipo_comision
            valor_comision = float(primera_asign.valor_comision or 0)
            
            comision_total_rol = 0.0
            if tipo_comision == "porcentaje":
                comision_total_rol = float(trabajo.monto_total or 0) * (valor_comision / 100)
            elif tipo_comision == "fijo":
                comision_total_rol = valor_comision
            elif tipo_comision == "por_m2":
                metros = float(trabajo.metros_cuadrados or 0)
                comision_total_rol = metros * valor_comision
            elif tipo_comision == "por_unidad":
                unidades = int(trabajo.unidades or 0)
                comision_total_rol = unidades * valor_comision
            
            comision_por_empleado = comision_total_rol / len(asignaciones_rol)
            
            for asign in asignaciones_rol:
                if asign.empleado:
                    emp_nombre = asign.empleado.nombre_completo
                    if emp_nombre not in comisiones_por_empleado:
                        comisiones_por_empleado[emp_nombre] = 0.0
                    comisiones_por_empleado[emp_nombre] += comision_por_empleado
                    total_comisiones_reporte += comision_por_empleado
    
    # Empleados y cálculo de pagos netos
    empleados = (await db.execute(
        select(models.Empleado).where(models.Empleado.activo == True)
    )).scalars().all()
    
    prestamos_semana = (await db.execute(
        select(models.Prestamo).where(
            models.Prestamo.fecha >= fecha_ini,
            models.Prestamo.fecha <= fecha_fin_obj
        )
    )).scalars().all()
    
    prestamos_por_empleado = {}
    prestamo_total_general = 0.0
    for p in prestamos_semana:
        emp_nombre = (await db.execute(
            select(models.Empleado.nombre_completo).where(models.Empleado.id == p.empleado_id)
        )).scalar() or f"Empleado ID {p.empleado_id}"
        monto = float(p.monto or 0)
        prestamos_por_empleado[emp_nombre] = prestamos_por_empleado.get(emp_nombre, 0.0) + monto
        prestamo_total_general += monto
    
    empleados_detalle = []
    total_sueldos_fijos = 0.0
    
    for emp in empleados:
        sueldo = float(emp.sueldo_fijo or 0.0)
        comision_emp = comisiones_por_empleado.get(emp.nombre_completo, 0.0)
        prestamo_emp = prestamos_por_empleado.get(emp.nombre_completo, 0.0)
        pago_neto = sueldo + comision_emp - prestamo_emp
        
        if sueldo > 0 or comision_emp > 0 or prestamo_emp > 0:
            empleados_detalle.append({
                "nombre": emp.nombre_completo,
                "sueldo_fijo": round(sueldo, 2),
                "comisiones": round(comision_emp, 2),
                "prestamos": round(prestamo_emp, 2),
                "pago_neto": round(pago_neto, 2)
            })
            total_sueldos_fijos += sueldo
    
    gastos = (await db.execute(
        select(models.GastoDiario).where(
            models.GastoDiario.fecha >= fecha_ini,
            models.GastoDiario.fecha <= fecha_fin_obj
        )
    )).scalars().all()
    
    gasto_total = sum(float(g.monto or 0) for g in gastos)
    
    pago_neto_empleados = total_sueldos_fijos + total_comisiones_reporte - prestamo_total_general
    caja_semanal = ingreso_bruto - pago_neto_empleados - gasto_total
    
    return {
        "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
        "resumen": {
            "ingreso_bruto": round(ingreso_bruto, 2),
            "total_sueldos_fijos": round(total_sueldos_fijos, 2),
            "total_comisiones": round(total_comisiones_reporte, 2),
            "total_prestamos": round(prestamo_total_general, 2),
            "pago_neto_empleados": round(pago_neto_empleados, 2),
            "gasto_total": round(gasto_total, 2),
            "caja_semanal": round(caja_semanal, 2)
        },
        "empleados": sorted(empleados_detalle, key=lambda x: x["pago_neto"], reverse=True),
        "trabajos": trabajos_detalle
    }


@app.get("/reportes/semanal/exportar/excel")
async def exportar_excel_semanal(
    request: Request,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    # Obtener datos COMPLETOS del reporte
    datos_reporte = await generar_datos_reporte_completo(fecha_inicio, fecha_fin, db)
    
    # Crear workbook de Excel
    wb = openpyxl.Workbook()
    
    # === HOJA 1: RESUMEN ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título
    ws_resumen['A1'] = f"REPORTE SEMANAL"
    ws_resumen['A2'] = f"Periodo: {datos_reporte['periodo']['inicio']} - {datos_reporte['periodo']['fin']}"
    ws_resumen.merge_cells('A1:F1')
    ws_resumen.merge_cells('A2:F2')
    
    # Estilos para título
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    ws_resumen['A1'].font = Font(bold=True, size=16)
    ws_resumen['A2'].font = Font(bold=True, size=12)
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    
    # Resumen
    ws_resumen['A4'] = "RESUMEN"
    ws_resumen['A4'].font = Font(bold=True, size=14)
    
    resumen_data = [
        ["Ingreso Bruto", f"${datos_reporte['resumen']['ingreso_bruto']:,.2f}"],
        ["Total Sueldos Fijos", f"${datos_reporte['resumen']['total_sueldos_fijos']:,.2f}"],
        ["Total Comisiones", f"${datos_reporte['resumen']['total_comisiones']:,.2f}"],
        ["Total Préstamos", f"${datos_reporte['resumen']['total_prestamos']:,.2f}"],
        ["Pago Neto Empleados", f"${datos_reporte['resumen']['pago_neto_empleados']:,.2f}"],
        ["Gasto Total", f"${datos_reporte['resumen']['gasto_total']:,.2f}"],
        ["CAJA SEMANAL", f"${datos_reporte['resumen']['caja_semanal']:,.2f}"]
    ]
    
    for i, (label, value) in enumerate(resumen_data, start=5):
        ws_resumen[f'A{i}'] = label
        ws_resumen[f'B{i}'] = value
        # Estilo para CAJA SEMANAL
        if "CAJA SEMANAL" in label:
            ws_resumen[f'A{i}'].font = Font(bold=True, color="FF0000")
            ws_resumen[f'B{i}'].font = Font(bold=True, color="FF0000")
    
    # === HOJA 2: DETALLE EMPLEADOS ===
    ws_empleados = wb.create_sheet("Detalle Empleados")
    ws_empleados['A1'] = "DETALLE DE PAGO NETO POR EMPLEADO"
    ws_empleados.merge_cells('A1:F1')
    ws_empleados['A1'].font = Font(bold=True, size=14)
    ws_empleados['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    empleados_headers = ["Empleado", "Sueldo Fijo", "Comisiones", "Préstamos", "Pago Neto"]
    for col, header in enumerate(empleados_headers, start=1):
        ws_empleados.cell(row=3, column=col, value=header)
        ws_empleados.cell(row=3, column=col).font = Font(bold=True)
        ws_empleados.cell(row=3, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos de empleados
    row = 4
    for emp in datos_reporte['empleados']:
        ws_empleados.cell(row=row, column=1, value=emp['nombre'])
        ws_empleados.cell(row=row, column=2, value=emp['sueldo_fijo'])
        ws_empleados.cell(row=row, column=3, value=emp['comisiones'])
        ws_empleados.cell(row=row, column=4, value=emp['prestamos'])
        ws_empleados.cell(row=row, column=5, value=emp['pago_neto'])
        row += 1
    
    # Formato de moneda para columnas numéricas
    for row in range(4, row):
        for col in range(2, 6):
            ws_empleados.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # === HOJA 3: TRABAJOS FINALIZADOS ===
    ws_trabajos = wb.create_sheet("Trabajos")
    ws_trabajos['A1'] = "TRABAJOS FINALIZADOS"
    ws_trabajos.merge_cells('A1:G1')
    ws_trabajos['A1'].font = Font(bold=True, size=14)
    ws_trabajos['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    trabajos_headers = ["Trabajo", "Monto", "Fecha", "m²", "Unidades", "Empleados"]
    for col, header in enumerate(trabajos_headers, start=1):
        ws_trabajos.cell(row=3, column=col, value=header)
        ws_trabajos.cell(row=3, column=col).font = Font(bold=True)
        ws_trabajos.cell(row=3, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos de trabajos
    row = 4
    for trabajo in datos_reporte['trabajos']:
        empleados_str = "; ".join([f"{emp['nombre']} ({emp['rol']})" for emp in trabajo['empleados']])
        ws_trabajos.cell(row=row, column=1, value=trabajo['nombre_trabajo'])
        ws_trabajos.cell(row=row, column=2, value=trabajo['monto_total'])
        ws_trabajos.cell(row=row, column=3, value=trabajo['fecha_inicio'])
        ws_trabajos.cell(row=row, column=4, value=trabajo['metros_cuadrados'])
        ws_trabajos.cell(row=row, column=5, value=trabajo['unidades'])
        ws_trabajos.cell(row=row, column=6, value=empleados_str)
        row += 1
    
    # Formato de moneda para montos
    for row in range(4, row):
        ws_trabajos.cell(row=row, column=2).number_format = '$#,##0.00'
        ws_trabajos.cell(row=row, column=4).number_format = '#,##0.00'
    
    # Ajustar ancho de columnas
    for ws in [ws_resumen, ws_empleados, ws_trabajos]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_semanal_{fecha_inicio}_{fecha_fin}.xlsx"}
    )

@app.get("/reportes/semanal/exportar/pdf")
async def exportar_pdf_semanal(
    request: Request,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    # Obtener datos COMPLETOS del reporte (incluyendo empleados)
    datos_reporte = await generar_datos_reporte_completo(fecha_inicio, fecha_fin, db)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("REPORTE SEMANAL", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(f"Periodo: {datos_reporte['periodo']['inicio']} - {datos_reporte['periodo']['fin']}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Resumen
    elements.append(Paragraph("RESUMEN", styles['Heading2']))
    resumen_data = [
        ["Concepto", "Monto"],
        ["Ingreso Bruto", f"${datos_reporte['resumen']['ingreso_bruto']:,.2f}"],
        ["Total Sueldos Fijos", f"${datos_reporte['resumen']['total_sueldos_fijos']:,.2f}"],
        ["Total Comisiones", f"${datos_reporte['resumen']['total_comisiones']:,.2f}"],
        ["Total Préstamos", f"${datos_reporte['resumen']['total_prestamos']:,.2f}"],
        ["Pago Neto Empleados", f"${datos_reporte['resumen']['pago_neto_empleados']:,.2f}"],
        ["Gasto Total", f"${datos_reporte['resumen']['gasto_total']:,.2f}"],
        ["CAJA SEMANAL", f"${datos_reporte['resumen']['caja_semanal']:,.2f}"]
    ]
    
    resumen_table = Table(resumen_data)
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 24))
    
    # DETALLE DE PAGO NETO POR EMPLEADO
    elements.append(Paragraph("DETALLE DE PAGO NETO POR EMPLEADO", styles['Heading2']))
    empleados_data = [["Empleado", "Sueldo Fijo", "Comisiones", "Préstamos", "Pago Neto"]]
    
    for emp in datos_reporte['empleados']:
        empleados_data.append([
            emp['nombre'],
            f"${emp['sueldo_fijo']:,.2f}",
            f"${emp['comisiones']:,.2f}",
            f"${emp['prestamos']:,.2f}",
            f"${emp['pago_neto']:,.2f}"
        ])
    
    if len(empleados_data) > 1:
        empleados_table = Table(empleados_data)
        empleados_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7)
        ]))
        elements.append(empleados_table)
        elements.append(Spacer(1, 24))
    
    # Trabajos
    elements.append(Paragraph("TRABAJOS FINALIZADOS", styles['Heading2']))
    trabajos_data = [["Trabajo", "Monto", "Fecha", "m²", "Unidades", "Empleados"]]
    
    for trabajo in datos_reporte['trabajos']:
        empleados_str = "; ".join([f"{emp['nombre']} ({emp['rol']})" for emp in trabajo['empleados']])
        trabajos_data.append([
            trabajo['nombre_trabajo'],
            f"${trabajo['monto_total']:,.2f}",
            trabajo['fecha_inicio'],
            f"{trabajo['metros_cuadrados']:.2f}",
            str(trabajo['unidades']),
            empleados_str
        ])
    
    if len(trabajos_data) > 1:
        trabajos_table = Table(trabajos_data)
        trabajos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7)
        ]))
        elements.append(trabajos_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_semanal_{fecha_inicio}_{fecha_fin}.pdf"}
    )
@app.post("/roles", response_class=HTMLResponse)
async def crear_rol_completo(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    form = await request.form()
    nombre = form.get("nombre", "").strip().lower()
    
    if not nombre:
        roles = (await db.execute(select(models.Rol).order_by(models.Rol.nombre))).scalars().all()
        return templates.TemplateResponse("roles.html", {
            "request": request,
            "roles": roles,
            "error": "El nombre del rol es requerido"
        })
    
    try:
        db_rol = models.Rol(nombre=nombre)
        db.add(db_rol)
        await db.commit()
        
        # ✅ ¡ÉXITO! Redirigir con parámetro success (en lugar de cookie)
        return RedirectResponse(url="/roles?success=1", status_code=303)
        
    except Exception:
        roles = (await db.execute(select(models.Rol).order_by(models.Rol.nombre))).scalars().all()
        return templates.TemplateResponse("roles.html", {
            "request": request,
            "roles": roles,
            "error": "Error: Ya existe un rol con ese nombre"
        })
        
@app.get("/roles", response_class=HTMLResponse)
async def listar_roles(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    roles = (await db.execute(
        select(models.Rol).order_by(models.Rol.nombre)
    )).scalars().all()
    
    return templates.TemplateResponse("roles.html", {
        "request": request,
        "roles": roles
    })


# === EDICIÓN Y ELIMINACIÓN DE ROLES ===
@app.get("/roles/{rol_id}/editar", response_class=HTMLResponse)
async def formulario_editar_rol(
    rol_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    rol = await db.get(models.Rol, rol_id)
    if not rol:
        return RedirectResponse(url="/roles")
    return templates.TemplateResponse("rol_editar.html", {
        "request": request,
        "rol": rol,
        "error": None
    })

@app.post("/roles/{rol_id}/editar", response_class=HTMLResponse)
async def editar_rol(
    rol_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    form = await request.form()
    nombre = form.get("nombre", "").strip().lower()
    
    if not nombre:
        rol = await db.get(models.Rol, rol_id)
        return templates.TemplateResponse("rol_editar.html", {
            "request": request,
            "rol": rol,
            "error": "El nombre es requerido"
        })
    
    try:
        rol = await db.get(models.Rol, rol_id)
        if rol:
            rol.nombre = nombre
            await db.commit()
            # ✅ ¡CAMBIO AQUÍ! Redirigir a la página de edición con ?success=1
            return RedirectResponse(
                url=f"/roles/{rol_id}/editar?success=1", 
                status_code=303
            )
        else:
            return RedirectResponse(url="/roles", status_code=303)
            
    except Exception:
        rol = await db.get(models.Rol, rol_id)
        return templates.TemplateResponse("rol_editar.html", {
            "request": request,
            "rol": rol,
            "error": "Error al actualizar"
        })

@app.get("/roles/{rol_id}/eliminar")
async def eliminar_rol(
    rol_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    try:
        rol = await db.get(models.Rol, rol_id)
        if rol:
            await db.delete(rol)
            await db.commit()

        return RedirectResponse(url="/roles?deleted=1", status_code=303)
    
    
    except Exception:
       
       await db.rollback()
       return RedirectResponse(url="/roles", status_code=303)

# === EMPLEADOS: LISTA + CREAR ===
# === EMPLEADOS: PÁGINA PRINCIPAL ===
@app.get("/empleados", response_class=HTMLResponse)
async def gestion_empleados(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    roles = (await db.execute(select(models.Rol).order_by(models.Rol.nombre))).scalars().all()
    empleados = (await db.execute(
        select(models.Empleado).where(models.Empleado.activo == True)
    )).scalars().all()
    
    # ✅ PASA LOS OBJETOS DIRECTAMENTE
    return templates.TemplateResponse("empleados/empleados.html", {
        "request": request,
        "empleados": empleados,  # ← Objetos, no diccionarios
        "roles": roles           # ← Objetos, no diccionarios
    })

@app.get("/empleados/nuevo", response_class=HTMLResponse)
async def nuevo_empleado_form(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    try:
        from sqlalchemy import select
        result = await db.execute(select(models.Rol))
        roles = result.scalars().all()
        
        return templates.TemplateResponse("empleados/empleados.html", {
            "request": request,
            "roles": roles,
            "modo": "crear"
        })
    except Exception as e:
        print(f"Error al cargar roles: {str(e)}")
        # En caso de error, redirigir con mensaje
        return RedirectResponse(url="/empleados?error=cargar_roles", status_code=303)
@app.post("/empleados")
async def crear_empleado(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    try:
        form = await request.form()
        nombre_completo = form.get("nombre_completo", "").strip()
        tipo_contrato = form.get("tipo_contrato", "")
        sueldo_fijo_str = form.get("sueldo_fijo", "0.0")
        rol_id_str = form.get("rol_id", "0")
        
        # Validaciones
        if not nombre_completo or not tipo_contrato or not rol_id_str.isdigit():
            return RedirectResponse(url="/empleados/nuevo?error=1", status_code=303)
        
        from decimal import Decimal, InvalidOperation
        try:
            sueldo_fijo = Decimal(sueldo_fijo_str) if sueldo_fijo_str else Decimal("0.0")
        except InvalidOperation:
            sueldo_fijo = Decimal("0.0")
        
        rol_id = int(rol_id_str)
        
        # Verificar rol
        from sqlalchemy import select
        rol = (await db.execute(select(models.Rol).where(models.Rol.id == rol_id))).scalar()
        if not rol:
            return RedirectResponse(url="/empleados/nuevo?error=1", status_code=303)
        
        # Crear empleado
        nuevo_empleado = models.Empleado(
            nombre_completo=nombre_completo,
            tipo_contrato=tipo_contrato,
            sueldo_fijo=sueldo_fijo,
            rol_id=rol_id,
            activo=True
        )
        
        db.add(nuevo_empleado)
        await db.commit()
        await db.refresh(nuevo_empleado)
        
        # ✅ ¡CAMBIO AQUÍ! Agregar ?success=1 para mostrar mensaje
        return RedirectResponse(url="/empleados?success=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"ERROR al crear empleado: {str(e)}")
        return RedirectResponse(url="/empleados/nuevo?error=1", status_code=303)

# === EMPLEADOS: EDITAR ===
@app.get("/empleados/{empleado_id}/editar", response_class=HTMLResponse)
async def formulario_editar_empleado(
    request: Request,
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    empleado = (await db.execute(
        select(models.Empleado).where(models.Empleado.id == empleado_id)
    )).scalar_one_or_none()
    
    if not empleado:
        return RedirectResponse(url="/empleados", status_code=303)
    
    roles = (await db.execute(select(models.Rol))).scalars().all()
    
    return templates.TemplateResponse("empleados/editar.html", {
        "request": request,
        "empleado": empleado,
        "roles": roles
    })

@app.post("/empleados/{empleado_id}/editar")
async def editar_empleado(
    request: Request,
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    form = await request.form()
    
    empleado = (await db.execute(
        select(models.Empleado).where(models.Empleado.id == empleado_id)
    )).scalar_one_or_none()
    
    if not empleado:
        return RedirectResponse(url="/empleados", status_code=303)
    
    try:
        from decimal import Decimal, InvalidOperation
        
        nombre_completo = form.get("nombre_completo", "").strip()
        tipo_contrato = form.get("tipo_contrato", "")
        sueldo_fijo_str = form.get("sueldo_fijo", "0.0")
        rol_id = int(form.get("rol_id", "0"))
        
        if not nombre_completo or not tipo_contrato or rol_id <= 0:
            raise ValueError("Datos incompletos")
        
        try:
            sueldo_fijo = Decimal(sueldo_fijo_str) if sueldo_fijo_str else Decimal("0.0")
        except InvalidOperation:
            sueldo_fijo = Decimal("0.0")
        
        # Verificar que el rol exista
        rol = (await db.execute(
            select(models.Rol).where(models.Rol.id == rol_id)
        )).scalar_one_or_none()
        if not rol:
            raise ValueError("Rol inválido")
        
        # Actualizar empleado
        empleado.nombre_completo = nombre_completo
        empleado.tipo_contrato = tipo_contrato
        empleado.sueldo_fijo = sueldo_fijo
        empleado.rol_id = rol_id
        
        await db.commit()
        
        # ✅ ¡ÉXITO! Redirigir a lista de empleados con mensaje
        return RedirectResponse(url="/empleados?success_edit=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"Error al editar empleado: {str(e)}")
        # ❌ En caso de error, redirigir a la página de edición
        return RedirectResponse(
            url=f"/empleados/{empleado_id}/editar?error=1", 
            status_code=303
        )

# === EMPLEADOS: ELIMINAR ===
@app.post("/empleados/{empleado_id}/eliminar")
async def eliminar_empleado(
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    try:
        empleado = (await db.execute(
            select(models.Empleado).where(models.Empleado.id == empleado_id)
        )).scalar_one_or_none()
        
        if empleado:
            await db.delete(empleado)
            await db.commit()
        
        # ✅ Redirigir a lista con mensaje de éxito
        return RedirectResponse(url="/empleados?deleted=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"Error al eliminar empleado: {str(e)}")
        return RedirectResponse(url="/empleados", status_code=303)

# === PRÉSTAMOS: PÁGINA PRINCIPAL ===

from datetime import date
from fastapi import Query

@app.get("/prestamos", response_class=HTMLResponse)
async def gestion_prestamos(
    request: Request,
    fecha: str = Query(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    hoy = date.today()
    fecha_actual = hoy.isoformat()
    
    # Determinar fecha de filtro
    if fecha:
        try:
            fecha_filtro = date.fromisoformat(fecha)
        except ValueError:
            fecha_filtro = hoy
    else:
        fecha_filtro = hoy
    
    # ✅ CONSULTA MEJORADA: Asegurar que se obtengan todos los préstamos
    prestamos_query = await db.execute(
        select(models.Prestamo)
        .where(models.Prestamo.fecha == fecha_filtro)
        .order_by(models.Prestamo.fecha.desc())
    )
    prestamos = prestamos_query.scalars().all()
    
    # ✅ Procesar cada préstamo individualmente
    prestamos_data = []
    for prestamo in prestamos:
        # Obtener el nombre del empleado (incluso si el empleado fue eliminado)
        empleado_nombre = "Empleado eliminado"
        if prestamo.empleado_id:
            empleado_result = await db.execute(
                select(models.Empleado.nombre_completo)
                .where(models.Empleado.id == prestamo.empleado_id)
            )
            empleado_nombre = empleado_result.scalar() or "Empleado eliminado"
        
        prestamos_data.append({
            "id": prestamo.id,
            "empleado_id": prestamo.empleado_id,
            "monto": float(prestamo.monto),
            "descripcion": prestamo.descripcion or "-",
            "fecha": prestamo.fecha.isoformat(),
            "empleado_nombre": empleado_nombre
        })
    
    # Obtener empleados activos para el formulario
    empleados_activos = (await db.execute(
        select(models.Empleado).where(models.Empleado.activo == True)
    )).scalars().all()
    empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados_activos]
    
    return templates.TemplateResponse("prestamos/prestamos.html", {
        "request": request,
        "prestamos": prestamos_data,
        "empleados": empleados_data,
        "fecha_actual": fecha_actual,
        "fecha_filtro": fecha_filtro.isoformat(),
        "error": None
    })

@app.post("/prestamos", response_class=HTMLResponse)
async def crear_prestamo(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    form = await request.form()
    empleado_id = form.get("empleado_id")
    monto = form.get("monto")
    descripcion = form.get("descripcion") or None

    # Validación mejorada
    error_msg = None
    if not empleado_id:
        error_msg = "Debe seleccionar un empleado"
    elif not monto:
        error_msg = "El monto es requerido"
    else:
        try:
            monto_val = float(monto)
            if monto_val <= 0:
                error_msg = "El monto debe ser mayor a 0"
        except ValueError:
            error_msg = "Monto inválido: debe ser un número"
        
        try:
            emp_id_val = int(empleado_id)
        except ValueError:
            error_msg = "Empleado inválido"

    if error_msg:
        empleados = (await db.execute(select(models.Empleado).where(models.Empleado.activo == True))).scalars().all()
        prestamos = (await db.execute(select(models.Prestamo).order_by(models.Prestamo.fecha.desc()))).scalars().all()
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados]
        prestamos_data = [{"id": p.id, "empleado_id": p.empleado_id, "monto": float(p.monto), "descripcion": p.descripcion or "-", "fecha": p.fecha.isoformat()} for p in prestamos]
        return templates.TemplateResponse("prestamos/prestamos.html", {
            "request": request,
            "prestamos": prestamos_data,
            "empleados": empleados_data,
            "error": error_msg
        })
    
    try:
        db_prestamo = models.Prestamo(
            empleado_id=emp_id_val,
            monto=monto_val,
            descripcion=descripcion,
            fecha=date.today()
        )
        db.add(db_prestamo)
        await db.commit()
        
        # ✅ ¡CAMBIO AQUÍ! Agregar ?success=1 para mostrar mensaje
        return RedirectResponse(url="/prestamos?success=1", status_code=303)
        
    except Exception as e:
        print("ERROR DB:", str(e))
        empleados = (await db.execute(select(models.Empleado).where(models.Empleado.activo == True))).scalars().all()
        prestamos = (await db.execute(select(models.Prestamo).order_by(models.Prestamo.fecha.desc()))).scalars().all()
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados]
        prestamos_data = [{"id": p.id, "empleado_id": p.empleado_id, "monto": float(p.monto), "descripcion": p.descripcion or "-", "fecha": p.fecha.isoformat()} for p in prestamos]
        return templates.TemplateResponse("prestamos/prestamos.html", {
            "request": request,
            "prestamos": prestamos_data,
            "empleados": empleados_data,
            "error": "Error al guardar en la base de datos"
        })

# === PRÉSTAMOS: EDITAR ===
@app.get("/prestamos/editar/{prestamo_id}", response_class=HTMLResponse)
async def editar_prestamo_form(
    prestamo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    # Cargar préstamo con nombre del empleado
    result = await db.execute(
        select(
            models.Prestamo.id,
            models.Prestamo.empleado_id,
            models.Prestamo.monto,
            models.Prestamo.descripcion,
            models.Empleado.nombre_completo
        )
        .select_from(models.Prestamo)
        .join(models.Empleado, models.Prestamo.empleado_id == models.Empleado.id, isouter=True)
        .where(models.Prestamo.id == prestamo_id)
    )
    row = result.fetchone()
    
    if not row:
        return RedirectResponse(url="/prestamos")
    
    prestamo_data = {
        "id": row.id,
        "empleado_id": row.empleado_id,
        "monto": float(row.monto),
        "descripcion": row.descripcion or "",
        "empleado_nombre": row.nombre_completo or "Empleado eliminado"
    }
    
    # Empleados activos para el select
    empleados = (await db.execute(
        select(models.Empleado).where(models.Empleado.activo == True)
    )).scalars().all()
    empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados]
    
    return templates.TemplateResponse("prestamos/editar.html", {
        "request": request,
        "prestamo": prestamo_data,
        "empleados": empleados_data,
        "error": None
    })

@app.post("/prestamos/editar/{prestamo_id}", response_class=HTMLResponse)
async def actualizar_prestamo(
    prestamo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    form = await request.form()
    empleado_id = form.get("empleado_id")
    monto = form.get("monto")
    descripcion = form.get("descripcion") or None

    if not empleado_id or not monto:
        prestamo = await db.get(models.Prestamo, prestamo_id)
        empleados = (await db.execute(select(models.Empleado).where(models.Empleado.activo == True))).scalars().all()
        prestamo_data = {
            "id": prestamo.id,
            "empleado_id": prestamo.empleado_id,
            "monto": float(prestamo.monto),
            "descripcion": prestamo.descripcion or ""
        }
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados]
        return templates.TemplateResponse("prestamos/editar.html", {
            "request": request,
            "prestamo": prestamo_data,
            "empleados": empleados_data,
            "error": "Empleado y monto son requeridos"
        })
    
    try:
        prestamo = await db.get(models.Prestamo, prestamo_id)
        if prestamo:
            prestamo.empleado_id = int(empleado_id)
            prestamo.monto = float(monto)
            prestamo.descripcion = descripcion
            await db.commit()
        # Después de guardar con éxito
            return RedirectResponse(url="/prestamos?success_edit=1", status_code=303)
    except Exception as e:
        print("ERROR:", e)
        prestamo = await db.get(models.Prestamo, prestamo_id)
        empleados = (await db.execute(select(models.Empleado).where(models.Empleado.activo == True))).scalars().all()
        prestamo_data = {
            "id": prestamo.id,
            "empleado_id": prestamo.empleado_id,
            "monto": float(prestamo.monto),
            "descripcion": prestamo.descripcion or ""
        }
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo} for e in empleados]
        return templates.TemplateResponse("prestamos/editar.html", {
            "request": request,
            "prestamo": prestamo_data,
            "empleados": empleados_data,
            "error": "Error al actualizar"
        })

# === PRÉSTAMOS: ELIMINAR ===
@app.get("/prestamos/eliminar/{prestamo_id}")
async def eliminar_prestamo(prestamo_id: int, db: AsyncSession = Depends(get_db)):
    prestamo = await db.get(models.Prestamo, prestamo_id)
    if prestamo:
        await db.delete(prestamo)
        await db.commit()

    return RedirectResponse(url="/prestamos?deleted=1", status_code=303)  
    return RedirectResponse(url="/prestamos", status_code=303)

# === GASTOS: PÁGINA PRINCIPAL ===
from datetime import date
from fastapi import Query
@app.get("/finanzas/gastos", response_class=HTMLResponse)
async def listar_gastos(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    hoy = date.today()
    gastos = (await db.execute(
        select(GastoDiario)
        .options(joinedload(GastoDiario.categoria), joinedload(GastoDiario.subcategoria))
        .where(GastoDiario.fecha == hoy)
        .order_by(GastoDiario.id.desc())
    )).scalars().all()
    
    # Calcular totales
    totales_categoria = {}
    total_general = 0.0
    for gasto in gastos:
        cat_nombre = gasto.categoria.nombre if gasto.categoria else "sin_categoría"
        if cat_nombre not in totales_categoria:
            totales_categoria[cat_nombre] = 0.0
        totales_categoria[cat_nombre] += float(gasto.monto)
        total_general += float(gasto.monto)
    
    return templates.TemplateResponse("finanzas/gastos/gastos_list.html", {  # ← ¡CAMBIA ESTA LÍNEA!
        "request": request,
        "user": user,
        "gastos": gastos,
        "totales_categoria": totales_categoria,
        "total_general": round(total_general, 2),
        "hoy": hoy
    })

@app.get("/finanzas/gastos/nuevo", response_class=HTMLResponse)
async def gasto_nuevo_form(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener categorías y subcategorías activas
    categorias = (await db.execute(
        select(CategoriaGasto).where(CategoriaGasto.activa == True)
    )).scalars().all()
    
    # Agrupar subcategorías por categoría
    subcategorias_por_categoria = {}
    for categoria in categorias:
        subcategorias = (await db.execute(
            select(SubcategoriaGasto).where(
                SubcategoriaGasto.categoria_id == categoria.id,
                SubcategoriaGasto.activa == True
            )
        )).scalars().all()
        subcategorias_por_categoria[categoria.id] = [
            {"id": s.id, "nombre": s.nombre, "descripcion": s.descripcion}
            for s in subcategorias
        ]
    
    return templates.TemplateResponse("finanzas/gastos/gasto_nuevo.html", {
        "request": request,
        "user": user,
        "categorias": categorias,
        "subcategorias_por_categoria": subcategorias_por_categoria
    })


@app.post("/finanzas/gastos/crear")
async def crear_gasto(
    request: Request,
    monto: float = Form(...),
    descripcion: str = Form(...),
    categoria_id: int = Form(...),
    subcategoria_id: int = Form(...),
    fecha: str = Form(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Validar que la subcategoría pertenezca a la categoría
    subcategoria = (await db.execute(
        select(SubcategoriaGasto).where(
            SubcategoriaGasto.id == subcategoria_id,
            SubcategoriaGasto.categoria_id == categoria_id
        )
    )).scalar_one_or_none()
    
    if not subcategoria:
        return RedirectResponse(url="/finanzas/gastos/nuevo?error=invalid_subcategoria", status_code=303)
    
    # Crear gasto SIN tipo_pago/tipo_gasto
    nuevo_gasto = GastoDiario(
        monto=monto,
        descripcion=descripcion,
        categoria_id=categoria_id,
        subcategoria_id=subcategoria_id,
        fecha=datetime.fromisoformat(fecha).date() if fecha else date.today(),
       
    )
    db.add(nuevo_gasto)
    await db.commit()
    
    return RedirectResponse(url="/finanzas/gastos?success=created", status_code=303)

# Vista detalle
@app.get("/finanzas/gastos/{gasto_id}", response_class=HTMLResponse)
async def ver_detalle_gasto(
    gasto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    gasto = (await db.execute(
        select(GastoDiario)
        .options(joinedload(GastoDiario.categoria), joinedload(GastoDiario.subcategoria))
        .where(GastoDiario.id == gasto_id)
    )).scalar_one_or_none()
    
    if not gasto:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    return templates.TemplateResponse("finanzas/gastos/detalle_gasto.html", {
        "request": request,
        "user": user,
        "gasto": gasto
    })

# Formulario edición
@app.get("/finanzas/gastos/editar/{gasto_id}", response_class=HTMLResponse)
async def formulario_editar_gasto(
    gasto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener el gasto específico
    gasto = (await db.execute(
        select(GastoDiario).where(GastoDiario.id == gasto_id)
    )).scalar_one_or_none()
    
    if not gasto:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    # Obtener categorías y subcategorías
    categorias = (await db.execute(
        select(CategoriaGasto).where(CategoriaGasto.activa == True)
    )).scalars().all()
    
    # Agrupar subcategorías por categoría
    subcategorias_por_categoria = {}
    for categoria in categorias:
        subcategorias = (await db.execute(
            select(SubcategoriaGasto).where(
                SubcategoriaGasto.categoria_id == categoria.id,
                SubcategoriaGasto.activa == True
            )
        )).scalars().all()
        subcategorias_por_categoria[categoria.id] = [
            {"id": s.id, "nombre": s.nombre}
            for s in subcategorias
        ]
    
    return templates.TemplateResponse("finanzas/gastos/editar_gasto.html", {
        "request": request,
        "user": user,
        "gasto": gasto,  # ← ¡Importante!
        "categorias": categorias,
        "subcategorias_por_categoria": subcategorias_por_categoria
    })

# Actualizar gasto
@app.post("/finanzas/gastos/actualizar/{gasto_id}")
async def actualizar_gasto(
    gasto_id: int,
    monto: float = Form(...),
    descripcion: str = Form(...),
    categoria_id: int = Form(...),
    subcategoria_id: int = Form(...),
    fecha: str = Form(...),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Validar subcategoría pertenece a categoría
    subcategoria = (await db.execute(
        select(SubcategoriaGasto).where(
            SubcategoriaGasto.id == subcategoria_id,
            SubcategoriaGasto.categoria_id == categoria_id
        )
    )).scalar_one_or_none()
    
    if not subcategoria:
        return RedirectResponse(url=f"/finanzas/gastos/editar/{gasto_id}?error=subcategoria_invalida", status_code=303)
    
    # Actualizar gasto
    gasto = (await db.execute(
        select(GastoDiario).where(GastoDiario.id == gasto_id)
    )).scalar_one_or_none()
    
    if not gasto:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    gasto.monto = monto
    gasto.descripcion = descripcion
    gasto.categoria_id = categoria_id
    gasto.subcategoria_id = subcategoria_id
    gasto.fecha = datetime.fromisoformat(fecha).date()
    
    await db.commit()
    
    return RedirectResponse(url=f"/finanzas/gastos/{gasto_id}?success=updated", status_code=303)

# Eliminar gasto
@app.post("/finanzas/gastos/eliminar/{gasto_id}")
async def eliminar_gasto(
    gasto_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    gasto = (await db.execute(
        select(GastoDiario).where(GastoDiario.id == gasto_id)
    )).scalar_one_or_none()
    
    if gasto:
        await db.delete(gasto)
        await db.commit()
    
    return RedirectResponse(url="/finanzas/gastos?deleted=1", status_code=303)



# === TRABAJOS: PÁGINA PRINCIPAL ==

@app.get("/trabajos/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_trabajo(
    request: Request,
    presupuesto_id: Optional[int] = None,
    error: str = None,
    success: str = None,
    msg: str = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    def safe_float(value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    try:
        # ✅ Obtener clientes con su empleado vinculado
        clientes_result = await db.execute(
            select(models.Cliente).where(models.Cliente.activo == True)
        )
        clientes = []
        for c in clientes_result.scalars().all():
            cliente_data = {
                "id": c.id,
                "nombre": c.nombre_razon_social,
                "empleado_id": c.empleado_id,
                "empleado_nombre": None
            }
            # Cargar empleado si existe
            if c.empleado_id:
                emp_result = await db.execute(
                    select(models.Empleado).where(models.Empleado.id == c.empleado_id)
                )
                empleado = emp_result.scalar_one_or_none()
                if empleado:
                    cliente_data["empleado_nombre"] = empleado.nombre_completo
            clientes.append(cliente_data)
        
        # Obtener roles
        roles_result = await db.execute(select(models.Rol).order_by(models.Rol.nombre))
        roles = [{"id": r.id, "nombre": r.nombre} for r in roles_result.scalars().all()]
        
        # Obtener empleados activos
        empleados_result = await db.execute(
            select(models.Empleado).where(models.Empleado.activo == True)
        )
        empleados = [
            {"id": e.id, "nombre": e.nombre_completo, "rol_id": e.rol_id} 
            for e in empleados_result.scalars().all()
        ]
        
        # Obtener presupuestos aprobados
        presupuestos_result = await db.execute(
            select(models.Presupuesto)
            .where(models.Presupuesto.estado == "aprobado")
            .join(models.Cliente)
            .order_by(models.Presupuesto.fecha_creacion.desc())
        )
        presupuestos_aprobados = []
        for pres in presupuestos_result.scalars().all():
            cliente = await db.get(models.Cliente, pres.cliente_id)
            presupuestos_aprobados.append({
                "id": pres.id,
                "nombre": pres.nombre_trabajo,
                "cliente_id": pres.cliente_id,
                "monto_usd": safe_float(pres.total_base),
                "monto_bs": safe_float(pres.total_cliente)
            })
        
        # ============================================================
        # ✅ NUEVO: OBTENER MATERIALES DEL INVENTARIO
        # ============================================================
        materiales_result = await db.execute(
            select(models.MaterialInventario)
            .where(models.MaterialInventario.activo == True)
            .order_by(models.MaterialInventario.nombre)
        )
        materiales_inventario = []
        for m in materiales_result.scalars().all():
            # Obtener categoría
            cat_result = await db.execute(
                select(models.CategoriaInventario)
                .where(models.CategoriaInventario.id == m.categoria_id)
            )
            categoria = cat_result.scalar_one_or_none()
            materiales_inventario.append({
                "id": m.id,
                "nombre": m.nombre,
                "stock_actual": float(m.stock_actual) if m.stock_actual else 0,
                "unidad_medida": m.unidad_medida or "unidades",
                "categoria": categoria
            })
        
        # Procesar presupuesto seleccionado
        datos_presupuesto = None
        if presupuesto_id:
            presupuesto = await db.get(models.Presupuesto, presupuesto_id)
            if presupuesto and presupuesto.estado == "aprobado":
                cliente = await db.get(models.Cliente, presupuesto.cliente_id)
                datos_presupuesto = {
                    "id": presupuesto.id,
                    "nombre_trabajo": presupuesto.nombre_trabajo,
                    "monto_total_usd": safe_float(presupuesto.total_base),
                    "monto_total_bs": safe_float(presupuesto.total_cliente),
                    "cliente_id": presupuesto.cliente_id,
                    "cliente_nombre": cliente.nombre_razon_social if cliente else ""
                }
        
        # Obtener tasa actual
        from app.services.currency_service import CurrencyService
        currency_service = CurrencyService(db)
        tasa_actual = await currency_service.get_tasa()
        
        return templates.TemplateResponse("trabajos/nuevo.html", {
            "request": request,
            "clientes": clientes,
            "roles": roles,
            "empleados": empleados,
            "presupuestos_aprobados": presupuestos_aprobados,
            "datos_presupuesto": datos_presupuesto,
            "materiales_inventario": materiales_inventario,  # ✅ NUEVO
            "tasa_cambio_actual": tasa_actual,
            "hoy": date.today(),
            "error": error,
            "success": success,
            "msg": msg,
            "tasa_actual": float(tasa_actual)
        })
        
    except Exception as e:
        print(f"Error al cargar formulario: {e}")
        import traceback
        traceback.print_exc()
        error_msg = "Error al cargar el formulario"
        return RedirectResponse(url=f"/trabajos/nuevo?error={quote(error_msg)}", status_code=303)


# --- Funciones de Utilidad ---
def safe_float(value, default=0.0):
    if value is None or value == "":
        return default
    try:
        # Maneja tanto puntos como comas
        if isinstance(value, str):
            value = value.replace(',', '.')
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def limpiar_nombre_archivo(filename):
    # Alternativa a secure_filename si no tienes werkzeug instalado
    return re.sub(r'[^a-zA-Z0-9._-]', '_', filename)

from urllib.parse import quote
import re
import os
from datetime import datetime
from fastapi import Request, File, UploadFile, Form, Depends
from fastapi.responses import RedirectResponse
from sqlalchemy import select

@app.post("/trabajos", response_model=None)
async def crear_trabajo(
    request: Request,
    archivos: List[UploadFile] = File(default=[]), 
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        form_data = await request.form()

        # ============================================================
        # 1. DATOS BÁSICOS
        # ============================================================
        monto_total_val = safe_float(form_data.get("monto_total"))
        tasa = safe_float(form_data.get("tasa_cambio"), default=1.0)
        porcentaje_pagado = safe_int(form_data.get("porcentaje_pagado", 0))
        monto_pagado_usd = (monto_total_val * porcentaje_pagado) / 100
        monto_pagado_bs = monto_pagado_usd * tasa

        metodo_pago = form_data.get("metodo_pago", "efectivo_usd")
        tipo_trabajo = form_data.get("tipo_trabajo", "rotulado_instalacion")
        
        # ============================================================
        # 2. CREAR TRABAJO
        # ============================================================
        unidades_val = safe_int(form_data.get("unidades", 0))
        print(f"🔍 DEBUG - unidades recibidas del formulario: {unidades_val}")
        
        nuevo_trabajo = models.Trabajo(
            cliente_id=safe_int(form_data.get("cliente_id")),
            nombre_trabajo=form_data.get("nombre_trabajo", "Sin Nombre"),
            descripcion=form_data.get("descripcion", ""),
            monto_total=monto_total_val,
            monto_total_usd=monto_total_val,
            monto_total_bs=monto_total_val * tasa,
            tasa_cambio=tasa,
            estado=form_data.get("estado", "pendiente"),
            fecha_inicio=datetime.now().date(),
            creado_por=user.id,
            metros_cuadrados=safe_float(form_data.get("metros_cuadrados")),
            unidades=unidades_val,
            porcentaje_pagado=porcentaje_pagado,
            monto_pagado=monto_pagado_usd,
            monto_pagado_usd=monto_pagado_usd,
            monto_pagado_bs=monto_pagado_bs,
            tasa_cambio_actual=tasa,
            prioridad=form_data.get("prioridad", "media"),
            metodo_pago=metodo_pago,
            fecha_pago=datetime.now() if porcentaje_pagado > 0 else None,
            tipo_trabajo=tipo_trabajo
        )
        
        db.add(nuevo_trabajo)
        await db.flush()
        
        print(f"📝 Trabajo creado ID: {nuevo_trabajo.id} - Tipo: {tipo_trabajo}")
        print(f"📊 Unidades guardadas en trabajo: {nuevo_trabajo.unidades}")

        # ============================================================
        # 3. FUNCIÓN PARA DESCONTAR INVENTARIO (CON DEBUG MEJORADO)
        # ============================================================
        import re
        
        async def descontar_inventario(nombre_material, cantidad_str, tipo_material):
            print(f"\n🔍 DESCONTAR INVENTARIO:")
            print(f"   - Material: '{nombre_material}'")
            print(f"   - Cantidad: '{cantidad_str}'")
            print(f"   - Tipo: '{tipo_material}'")
            
            if not nombre_material or not cantidad_str:
                print(f"   ⚠️ Material o cantidad vacío")
                return False
            
            # Extraer cantidad numérica
            match = re.search(r'(\d+(?:\.\d+)?)', str(cantidad_str).replace(',', '.'))
            cantidad = float(match.group(1)) if match else 0
            
            print(f"   - Cantidad numérica: {cantidad}")
            
            if cantidad <= 0:
                print(f"   ⚠️ Cantidad <= 0, no se descuenta")
                return False
            
            # 🔥 PRIMERO: Buscar por nombre exacto (case insensitive)
            print(f"   - Buscando exacto: '{nombre_material.strip()}'")
            result = await db.execute(
                select(models.MaterialInventario)
                .where(func.lower(models.MaterialInventario.nombre) == func.lower(nombre_material.strip()))
                .where(models.MaterialInventario.activo == True)
                .limit(1)
            )
            material_row = result.first()
            
            # 🔥 SEGUNDO: Buscar por coincidencia parcial (contiene el nombre)
            if not material_row:
                print(f"   - No encontrado exacto, buscando LIKE '%{nombre_material.strip()}%'")
                result = await db.execute(
                    select(models.MaterialInventario)
                    .where(models.MaterialInventario.nombre.ilike(f"%{nombre_material.strip()}%"))
                    .where(models.MaterialInventario.activo == True)
                    .limit(1)
                )
                material_row = result.first()
            
            # 🔥 TERCERO: Buscar palabra por palabra (más flexible)
            if not material_row:
                palabras = nombre_material.strip().split()
                print(f"   - Buscando por palabras: {palabras}")
                for palabra in palabras:
                    if len(palabra) > 2:
                        result = await db.execute(
                            select(models.MaterialInventario)
                            .where(models.MaterialInventario.nombre.ilike(f"%{palabra}%"))
                            .where(models.MaterialInventario.activo == True)
                            .limit(1)
                        )
                        material_row = result.first()
                        if material_row:
                            print(f"   - Encontrado por palabra '{palabra}'")
                            break
            
            if not material_row:
                print(f"   ❌ Material NO ENCONTRADO: {nombre_material}")
                # 🔥 Mostrar materiales disponibles para DEBUG
                all_materials = await db.execute(
                    select(models.MaterialInventario.nombre).where(models.MaterialInventario.activo == True).limit(20)
                )
                materiales = [m[0] for m in all_materials.all()]
                print(f"   📋 Materiales disponibles: {materiales}")
                return False
            
            material = material_row[0]
            print(f"   ✅ Material encontrado: {material.nombre} (ID: {material.id})")
            
            stock_actual = float(material.stock_actual)
            print(f"   - Stock actual: {stock_actual}")
            
            if stock_actual >= cantidad:
                material.stock_actual = stock_actual - cantidad
                material.fecha_actualizacion = datetime.utcnow()
                
                movimiento = models.MovimientoInventario(
                    material_id=material.id,
                    cantidad=cantidad,
                    tipo='salida',
                    motivo=f"Consumo en trabajo #{nuevo_trabajo.id} - {tipo_material}",
                    referencia=f"TRABAJO_{nuevo_trabajo.id}",
                    usuario_id=user.id,
                    trabajo_id=nuevo_trabajo.id
                )
                db.add(movimiento)
                print(f"   ✅ Descontado: {material.nombre} - {cantidad} {material.unidad_medida}")
                print(f"   - Nuevo stock: {material.stock_actual}")
                return True
            else:
                print(f"   ❌ Stock INSUFICIENTE: {material.nombre} - disponible: {stock_actual}, requerido: {cantidad}")
                return False

        # ============================================================
        # 4A. ROTULADO - Procesar materiales
        # ============================================================
        if tipo_trabajo == 'rotulado_instalacion':
            m_conceptos = form_data.getlist("materiales_concepto[]")
            m_cantidades = form_data.getlist("materiales_cantidad[]")
            
            print(f"\n📦 PROCESANDO ROTULADO - {len(m_conceptos)} materiales")
            
            for i in range(len(m_conceptos)):
                concepto = m_conceptos[i].strip() if m_conceptos[i] else ""
                cantidad_str = m_cantidades[i].strip() if m_cantidades[i] else ""
                
                if concepto and cantidad_str:
                    match = re.search(r'(\d+(?:\.\d+)?)', cantidad_str)
                    cantidad_num = float(match.group(1)) if match else 0
                    
                    material_usado = models.MaterialUsado(
                        trabajo_id=nuevo_trabajo.id,
                        concepto=concepto,
                        cantidad_usada=cantidad_num,
                        costo_unitario=0,
                        material_id=None
                    )
                    db.add(material_usado)
                    
                    await descontar_inventario(concepto, cantidad_str, "rotulado")

        # ============================================================
        # 4B. TEXTIL - Procesar materiales (CON ACUMULADOR DE TELA)
        # ============================================================
        elif tipo_trabajo == 'textil':
            print(f"\n👕 PROCESANDO TEXTIL")
            
            prendas = form_data.getlist("textil_prenda[]")
            tallas = form_data.getlist("textil_talla[]")
            cantidades_prendas = form_data.getlist("textil_cantidad[]")
            telas_nombre = form_data.getlist("textil_tela_nombre[]")
            telas_cantidad = form_data.getlist("textil_tela_cantidad[]")
            m2_por_prenda = form_data.getlist("textil_m2_por_prenda[]")
            papeles_nombre = form_data.getlist("textil_papel_nombre[]")
            papeles_cantidad = form_data.getlist("textil_papel_cantidad[]")
            tinta_cyan_nombre = form_data.getlist("textil_tinta_cyan_nombre[]")
            tinta_cyan_cantidad = form_data.getlist("textil_tinta_cyan_cantidad[]")
            tinta_magenta_nombre = form_data.getlist("textil_tinta_magenta_nombre[]")
            tinta_magenta_cantidad = form_data.getlist("textil_tinta_magenta_cantidad[]")
            tinta_amarilla_nombre = form_data.getlist("textil_tinta_amarilla_nombre[]")
            tinta_amarilla_cantidad = form_data.getlist("textil_tinta_amarilla_cantidad[]")
            tinta_negra_nombre = form_data.getlist("textil_tinta_negra_nombre[]")
            tinta_negra_cantidad = form_data.getlist("textil_tinta_negra_cantidad[]")
            
            # 🔥 DICCIONARIO PARA ACUMULAR TELA POR TIPO
            acumulador_tela = {}
            
            print(f"📊 m² recibidos: {m2_por_prenda}")
            
            for idx in range(len(prendas)):
                if not prendas[idx]:
                    continue
                
                print(f"\n--- Prenda {idx+1}: {prendas[idx]} ---")
                
                # Extraer cantidad de prendas
                cantidad_prendas = 1
                if idx < len(cantidades_prendas):
                    try:
                        cantidad_prendas = int(cantidades_prendas[idx]) if cantidades_prendas[idx] else 1
                    except ValueError:
                        cantidad_prendas = 1
                print(f"   - Cantidad prendas: {cantidad_prendas}")
                
                # Extraer m² por prenda
                m2_valor = 0.0
                if idx < len(m2_por_prenda):
                    try:
                        m2_valor = float(m2_por_prenda[idx]) if m2_por_prenda[idx] else 0.0
                        print(f"   - m² valor: {m2_valor}")
                    except ValueError:
                        m2_valor = 0.0
                
                # 🔥 OBTENER TELA Y ACUMULAR
                tela_nombre = telas_nombre[idx] if idx < len(telas_nombre) else ''
                tela_cantidad_str = telas_cantidad[idx] if idx < len(telas_cantidad) else ''
                print(f"   - Tela: '{tela_nombre}' - Cantidad: '{tela_cantidad_str}'")
                
                # 🔥 ACUMULAR TELA
                if tela_nombre and tela_cantidad_str:
                    import re
                    match = re.search(r'(\d+(?:\.\d+)?)', str(tela_cantidad_str).replace(',', '.'))
                    cantidad_tela = float(match.group(1)) if match else 0
                    
                    if tela_nombre in acumulador_tela:
                        acumulador_tela[tela_nombre] += cantidad_tela
                        print(f"   - Tela acumulada: {tela_nombre} -> {acumulador_tela[tela_nombre]}")
                    else:
                        acumulador_tela[tela_nombre] = cantidad_tela
                        print(f"   - Tela nueva: {tela_nombre} -> {acumulador_tela[tela_nombre]}")
                
                detalle = models.TrabajoMaterialTextil(
                    trabajo_id=nuevo_trabajo.id,
                    prenda=prendas[idx],
                    talla=tallas[idx] if idx < len(tallas) else '',
                    cantidad=cantidad_prendas,
                    tela_nombre=tela_nombre,
                    tela_cantidad=tela_cantidad_str,
                    m2_por_prenda=m2_valor,
                    papel_nombre=papeles_nombre[idx] if idx < len(papeles_nombre) else '',
                    papel_cantidad=papeles_cantidad[idx] if idx < len(papeles_cantidad) else '',
                    tinta_nombre=f"Cyan: {tinta_cyan_nombre[idx] if idx < len(tinta_cyan_nombre) else ''}",
                    tinta_cantidad=tinta_cyan_cantidad[idx] if idx < len(tinta_cyan_cantidad) else ''
                )
                db.add(detalle)
                print(f"✅ Prenda guardada: {detalle.prenda} x {detalle.cantidad} - m²: {detalle.m2_por_prenda}")
                
                # Descontar papeles (si hay)
                if idx < len(papeles_nombre) and papeles_nombre[idx]:
                    await descontar_inventario(papeles_nombre[idx], papeles_cantidad[idx] if idx < len(papeles_cantidad) else "", "papel")
                
                # Descontar tintas (si hay)
                if idx < len(tinta_cyan_nombre) and tinta_cyan_nombre[idx]:
                    await descontar_inventario(tinta_cyan_nombre[idx], tinta_cyan_cantidad[idx] if idx < len(tinta_cyan_cantidad) else "", "tinta cyan")
                
                if idx < len(tinta_magenta_nombre) and tinta_magenta_nombre[idx]:
                    await descontar_inventario(tinta_magenta_nombre[idx], tinta_magenta_cantidad[idx] if idx < len(tinta_magenta_cantidad) else "", "tinta magenta")
                
                if idx < len(tinta_amarilla_nombre) and tinta_amarilla_nombre[idx]:
                    await descontar_inventario(tinta_amarilla_nombre[idx], tinta_amarilla_cantidad[idx] if idx < len(tinta_amarilla_cantidad) else "", "tinta amarilla")
                
                if idx < len(tinta_negra_nombre) and tinta_negra_nombre[idx]:
                    await descontar_inventario(tinta_negra_nombre[idx], tinta_negra_cantidad[idx] if idx < len(tinta_negra_cantidad) else "", "tinta negra")
            
            # ============================================================
            # 🔥 DESCONTAR TELA ACUMULADA (UNA SOLA VEZ POR TIPO)
            # ============================================================
            print(f"\n📊 ACUMULADOR DE TELA FINAL: {acumulador_tela}")
            for tela_nombre, cantidad_total in acumulador_tela.items():
                cantidad_str = str(cantidad_total)
                print(f"   🔥 Descontando tela acumulada: {tela_nombre} - {cantidad_str}")
                await descontar_inventario(tela_nombre, cantidad_str, "tela textil (acumulada)")
            
            # ============================================================
            # PAPEL DE SUBLIMACIÓN (TEXTIL)
            # ============================================================
            textil_papel_nombre = form_data.get("textil_papel_nombre", "")
            textil_papel_cantidad = form_data.get("textil_papel_cantidad", "")
            
            if textil_papel_nombre and textil_papel_cantidad:
                nuevo_trabajo.papel_sublimacion_nombre = textil_papel_nombre
                nuevo_trabajo.papel_sublimacion_cantidad = textil_papel_cantidad
                print(f"📄 Papel guardado en trabajo: {textil_papel_nombre} - {textil_papel_cantidad}")
                
                await descontar_inventario(textil_papel_nombre, textil_papel_cantidad, "papel sublimación textil")
                print(f"📄 Papel descontado: {textil_papel_nombre} - {textil_papel_cantidad}")
            else:
                print(f"⚠️ No se procesó papel - nombre: '{textil_papel_nombre}', cantidad: '{textil_papel_cantidad}'")

        # ============================================================
        # 5. SERVICIOS EXTERNOS
        # ============================================================
        s_conceptos = form_data.getlist("servicios_concepto[]")
        for concepto in s_conceptos:
            if concepto and concepto.strip():
                db.add(models.ServicioExterno(
                    trabajo_id=nuevo_trabajo.id,
                    concepto=concepto.strip(),
                    proveedor=None,
                    costo=0
                ))

        # ============================================================
        # 6. COMISIONES
        # ============================================================
        roles_res = await db.execute(select(models.Rol))
        roles = roles_res.scalars().all()
        total_com_usd = 0.0
        
        print(f"\n📊 CALCULANDO COMISIONES - Unidades disponibles: {nuevo_trabajo.unidades}")
        
        for r in roles:
            emps = form_data.getlist(f"empleado_{r.id}[]")
            if emps:
                valor_unitario_usuario = safe_float(form_data.get(f"valor_comision_{r.id}"))
                tipo_comision = form_data.get(f"tipo_comision_{r.id}", "fijo")
                
                total_rol = 0.0
                if tipo_comision == "porcentaje": 
                    total_rol = (monto_total_val * valor_unitario_usuario) / 100
                elif tipo_comision == "porcentaje_ganancia":
                    total_rol = (monto_total_val * valor_unitario_usuario) / 100
                elif tipo_comision == "metro": 
                    total_rol = (nuevo_trabajo.metros_cuadrados or 0) * valor_unitario_usuario
                elif tipo_comision == "unidad": 
                    total_rol = (nuevo_trabajo.unidades or 0) * valor_unitario_usuario
                    print(f"🔍 Comisión x unidad - Rol: {r.nombre}, Unidades: {nuevo_trabajo.unidades}, Valor: {valor_unitario_usuario}, Total: {total_rol}")
                elif tipo_comision == "fijo": 
                    total_rol = valor_unitario_usuario
                
                total_com_usd += total_rol
                valor_por_empleado = total_rol / len(emps) if total_rol > 0 and len(emps) > 0 else 0
                
                for eid in emps:
                    db.add(models.Asignacion(
                        trabajo_id=nuevo_trabajo.id,
                        rol_id=r.id,
                        empleado_id=int(eid),
                        tipo_comision=tipo_comision,
                        valor_unitario=valor_unitario_usuario,
                        valor_comision=valor_por_empleado 
                    ))

        nuevo_trabajo.total_comisiones_usd = total_com_usd
        nuevo_trabajo.total_materiales_usd = 0
        nuevo_trabajo.servicios_externos_usd = 0
        nuevo_trabajo.ganancia_neta_usd = monto_total_val - total_com_usd

        # ============================================================
        # 7. COMPRA POR NÓMINA
        # ============================================================
        if metodo_pago == 'nomina':
            cliente = await db.get(models.Cliente, nuevo_trabajo.cliente_id)
            if cliente and cliente.empleado_id:
                compra = models.CompraEmpleado(
                    empleado_id=cliente.empleado_id,
                    trabajo_id=nuevo_trabajo.id,
                    tipo_producto='trabajo',
                    descripcion_producto=f"Trabajo: {nuevo_trabajo.nombre_trabajo}",
                    cantidad=1,
                    precio_unitario=monto_total_val,
                    subtotal=monto_total_val,
                    descuento_aplicado=0,
                    total_a_descontar=monto_total_val,
                    estado_pago='pendiente',
                    fecha_compra=datetime.utcnow()
                )
                db.add(compra)

        # ============================================================
        # 8. ARCHIVOS
        # ============================================================
        print(f"\n📁 Procesando {len(archivos)} archivos...")
        
        for i, archivo in enumerate(archivos):
            if archivo.filename:
                try:
                    folder = f"uploads/trabajos/{nuevo_trabajo.id}"
                    os.makedirs(folder, exist_ok=True)
                    nombre_seguro = re.sub(r'[^\w\-.]', '_', archivo.filename) or f"archivo_{i+1}"
                    if '.' not in nombre_seguro:
                        nombre_seguro += ".bin"
                    filepath = f"{folder}/{nombre_seguro}"
                    contenido = await archivo.read()
                    with open(filepath, "wb") as f:
                        f.write(contenido)
                    db.add(models.ArchivoTrabajo(
                        trabajo_id=nuevo_trabajo.id,
                        nombre_original=archivo.filename,
                        nombre_guardado=nombre_seguro,
                        ruta_completa=filepath,
                        tipo_mime=getattr(archivo, 'content_type', 'application/octet-stream'),
                        tamano_bytes=len(contenido)
                    ))
                    print(f"✅ Archivo guardado: {archivo.filename}")
                except Exception as e:
                    print(f"❌ Error archivo {archivo.filename}: {e}")

        # ============================================================
        # 9. DESCUENTO DE TINTA CMYK
        # ============================================================
        tinta_total_litros = safe_float(form_data.get("tinta_total_litros", 0))
        
        if tinta_total_litros > 0:
            tinta_por_color = tinta_total_litros / 4
            print(f"\n🎨 DESCONTANDO TINTA: {tinta_total_litros} litros totales -> {tinta_por_color:.4f} litros por color")
            
            tintas = [
                {"nombre": "Tinta Cyan", "color": "cyan"},
                {"nombre": "Tinta Magenta", "color": "magenta"},
                {"nombre": "Tinta Amarilla", "color": "yellow"},
                {"nombre": "Tinta Negra", "color": "black"}
            ]
            
            for tinta in tintas:
                result = await db.execute(
                    select(models.MaterialInventario)
                    .where(func.lower(models.MaterialInventario.nombre) == func.lower(tinta["nombre"]))
                    .where(models.MaterialInventario.activo == True)
                )
                material = result.scalar_one_or_none()
                
                if material:
                    stock_actual = float(material.stock_actual)
                    if stock_actual >= tinta_por_color:
                        material.stock_actual = stock_actual - tinta_por_color
                        material.fecha_actualizacion = datetime.utcnow()
                        
                        movimiento = models.MovimientoInventario(
                            material_id=material.id,
                            cantidad=tinta_por_color,
                            tipo='salida',
                            motivo=f"Consumo impresión - Trabajo #{nuevo_trabajo.id}",
                            referencia=f"TRABAJO_{nuevo_trabajo.id}",
                            usuario_id=user.id,
                            trabajo_id=nuevo_trabajo.id,
                            observaciones=f"Tinta {tinta['color']} - Total: {tinta_total_litros} L"
                        )
                        db.add(movimiento)
                        print(f"✅ Descontado: {material.nombre} - {tinta_por_color:.4f} litros")
                    else:
                        print(f"⚠️ Stock insuficiente para {material.nombre}: {stock_actual} litros disponibles")
                else:
                    print(f"⚠️ Material no encontrado: {tinta['nombre']}")

        # ============================================================
        # 10. COMMIT FINAL
        # ============================================================
        await db.commit()
        
        print(f"\n🎉 TRABAJO #{nuevo_trabajo.id} CREADO EXITOSAMENTE")
        print(f"📊 Comisiones totales: {total_com_usd}")
        return RedirectResponse(url="/trabajos?success=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url=f"/trabajos/nuevo?error={quote(str(e))}", status_code=303)
    
@app.post("/trabajos/{trabajo_id}/actualizar-pago")
async def actualizar_pago_trabajo(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        form_data = await request.form()
        nuevo_porcentaje = safe_int(form_data.get("porcentaje_pagado"))
        tasa_actual = safe_float(form_data.get("tasa_cambio_actual"), default=1.0)
        
        # Obtener trabajo existente
        trabajo = await db.get(models.Trabajo, trabajo_id)
        if not trabajo:
            raise HTTPException(status_code=404, detail="Trabajo no encontrado")
        
        # Calcular nuevo monto pagado en USD
        nuevo_monto_usd = (trabajo.monto_total * nuevo_porcentaje) / 100
        
        # Calcular en Bs usando la tasa ACTUAL
        nuevo_monto_bs = nuevo_monto_usd * tasa_actual
        
        # Actualizar
        trabajo.porcentaje_pagado = nuevo_porcentaje
        trabajo.monto_pagado_usd = nuevo_monto_usd
        trabajo.monto_pagado_bs = nuevo_monto_bs
        trabajo.tasa_cambio_actual = tasa_actual
        
        await db.commit()
        
        return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}?success=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}?error={quote(str(e))}", status_code=303)

@app.post("/trabajos/{trabajo_id}/actualizar-estado")
async def actualizar_estado_trabajo(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    nuevo_estado = form_data.get("estado")
    
    estados_validos = ["pendiente", "en_ejecucion", "finalizado", "entregado", "completado"]
    if nuevo_estado not in estados_validos:
        return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}?error=Estado+no+válido", status_code=303)
    
    trabajo = await db.get(models.Trabajo, trabajo_id)
    if trabajo:
        trabajo.estado = nuevo_estado
        await db.commit()
    
    return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}", status_code=303)

# Ruta para actualizar fecha de entrega
@app.post("/trabajos/{trabajo_id}/actualizar-fecha-entrega")
async def actualizar_fecha_entrega(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    fecha_entrega_str = form_data.get("fecha_entrega")
    
    trabajo = await db.get(models.Trabajo, trabajo_id)
    if trabajo:
        if fecha_entrega_str:
            try:
                fecha_entrega = datetime.strptime(fecha_entrega_str, "%Y-%m-%dT%H:%M")
                trabajo.fecha_entrega = fecha_entrega
            except ValueError:
                pass
        else:
            trabajo.fecha_entrega = None
        
        await db.commit()
    
    return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}", status_code=303)

# Ruta para actualizar pago (ya la tienes, pero asegúrate que incluya la lógica de estados)
@app.post("/trabajos/{trabajo_id}/actualizar-pago")
async def actualizar_pago_trabajo(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    porcentaje = int(form_data.get("porcentaje_pagado", 0))
    tasa_actual = float(form_data.get("tasa_cambio_actual", 36.5))
    
    porcentaje = max(0, min(100, porcentaje))
    
    trabajo = await db.get(models.Trabajo, trabajo_id)
    if trabajo:
        trabajo.porcentaje_pagado = porcentaje
        monto_total_usd = float(trabajo.monto_total or 0)
        monto_pagado_usd = (monto_total_usd * porcentaje / 100)
        
        trabajo.monto_pagado_usd = monto_pagado_usd
        trabajo.monto_pagado_bs = monto_pagado_usd * tasa_actual
        trabajo.tasa_cambio_actual = tasa_actual
        
        # Actualizar estado según pago
        if porcentaje == 100:
            trabajo.estado = "completado"
        elif porcentaje > 0 and trabajo.estado == "pendiente":
            trabajo.estado = "en_ejecucion"
        
        await db.commit()
    
    return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}", status_code=303)


async def calcular_resumen_trabajo(trabajo, async_db=None):
    """
    Calcula el resumen completo de un trabajo:
    - Comisiones
    - Materiales  
    - Servicios externos
    - Ganancia neta
    """
    if not async_db:
        return {
            "comisiones": [],
            "materiales": [],
            "servicios_externos": [],
            "total_comisiones": 0.0,
            "total_materiales": 0.0,
            "total_servicios_externos": 0.0,
            "ganancia_neta": float(trabajo.monto_total) if trabajo.monto_total else 0.0
        }
    
    # === COMISIONES ===
    asignaciones_result = await async_db.execute(
        select(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo.id)
    )
    asignaciones = asignaciones_result.scalars().all()
    
    comisiones_detalles = []
    total_comisiones = 0.0
    
    for asignacion in asignaciones:
        empleado = await async_db.get(models.Empleado, asignacion.empleado_id)
        emp_nombre = empleado.nombre_completo if empleado else "Empleado desconocido"
        
        comision_valor = float(asignacion.valor_comision) if asignacion.valor_comision else 0.0
        
        comisiones_detalles.append({
            "empleado": emp_nombre,
            "rol": asignacion.rol.nombre if hasattr(asignacion, 'rol') else "Rol desconocido",
            "tipo_comision": asignacion.tipo_comision,
            "valor_comision": comision_valor
        })
        
        total_comisiones += comision_valor
    
    # === MATERIALES ===
    materiales_result = await async_db.execute(
        select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == trabajo.id)
    )
    materiales = materiales_result.scalars().all()
    
    materiales_detalles = []
    total_materiales = 0.0
    
    for material in materiales:
        costo_material = float(material.costo_unitario or 0) * float(material.cantidad_usada or 0)
        total_materiales += costo_material
        
        materiales_detalles.append({
            "concepto": material.concepto or f"Material ID {material.material_id}",
            "cantidad": float(material.cantidad_usada or 0),
            "costo_unitario": float(material.costo_unitario or 0),
            "costo_total": costo_material
        })
    
    # === SERVICIOS EXTERNOS ===
    # Si tienes una tabla separada para servicios externos, agrégala aquí
    # Por ahora, asumimos que están en el mismo trabajo
    total_servicios_externos = float(getattr(trabajo, 'servicios_externos', 0.0))
    
    # === GANANCIA NETA ===
    monto_total = float(trabajo.monto_total) if trabajo.monto_total else 0.0
    ganancia_neta = monto_total - (total_comisiones + total_materiales + total_servicios_externos)
    
    return {
        "comisiones": comisiones_detalles,
        "materiales": materiales_detalles,
        "servicios_externos": [],  # Ajusta según tu estructura
        "total_comisiones": total_comisiones,
        "total_materiales": total_materiales,
        "total_servicios_externos": total_servicios_externos,
        "ganancia_neta": ganancia_neta
    }

@app.get("/trabajos/editar/{id}")
async def editar_trabajo_form(
    request: Request,
    id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    """Mostrar formulario para editar trabajo - CORREGIDO AL 100%"""
    try:
        if not user:
            return RedirectResponse(url="/login", status_code=303)
        
        # Obtener trabajo
        trabajo_db = await db.execute(
            select(models.Trabajo).where(models.Trabajo.id == id)
        )
        trabajo = trabajo_db.scalar_one_or_none()
        if not trabajo:
            raise HTTPException(status_code=404, detail="Trabajo no encontrado")
        
        # Obtener asignaciones
        asignaciones_db = await db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == id)
        )
        asignaciones = list(asignaciones_db.scalars().all())
        
        # Procesar asignaciones individuales
        asignaciones_list = []
        empleados_por_rol = {}
        
        for asig in asignaciones:
            if not asig.empleado_id or not asig.rol_id:
                continue
                
            empleado_db = await db.get(models.Empleado, asig.empleado_id)
            rol_db = await db.get(models.Rol, asig.rol_id)
            
            # Determinar el valor unitario correcto con fallback seguro
            valor_unitario = 0.0
            if asig.valor_unitario is not None:
                valor_unitario = float(asig.valor_unitario)
            elif asig.valor_comision is not None:
                # Usar valor_comision como fallback si valor_unitario no existe
                valor_unitario = float(asig.valor_comision)
            
            # Nombre del empleado seguro - PRIORIDAD: nombre_completo > nombre
            empleado_nombre = f"Empleado {asig.empleado_id}"
            if empleado_db:
                if hasattr(empleado_db, 'nombre_completo') and empleado_db.nombre_completo:
                    empleado_nombre = empleado_db.nombre_completo
                elif hasattr(empleado_db, 'nombre') and empleado_db.nombre:
                    empleado_nombre = empleado_db.nombre
            
            # Nombre del rol seguro  
            rol_nombre = f"Rol {asig.rol_id}"
            if rol_db and hasattr(rol_db, 'nombre') and rol_db.nombre:
                rol_nombre = rol_db.nombre
            
            asignacion_data = {
                'id': asig.id,
                'empleado_id': asig.empleado_id,
                'empleado_nombre': empleado_nombre,
                'rol_id': asig.rol_id,
                'rol_nombre': rol_nombre,
                'tipo_comision': asig.tipo_comision or 'fijo',
                'valor_comision': float(asig.valor_comision) if asig.valor_comision is not None else 0.0,
                'valor_unitario': valor_unitario
            }
            asignaciones_list.append(asignacion_data)
            
            # Registrar empleados por rol
            if asig.rol_id not in empleados_por_rol:
                empleados_por_rol[asig.rol_id] = []
            empleados_por_rol[asig.rol_id].append(asig.empleado_id)

        # Obtener todos los roles
        roles_db = await db.execute(select(models.Rol).order_by(models.Rol.nombre))
        roles_db_list = list(roles_db.scalars().all())
        
        # Procesar roles para la lista - CON MANEJO SEGURO DE VALORES
        roles_list = []
        for rol in roles_db_list:
            # Nombre seguro del rol
            rol_nombre = f"Rol {rol.id}"
            if hasattr(rol, 'nombre') and rol.nombre and str(rol.nombre).strip():
                rol_nombre = str(rol.nombre).strip()
            
            # Buscar el valor unitario real para este rol
            valor_unitario_rol = 0.0
            tipo_comision_rol = 'fijo'
            
            # Buscar en las asignaciones reales
            for asig in asignaciones:
                if asig.rol_id == rol.id:
                    if asig.valor_unitario is not None:
                        valor_unitario_rol = float(asig.valor_unitario)
                    elif asig.valor_comision is not None:
                        valor_unitario_rol = float(asig.valor_comision)
                    tipo_comision_rol = asig.tipo_comision or 'fijo'
                    break  # Tomar el primer valor encontrado
            
            # Empleados asignados a este rol
            empleados_rol = empleados_por_rol.get(rol.id, [])
            
            rol_data = {
                'id': rol.id,
                'nombre': rol_nombre,
                'tipo_comision': tipo_comision_rol,
                'valor_unitario': valor_unitario_rol,
                'empleados_asignados': empleados_rol
            }
            roles_list.append(rol_data)
        
        # Obtener clientes activos
        clientes_db = await db.execute(
            select(models.Cliente).where(models.Cliente.activo == True)
        )
        clientes = list(clientes_db.scalars().all())
        
        # Obtener empleados activos
        empleados_db = await db.execute(
            select(models.Empleado).where(models.Empleado.activo == True)
        )
        empleados = list(empleados_db.scalars().all())
        
        empleados_list = []
        for emp in empleados:
            rol_empleado = await db.get(models.Rol, emp.rol_id)
            # Nombre del empleado - PRIORIDAD: nombre_completo > nombre
            empleado_nombre = f"Empleado {emp.id}"
            if hasattr(emp, 'nombre_completo') and emp.nombre_completo:
                empleado_nombre = emp.nombre_completo
            elif hasattr(emp, 'nombre') and emp.nombre:
                empleado_nombre = emp.nombre
            
            rol_nombre_emp = "Sin rol"
            if rol_empleado and hasattr(rol_empleado, 'nombre') and rol_empleado.nombre:
                rol_nombre_emp = rol_empleado.nombre
            
            empleado_data = {
                'id': emp.id,
                'nombre': empleado_nombre,
                'rol_id': emp.rol_id,
                'rol_nombre': rol_nombre_emp
            }
            empleados_list.append(empleado_data)
        
        # Obtener materiales
        materiales_db = await db.execute(
            select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == id)
        )
        materiales = list(materiales_db.scalars().all())
        
        # Obtener servicios externos
        servicios_db = await db.execute(
            select(models.ServicioExterno).where(models.ServicioExterno.trabajo_id == id)
        )
        servicios = list(servicios_db.scalars().all())
        
        # Obtener archivos
        archivos_db = await db.execute(
            select(models.ArchivoTrabajo).where(models.ArchivoTrabajo.trabajo_id == id)
        )
        archivos = list(archivos_db.scalars().all())
        
        # Convertir trabajo a dict
        trabajo_dict = {
            'id': trabajo.id,
            'cliente_id': trabajo.cliente_id,
            'nombre_trabajo': trabajo.nombre_trabajo or "",
            'monto_total': float(trabajo.monto_total) if trabajo.monto_total else 0.0,
            'fecha_inicio': trabajo.fecha_inicio,
            'estado': trabajo.estado or "pendiente",
            'metros_cuadrados': float(trabajo.metros_cuadrados) if trabajo.metros_cuadrados else 0.0,
            'unidades': trabajo.unidades or 0,
            'descripcion': trabajo.descripcion or "",
            'prioridad': trabajo.prioridad or "media",
            'monto_pagado_usd': float(trabajo.monto_pagado_usd) if trabajo.monto_pagado_usd else 0.0,
            'monto_pagado_bs': float(trabajo.monto_pagado_bs) if trabajo.monto_pagado_bs else 0.0,
            'porcentaje_pagado': trabajo.porcentaje_pagado or 0,
            'tasa_cambio': float(trabajo.tasa_cambio) if trabajo.tasa_cambio else 400.0,
            'entregado': trabajo.entregado,
            'fecha_entrega': trabajo.fecha_entrega,
            'total_materiales_usd': float(trabajo.total_materiales_usd) if trabajo.total_materiales_usd else 0.0,
            'total_comisiones_usd': float(trabajo.total_comisiones_usd) if trabajo.total_comisiones_usd else 0.0,
            'servicios_externos_usd': float(trabajo.servicios_externos_usd) if trabajo.servicios_externos_usd else 0.0,
            'ganancia_neta_usd': float(trabajo.ganancia_neta_usd) if trabajo.ganancia_neta_usd else 0.0
        }
        
        # Calcular totales
        total_comisiones = float(trabajo.total_comisiones_usd) if trabajo.total_comisiones_usd else 0.0
        total_materiales = float(trabajo.total_materiales_usd) if trabajo.total_materiales_usd else 0.0
        total_servicios = float(trabajo.servicios_externos_usd) if trabajo.servicios_externos_usd else 0.0
        monto_presupuesto = float(trabajo.monto_total) if trabajo.monto_total else 0.0
        costo_total = total_comisiones + total_materiales + total_servicios
        ganancia_neta = monto_presupuesto - costo_total
        monto_pagado = float(trabajo.monto_pagado_usd) if trabajo.monto_pagado_usd else 0.0
        monto_pendiente = monto_presupuesto - monto_pagado

        totales = {
            "monto_total": monto_presupuesto,
            "monto_pendiente": monto_pendiente,
            "total_comisiones": total_comisiones,
            "total_materiales": total_materiales,
            "total_servicios_externos": total_servicios,
            "costo_total": costo_total,
            "ganancia_neta": ganancia_neta
        }
        
        tasa_cambio = trabajo_dict['tasa_cambio']
        
        context = {
            "request": request,
            "current_user": user,
            "trabajo": trabajo_dict,
            "asignaciones": asignaciones_list,
            "roles": roles_list,
            "empleados": empleados_list,
            "clientes": clientes,
            "materiales": materiales,
            "servicios_externos": servicios,
            "archivos": archivos,
            "tasa_cambio": tasa_cambio,
            "hoy": datetime.now().date(),
            "es_edicion": True,
            "totales": totales
        }
        
        return templates.TemplateResponse("trabajos/editar.html", context)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error en editar_trabajo_form: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

@app.post("/editar/{id}", response_class=RedirectResponse)
async def actualizar_trabajo(
    request: Request,
    id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    """Actualizar trabajo existente con automatización de inventario y registro unificado en MovimientoInventario"""
    try:
        print("🔍 ¡RUTA DE ACTUALIZACIÓN EJECUTADA!")
        print(f"ID del trabajo: {id}")
        
        if not user:
            print("❌ Usuario no autenticado")
            return RedirectResponse(url="/login", status_code=303)
        
        form_data = await request.form()
        print(f"📝 Nombre del trabajo recibido: {form_data.get('nombre_trabajo')}")
        print(f"📝 Cliente ID recibido: {form_data.get('cliente_id')}")
        
        # Parsear datos básicos
        cliente_id_raw = form_data.get("cliente_id")
        if cliente_id_raw and cliente_id_raw != "":
            cliente_id = int(cliente_id_raw)
        else:
            # Usar el cliente actual del trabajo
            trabajo_temp = await db.get(models.Trabajo, id)
            cliente_id = trabajo_temp.cliente_id if trabajo_temp else 0

        nombre_trabajo = form_data.get("nombre_trabajo", "").strip()
        monto_total = parse_decimal(form_data.get("monto_total", 0))
        fecha_inicio = form_data.get("fecha_inicio")
        
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        
        if not nombre_trabajo:
            raise HTTPException(status_code=400, detail="Nombre del trabajo es requerido")
        if not cliente_id:
            raise HTTPException(status_code=400, detail="Cliente es requerido")
        
        # Obtener el trabajo existente
        trabajo_db = await db.execute(
            select(models.Trabajo).where(models.Trabajo.id == id)
        )
        trabajo = trabajo_db.scalar_one_or_none()
        
        if not trabajo:
            raise HTTPException(status_code=404, detail="Trabajo no encontrado")
        
        # Actualizar trabajo
        trabajo.cliente_id = cliente_id
        trabajo.nombre_trabajo = nombre_trabajo
        trabajo.monto_total = monto_total
        
        if fecha_inicio:
            trabajo.fecha_inicio = fecha_inicio
        
        trabajo.estado = form_data.get("estado", "pendiente")
        
        # Manejo seguro de campos numéricos
        metros_raw = form_data.get("metros_cuadrados", "0")
        trabajo.metros_cuadrados = parse_decimal(metros_raw)
        
        unidades_raw = form_data.get("unidades", "0")
        trabajo.unidades = int(unidades_raw) if unidades_raw and str(unidades_raw).strip() != "" else 0
        
        trabajo.descripcion = form_data.get("descripcion", "")
        trabajo.prioridad = form_data.get("prioridad", "media")
        trabajo.monto_pagado_usd = parse_decimal(form_data.get("monto_pagado_usd", 0))
        trabajo.monto_pagado_bs = parse_decimal(form_data.get("monto_pagado", 0))
        trabajo.entregado = form_data.get("entregado") == "on"
        
        if trabajo.entregado and not trabajo.fecha_entrega:
            trabajo.fecha_entrega = datetime.now()
        
        # Usar la tasa existente del trabajo (no cambiarla al editar)
        tasa_cambio = trabajo.tasa_cambio
        trabajo.tasa_cambio = tasa_cambio
        trabajo.monto_total_usd = monto_total
        
        # --- GESTIÓN DE ASIGNACIONES ---
        asignaciones_existentes = await db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == id)
        )
        asignaciones_dict = {a.id: a for a in asignaciones_existentes.scalars().all()}
        ids_procesados = set()
        
        roles_db = await db.execute(select(models.Rol))
        roles = list(roles_db.scalars().all())
        empleados_asignados_global = set()
        
        for rol in roles:
            empleados_ids_raw = form_data.getlist(f"empleado_{rol.id}[]")
            empleados_ids = []
            
            for emp_id in empleados_ids_raw:
                if emp_id:
                    emp_id_int = int(emp_id)
                    if emp_id_int in empleados_asignados_global:
                        raise HTTPException(
                            status_code=400, 
                            detail=f"El empleado ID {emp_id_int} está asignado a múltiples roles"
                        )
                    empleados_ids.append(emp_id_int)
                    empleados_asignados_global.add(emp_id_int)
            
            if empleados_ids:
                tipo_comision = form_data.get(f"tipo_comision_{rol.id}", "fijo")
                valor_unitario = parse_decimal(form_data.get(f"valor_comision_{rol.id}", "0"))
                
                if tipo_comision == "porcentaje":
                    total_rol = (trabajo.monto_total * valor_unitario) / 100
                elif tipo_comision == "metro":
                    total_rol = (trabajo.metros_cuadrados or 0) * valor_unitario
                elif tipo_comision == "unidad":
                    total_rol = (trabajo.unidades or 0) * valor_unitario
                else:  # fijo
                    total_rol = valor_unitario
                
                valor_por_empleado = total_rol / len(empleados_ids)
                
                asignaciones_rol_existentes = [
                    a for a in asignaciones_dict.values() 
                    if a.rol_id == rol.id and a.empleado_id in empleados_ids
                ]
                
                for emp_id in empleados_ids:
                    asignacion_existente = next(
                        (a for a in asignaciones_rol_existentes if a.empleado_id == emp_id), 
                        None
                    )
                    
                    if asignacion_existente:
                        asignacion_existente.tipo_comision = tipo_comision
                        asignacion_existente.valor_comision = valor_por_empleado
                        asignacion_existente.valor_unitario = valor_unitario
                        ids_procesados.add(asignacion_existente.id)
                    else:
                        nueva_asignacion = models.Asignacion(
                            trabajo_id=id,
                            empleado_id=emp_id,
                            rol_id=rol.id,
                            tipo_comision=tipo_comision,
                            valor_unitario=valor_unitario,
                            valor_comision=valor_por_empleado
                        )
                        db.add(nueva_asignacion)
        
        # Eliminar asignaciones que ya no están en el formulario
        for asignacion_id, asignacion in asignaciones_dict.items():
            if asignacion_id not in ids_procesados:
                await db.delete(asignacion)
        
        # --- AUTOMATIZACIÓN DE MATERIALES E INVENTARIO ---
        # Paso 1: Obtener materiales usados anteriormente
        materiales_anteriores = await db.execute(
            select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == id)
        )
        materiales_ant_dict = {}
        for mat_ant in materiales_anteriores.scalars().all():
            materiales_ant_dict[mat_ant.concepto] = float(mat_ant.cantidad_usada)
        
        # Paso 2: Eliminar materiales anteriores
        await db.execute(
            delete(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == id)
        )
        
        # Paso 3: Procesar nuevos materiales
        materiales_raw = form_data.getlist("materiales_concepto[]")
        cantidades_raw = form_data.getlist("materiales_cantidad[]")
        costos_raw = form_data.getlist("materiales_costo[]")
        
        nuevos_materiales_dict = {}
        min_len = min(len(materiales_raw), len(cantidades_raw), len(costos_raw))
        
        if min_len > 0:
            for i in range(min_len):
                concepto = materiales_raw[i].strip()
                cantidad_raw = cantidades_raw[i]
                costo_raw = costos_raw[i]
                
                if concepto and cantidad_raw:
                    cantidad = parse_decimal(cantidad_raw)
                    costo_unitario = parse_decimal(costo_raw)
                    if cantidad > 0:
                        material_usado = models.MaterialUsado(
                            trabajo_id=id,
                            concepto=concepto,
                            cantidad_usada=cantidad,
                            costo_unitario=costo_unitario
                        )
                        db.add(material_usado)
                        nuevos_materiales_dict[concepto] = cantidad
                        
                        material_inv = await db.execute(
                            select(models.MaterialInventario).where(models.MaterialInventario.nombre == concepto)
                        )
                        material_inv_obj = material_inv.scalar_one_or_none()
                        
                        if material_inv_obj:
                            cantidad_anterior = materiales_ant_dict.get(concepto, 0)
                            diferencia = cantidad - cantidad_anterior
                            
                            if diferencia > 0:
                                # 🔥 REGISTRAR EN MOVIMIENTOINVENTARIO (UNIFICADO)
                                movimiento_salida = models.MovimientoInventario(
                                    material_id=material_inv_obj.id,
                                    tipo="salida",
                                    cantidad=diferencia,
                                    motivo=f"Edición de trabajo: {trabajo.nombre_trabajo} (ID {id})",
                                    referencia=f"TRABAJO_{id}",
                                    usuario_id=user.id,
                                    trabajo_id=id
                                )
                                db.add(movimiento_salida)
                                material_inv_obj.stock_actual = float(material_inv_obj.stock_actual) - diferencia
                                material_inv_obj.fecha_actualizacion = datetime.utcnow()
                            
                            elif diferencia < 0:
                                # Devolución al inventario
                                material_inv_obj.stock_actual = float(material_inv_obj.stock_actual) + abs(diferencia)
                                material_inv_obj.fecha_actualizacion = datetime.utcnow()

        # Paso 4: Restaurar materiales eliminados
        for concepto_ant, cantidad_ant in materiales_ant_dict.items():
            if concepto_ant not in nuevos_materiales_dict:
                material_inv = await db.execute(
                    select(models.MaterialInventario).where(models.MaterialInventario.nombre == concepto_ant)
                )
                material_inv_obj = material_inv.scalar_one_or_none()
                
                if material_inv_obj:
                    # 🔥 REGISTRAR DEVOLUCIÓN EN MOVIMIENTOINVENTARIO
                    movimiento_devolucion = models.MovimientoInventario(
                        material_id=material_inv_obj.id,
                        tipo="entrada",
                        cantidad=cantidad_ant,
                        motivo=f"Eliminación de material en edición de trabajo ID {id}",
                        referencia=f"TRABAJO_{id}_DEVOLUCION",
                        usuario_id=user.id,
                        trabajo_id=id
                    )
                    db.add(movimiento_devolucion)
                    material_inv_obj.stock_actual = float(material_inv_obj.stock_actual) + cantidad_ant
                    material_inv_obj.fecha_actualizacion = datetime.utcnow()
        
        # --- PROCESAR SERVICIOS EXTERNOS ---
        await db.execute(
            delete(models.ServicioExterno).where(models.ServicioExterno.trabajo_id == id)
        )
        
        s_conceptos = form_data.getlist("servicios_concepto[]")
        s_costos = form_data.getlist("servicios_costo[]")
        min_len_serv = min(len(s_conceptos), len(s_costos))
        
        for i in range(min_len_serv):
            concepto = s_conceptos[i].strip()
            costo = parse_decimal(s_costos[i])
            if concepto and costo > 0:
                db.add(models.ServicioExterno(
                    trabajo_id=id,
                    concepto=concepto,
                    proveedor=None,
                    costo=costo
                ))
        
        # --- PROCESAR ARCHIVOS ---
        archivos_procesados = 0
        for field_name in form_data.keys():
            if field_name.startswith('archivo_'):
                file_item = form_data.get(field_name)
                if isinstance(file_item, UploadFile) and file_item.filename:
                    upload_dir = f"uploads/trabajos/{id}"
                    os.makedirs(upload_dir, exist_ok=True)
                    file_ext = os.path.splitext(file_item.filename)[1]
                    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
                    file_path = os.path.join(upload_dir, unique_filename)
                    with open(file_path, "wb") as buffer:
                        content = await file_item.read()
                        buffer.write(content)
                    archivo_db = models.ArchivoTrabajo(
                        trabajo_id=id,
                        nombre_original=file_item.filename,
                        nombre_guardado=unique_filename,
                        ruta_completa=file_path,
                        tipo_mime=file_item.content_type,
                        tamano_bytes=len(content),
                        fecha_subida=datetime.now()
                    )
                    db.add(archivo_db)
                    archivos_procesados += 1
        
        await db.commit()
        
        # --- RECALCULAR TOTALES ---
        total_com_usd = 0.0
        asignaciones_actualizadas = await db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == id)
        )
        for asig in asignaciones_actualizadas.scalars().all():
            if asig.tipo_comision == "porcentaje":
                monto = (trabajo.monto_total * asig.valor_comision) / 100
            elif asig.tipo_comision == "metro":
                monto = (trabajo.metros_cuadrados or 0) * asig.valor_comision
            elif asig.tipo_comision == "unidad":
                monto = (trabajo.unidades or 0) * asig.valor_comision
            elif asig.tipo_comision == "fijo":
                monto = asig.valor_comision
            else:
                monto = 0.0
            total_com_usd += monto

        total_mat_usd = 0.0
        materiales_actualizados = await db.execute(
            select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == id)
        )
        for mat in materiales_actualizados.scalars().all():
            total_mat_usd += float(mat.cantidad_usada or 0) * float(mat.costo_unitario or 0)

        total_serv_usd = 0.0
        servicios_actualizados = await db.execute(
            select(models.ServicioExterno).where(models.ServicioExterno.trabajo_id == id)
        )
        for serv in servicios_actualizados.scalars().all():
            total_serv_usd += float(serv.costo or 0)

        trabajo.total_comisiones_usd = total_com_usd
        trabajo.total_materiales_usd = total_mat_usd
        trabajo.servicios_externos_usd = total_serv_usd
        gastos_totales_usd = total_com_usd + total_mat_usd + total_serv_usd
        trabajo.ganancia_neta_usd = trabajo.monto_total - gastos_totales_usd
        
        tasa_cambio_float = float(trabajo.tasa_cambio) if trabajo.tasa_cambio else 1.0
        trabajo.total_comisiones_bs = float(total_com_usd) * tasa_cambio_float
        trabajo.total_materiales_bs = float(total_mat_usd) * tasa_cambio_float
        trabajo.servicios_externos_bs = float(total_serv_usd) * tasa_cambio_float
        trabajo.ganancia_neta_bs = float(trabajo.ganancia_neta_usd) * tasa_cambio_float
        
        await db.commit()
        
        return RedirectResponse(
            url=f"/trabajos?success_edit=1",
            status_code=303
        )
        
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        print(f"Error en actualizar_trabajo: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al actualizar trabajo: {str(e)}")
    
@app.get("/trabajos/ver/{trabajo_id}")
async def ver_detalle_trabajo(
    request: Request,
    trabajo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    print(f"🔍 ID del trabajo solicitado: {trabajo_id}")
    
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        print("🔍 Paso 1: Obteniendo trabajo...")
        trabajo_query = (
            select(models.Trabajo, models.Cliente.nombre_razon_social.label("cliente_nombre"))
            .join(models.Cliente, models.Trabajo.cliente_id == models.Cliente.id)
            .where(models.Trabajo.id == trabajo_id)
        )
        resultado = await db.execute(trabajo_query)
        row = resultado.fetchone()
        
        if not row:
            print(f"❌ Trabajo {trabajo_id} no encontrado")
            return RedirectResponse(url="/trabajos?error=Trabajo+no+encontrado", status_code=303)
        
        trabajo, cliente_nombre = row[0], row.cliente_nombre
        print(f"✅ Trabajo encontrado: {trabajo.nombre_trabajo}")
        
        # 2. Asignaciones (Comisiones)
        print("🔍 Paso 2: Obteniendo asignaciones...")
        asignaciones_result = await db.execute(
            select(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo_id)
        )
        asignaciones = asignaciones_result.scalars().all()
        asignaciones_detalles = []
        total_comisiones = 0.0
        
        for asig in asignaciones:
            empleado = await db.get(models.Empleado, asig.empleado_id)
            rol = await db.get(models.Rol, asig.rol_id)

            monto_total_empleado = float(asig.valor_comision) if asig.valor_comision else 0.0
            
            total_comisiones += monto_total_empleado
            asignaciones_detalles.append({
                "empleado_nombre": empleado.nombre_completo if empleado else f"ID {asig.empleado_id}",
                "rol_nombre": rol.nombre.title() if rol else "N/A",
                "tipo_comision": asig.tipo_comision,
                "valor_comision": monto_total_empleado,  
                "valor_unitario": float(asig.valor_unitario) if hasattr(asig, 'valor_unitario') and asig.valor_unitario else 0.0  
            })
        
        print(f"✅ Asignaciones: {len(asignaciones_detalles)}")

        # 3. Materiales de Rotulado (MaterialUsado)
        print("🔍 Paso 3: Obteniendo materiales de rotulado...")
        materiales_detalles = []
        total_materiales = 0.0
        materiales_result = await db.execute(
            select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == trabajo_id)
        )
        
        for m in materiales_result.scalars().all():
            c_unit = float(m.costo_unitario or 0)
            cant = float(m.cantidad_usada or 0)
            subtotal = c_unit * cant
            total_materiales += subtotal
            materiales_detalles.append({
                "concepto": m.concepto or "Material sin nombre",
                "cantidad": cant,
                "costo_unitario": c_unit,
                "costo_total": subtotal,
                "unidad": "m²"  # Unidad por defecto para rotulado
            })
        
        print(f"✅ Materiales rotulado: {len(materiales_detalles)}")
        
        # ============================================================
        # 3B. MATERIALES TEXTIL
        # ============================================================
        materiales_textil = []
        if trabajo.tipo_trabajo == 'textil':
            print("🔍 Paso 3B: Obteniendo materiales textil...")
            try:
                textil_result = await db.execute(
                    select(models.TrabajoMaterialTextil).where(models.TrabajoMaterialTextil.trabajo_id == trabajo_id)
                )
                materiales_textil = textil_result.scalars().all()
                print(f"✅ Materiales textil encontrados: {len(materiales_textil)}")
                for mt in materiales_textil:
                    print(f"   - Prenda: {mt.prenda}, Tela: {mt.tela_nombre}, m²: {mt.m2_por_prenda if hasattr(mt, 'm2_por_prenda') else 'N/A'}")
            except Exception as e:
                print(f"⚠️ Error al obtener materiales textil: {e}")
                materiales_textil = []
        
        # ============================================================
        # PAPEL DE SUBLIMACIÓN (TEXTIL) - BUSCAR EN EL TRABAJO
        # ============================================================
        textil_papel = None
        if trabajo.tipo_trabajo == 'textil':
            # Buscar el papel en el trabajo (donde lo guardamos en el POST)
            if hasattr(trabajo, 'papel_sublimacion_nombre') and trabajo.papel_sublimacion_nombre:
                textil_papel = {
                    "nombre": trabajo.papel_sublimacion_nombre,
                    "cantidad": trabajo.papel_sublimacion_cantidad
                }
                print(f"✅ Papel de sublimación encontrado en trabajo: {textil_papel}")
            else:
                print(f"⚠️ No se encontró papel de sublimación para el trabajo {trabajo_id}")

        # 4. Servicios Externos
        print("🔍 Paso 4: Procesando servicios externos...")
        servicios_externos_result = await db.execute(
            select(models.ServicioExterno).where(models.ServicioExterno.trabajo_id == trabajo_id)
        )
        servicios_externos = [
            {
                'concepto': s.concepto,
                'costo': float(s.costo or 0)
            }
            for s in servicios_externos_result.scalars().all()
        ]
        total_servicios = sum(s['costo'] for s in servicios_externos)

        # 5. Archivos
        print("🔍 Paso 5: Obteniendo archivos...")
        try:
            archivos_query = await db.execute(
                select(models.ArchivoTrabajo)
                .where(models.ArchivoTrabajo.trabajo_id == trabajo_id)
                .order_by(models.ArchivoTrabajo.fecha_subida.desc())
            )
            archivos = archivos_query.scalars().all()
            print(f"✅ Archivos encontrados: {len(archivos)}")
        except Exception as e:
            print(f"❌ Error obteniendo archivos: {e}")
            archivos = []

        # 6. Totales
        print("🔍 Paso 6: Calculando totales...")
        monto_presupuesto = float(trabajo.monto_total or 0)
        costo_total = total_comisiones + total_materiales + total_servicios
        monto_pagado = float(trabajo.monto_pagado_usd or 0.0)
        monto_pendiente = monto_presupuesto - monto_pagado
        
        print(f"✅ Cálculos completados")
        print(f"  Monto total: {monto_presupuesto}")
        print(f"  Total comisiones: {total_comisiones}")
        print(f"  Total materiales: {total_materiales}")
        print(f"  Total servicios: {total_servicios}")
        
        # Obtener lista de empleados únicos para mostrar
        empleados_lista = list(set([a['empleado_nombre'] for a in asignaciones_detalles]))
        
        # 🔥 OBTENER LA TASA ACTIVA
        try:
            from app.services.currency_service import CurrencyService
            currency_service = CurrencyService(db)
            tasa_actual = await currency_service.get_tasa()
            tasa_actual = float(tasa_actual) if tasa_actual else float(trabajo.tasa_cambio_actual or 40.0)
        except:
            tasa_actual = float(trabajo.tasa_cambio_actual or 40.0)

        context = {
            "request": request,
            "trabajo": trabajo,
            "cliente_nombre": cliente_nombre,
            "asignaciones": asignaciones_detalles,
            "materiales": materiales_detalles,
            "materiales_textil": materiales_textil,
            "textil_papel": textil_papel,
            "servicios_externos": servicios_externos, 
            "archivos": archivos,
            "empleados": empleados_lista,
            "hoy": date.today(),
            "tasa_actual": tasa_actual,  # 🔥 PASAR LA TASA COMO VARIABLE INDEPENDIENTE
            "totales": {
                "monto_total": monto_presupuesto,
                "monto_pendiente": monto_pendiente,
                "total_comisiones": total_comisiones,
                "total_materiales": total_materiales,
                "total_servicios_externos": total_servicios,
                "costo_total": costo_total,
                "ganancia_neta": monto_presupuesto - costo_total
            }
        }
        
        return templates.TemplateResponse("trabajos/ver.html", context)
        
    except Exception as e:
        print(f"❌ ERROR EN VER TRABAJO {trabajo_id}: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/trabajos?error=Error+al+cargar+detalle", status_code=303)
    
@app.get("/trabajos", response_class=HTMLResponse)
async def listar_trabajos(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db),
    fecha_filtro: Optional[str] = Query(None),
    estado_filtro: Optional[str] = Query(None)
):
    """
    Muestra la lista de todos los trabajos con filtros
    Por defecto, muestra trabajos de la fecha actual
    """
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # 1. Si no hay fecha_filtro, usar la fecha actual
        if fecha_filtro is None:
            fecha_actual = datetime.now().date()
            fecha_filtro = fecha_actual.strftime("%Y-%m-%d")
        else:
            try:
                fecha_actual = datetime.strptime(fecha_filtro, "%Y-%m-%d").date()
            except ValueError:
                # Si la fecha es inválida, usar hoy
                fecha_actual = datetime.now().date()
                fecha_filtro = fecha_actual.strftime("%Y-%m-%d")
        
        # 2. Construir query base con join a cliente
        query = (
            select(
                models.Trabajo,
                models.Cliente.nombre_razon_social.label("cliente_nombre")
            )
            .join(models.Cliente, models.Trabajo.cliente_id == models.Cliente.id)
        )
        
        # 3. Aplicar filtro de fecha (siempre filtrar por fecha cuando se especifica)
        # Si hay fecha_filtro, filtrar por esa fecha
        if fecha_filtro:
            query = query.where(models.Trabajo.fecha_inicio == fecha_actual)
        
        # 4. Aplicar filtro de estado si se especifica
        if estado_filtro:
            query = query.where(models.Trabajo.estado == estado_filtro)
        
        # 5. Ordenar por prioridad y luego por fecha
        # Mapeo de prioridades para ordenar correctamente
        prioridad_order = {
            "alta": 1,
            "media": 2,
            "baja": 3
        }
        
        # Usar SQLAlchemy case para ordenar por prioridad
        from sqlalchemy import case
        prioridad_case = case(
            (models.Trabajo.prioridad == "alta", 1),
            (models.Trabajo.prioridad == "media", 2),
            (models.Trabajo.prioridad == "baja", 3),
            else_=4
        )
        
        query = query.order_by(
            prioridad_case,
            models.Trabajo.fecha_inicio.desc(),
            models.Trabajo.id.desc()
        )
        
        # 6. Ejecutar query
        result = await db.execute(query)
        rows = result.fetchall()
        
        # 7. Procesar cada trabajo con cálculos detallados
        trabajos_data = []
        for trabajo, cliente_nombre in rows:
            # === CALCULAR COMISIONES ===
            asignaciones_result = await db.execute(
                select(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo.id)
            )
            asignaciones = asignaciones_result.scalars().all()
            total_comisiones = sum(float(a.valor_comision or 0) for a in asignaciones)
            
            # === CALCULAR MATERIALES ===
            materiales_result = await db.execute(
                select(models.MaterialUsado).where(models.MaterialUsado.trabajo_id == trabajo.id)
            )
            materiales = materiales_result.scalars().all()
            total_materiales = sum(
                float(m.cantidad_usada or 0) * float(m.costo_unitario or 0) 
                for m in materiales
            )
            
            # === CALCULAR SERVICIOS EXTERNOS ===
            total_servicios_externos = float(getattr(trabajo, 'servicios_externos_usd', 0) or 0)
            
            # === CALCULAR TOTALES ===
            monto_total = float(trabajo.monto_total or 0)
            costo_total = total_comisiones + total_materiales + total_servicios_externos
            ganancia_neta = monto_total - costo_total
            
            # === DETERMINAR SI REQUIERE ATENCIÓN ===
            # En ejecución pero sin pago o con menos del 50% de pago
            requiere_atencion = (
                trabajo.estado == "en_ejecucion" and 
                (trabajo.porcentaje_pagado or 0) < 50
            )
            
            # === FORMATEAR DATOS PARA EL TEMPLATE ===
            trabajo_dict = {
                "id": trabajo.id,
                "nombre_trabajo": trabajo.nombre_trabajo or "Sin nombre",
                "cliente_nombre": cliente_nombre or "Sin cliente",
                "porcentaje_pagado": trabajo.porcentaje_pagado or 0,
                "monto_total": monto_total,
                "total_materiales": total_materiales,
                "total_servicios_externos": total_servicios_externos,
                "total_comisiones": total_comisiones,
                "costo_total": costo_total,
                "ganancia_neta": ganancia_neta,
                "estado": trabajo.estado or "pendiente",
                "requiere_atencion": requiere_atencion,
                "fecha_inicio": trabajo.fecha_inicio,
                "prioridad": trabajo.prioridad or "media",
                "prioridad_color": {
                    "alta": "text-red-600",
                    "media": "text-amber-600", 
                    "baja": "text-green-600"
                }.get(trabajo.prioridad or "media", "text-gray-600")
            }
            
            trabajos_data.append(trabajo_dict)
        
        # 8. Calcular estadísticas para el día
        trabajos_hoy = len(trabajos_data)
        facturado_hoy = sum(t["monto_total"] for t in trabajos_data)
        ganancia_hoy = sum(t["ganancia_neta"] for t in trabajos_data if t["ganancia_neta"] > 0)
        
        trabajos_pendientes = sum(1 for t in trabajos_data if t["estado"] in ["pendiente", "en_ejecucion"])
        trabajos_finalizados = sum(1 for t in trabajos_data if t["estado"] == "finalizado")
        
        # 9. Obtener trabajos de otros días para estadísticas comparativas
        query_todos = (
            select(models.Trabajo)
            .order_by(models.Trabajo.fecha_inicio.desc())
            .limit(100)  # Últimos 100 trabajos para estadísticas
        )
        result_todos = await db.execute(query_todos)
        todos_trabajos = result_todos.scalars().all()
        
        # Calcular estadísticas generales
        total_general = len(todos_trabajos)
        facturado_total = sum(float(t.monto_total or 0) for t in todos_trabajos)
        
        # 10. Renderizar template con los datos
        return templates.TemplateResponse("trabajos/listar.html", {
            "request": request,
            "trabajos": trabajos_data,
            "fecha_filtro": fecha_filtro,  # Fecha actual formateada
            "estado_filtro": estado_filtro,
            "total_trabajos": trabajos_hoy,
            "total_facturado": facturado_hoy,
            "total_ganancia": ganancia_hoy,
            "trabajos_pendientes": trabajos_pendientes,
            "trabajos_finalizados": trabajos_finalizados,
            "estadisticas_generales": {
                "total_trabajos": total_general,
                "facturado_total": facturado_total,
                "fecha_actual": fecha_actual.strftime("%d/%m/%Y")
            },
            "hoy_str": fecha_actual.strftime("%A, %d de %B").replace(
                "Monday", "Lunes").replace("Tuesday", "Martes").replace(
                "Wednesday", "Miércoles").replace("Thursday", "Jueves").replace(
                "Friday", "Viernes").replace("Saturday", "Sábado").replace(
                "Sunday", "Domingo").replace("January", "Enero").replace(
                "February", "Febrero").replace("March", "Marzo").replace(
                "April", "Abril").replace("May", "Mayo").replace(
                "June", "Junio").replace("July", "Julio").replace(
                "August", "Agosto").replace("September", "Septiembre").replace(
                "October", "Octubre").replace("November", "Noviembre").replace(
                "December", "Diciembre")
        })
        
    except Exception as e:
        print(f"❌ Error al listar trabajos: {e}")
        import traceback
        traceback.print_exc()
        
        # En caso de error, mostrar página con fecha actual
        fecha_actual = datetime.now().date()
        
        return templates.TemplateResponse("trabajos/listar.html", {
            "request": request,
            "trabajos": [],
            "fecha_filtro": fecha_actual.strftime("%Y-%m-%d"),
            "estado_filtro": estado_filtro,
            "total_trabajos": 0,
            "error": f"Error al cargar los trabajos: {str(e)[:50]}",
            "hoy_str": fecha_actual.strftime("%A, %d de %B")
        })

@app.post("/trabajos/{trabajo_id}/subir-archivos")
async def subir_archivos_trabajo(
    request: Request,
    trabajo_id: int,
    archivos: List[UploadFile] = File(...),
    descripcion: str = Form(None),
    user = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    """
    Agrega archivos a un trabajo existente
    """
    print(f"📤 INICIANDO: Subiendo archivos para trabajo {trabajo_id}")
    print(f"📤 Usuario: {user.id if user else 'No autenticado'}")
    print(f"📤 Archivos recibidos: {len(archivos)}")
    print(f"📤 Descripción: {descripcion}")
    
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    
    try:
        # 1. VERIFICAR QUE EL TRABAJO EXISTE
        print(f"🔍 Verificando existencia del trabajo {trabajo_id}...")
        trabajo_query = await db.execute(
            select(models.Trabajo).where(models.Trabajo.id == trabajo_id)
        )
        trabajo = trabajo_query.scalar_one_or_none()
        
        if not trabajo:
            print(f"❌ Trabajo {trabajo_id} no encontrado")
            raise HTTPException(status_code=404, detail="Trabajo no encontrado")
        
        print(f"✅ Trabajo encontrado: {trabajo.nombre_trabajo}")
        
        # 2. PROCESAR CADA ARCHIVO
        archivos_subidos = []
        errores = []
        
        for i, archivo in enumerate(archivos):
            if not archivo.filename or archivo.filename == "":
                print(f"⚠️  Archivo {i+1}: Sin nombre, omitiendo...")
                continue
                
            try:
                print(f"📄 Procesando archivo {i+1}: {archivo.filename}")
                
                # Generar nombre único para el archivo
                nombre_original = archivo.filename
                extension = os.path.splitext(nombre_original)[1]
                
                # Si no tiene extensión, usar .bin
                if not extension:
                    extension = ".bin"
                    print(f"⚠️  Archivo sin extensión, usando .bin")
                
                # Generar nombre único con UUID
                nombre_guardado = f"{uuid.uuid4().hex}{extension}"
                print(f"📝 Nuevo nombre: {nombre_guardado}")
                
                # Crear carpeta para el trabajo si no existe
                folder = f"uploads/trabajos/{trabajo_id}"
                os.makedirs(folder, exist_ok=True)
                print(f"📁 Carpeta: {folder}")
                
                # Ruta completa del archivo
                filepath = os.path.join(folder, nombre_guardado)
                print(f"📍 Ruta completa: {filepath}")
                
                # Leer y guardar el archivo en disco
                contenido = await archivo.read()
                print(f"💾 Tamaño del archivo: {len(contenido)} bytes")
                
                with open(filepath, "wb") as f:
                    f.write(contenido)
                
                print(f"✅ Archivo guardado físicamente")
                
                # Crear registro en la base de datos
                # Usar la descripción solo para el primer archivo si hay múltiples
                descripcion_archivo = descripcion if i == 0 and descripcion else None
                
                nuevo_archivo = models.ArchivoTrabajo(
                    trabajo_id=trabajo_id,
                    nombre_original=nombre_original,
                    nombre_guardado=nombre_guardado,
                    ruta_completa=filepath,
                    descripcion=descripcion_archivo,  # Asegúrate de que esta columna existe
                    tipo_mime=archivo.content_type or "application/octet-stream",
                    tamano_bytes=len(contenido),
                    fecha_subida=datetime.now()
                )
                
                db.add(nuevo_archivo)
                archivos_subidos.append({
                    "nombre_original": nombre_original,
                    "nombre_guardado": nombre_guardado,
                    "tamaño": len(contenido)
                })
                
                print(f"✅ Registro creado en BD para {nombre_original}")
                
            except Exception as e:
                error_msg = f"Error con archivo {archivo.filename}: {str(e)}"
                print(f"❌ {error_msg}")
                errores.append(error_msg)
                import traceback
                traceback.print_exc()
                continue
        
        # 3. GUARDAR CAMBIOS EN LA BASE DE DATOS
        if archivos_subidos:
            print(f"💾 Guardando {len(archivos_subidos)} archivo(s) en la base de datos...")
            await db.commit()
            print(f"✅ {len(archivos_subidos)} archivo(s) guardado(s) exitosamente")
            
            # Redirigir de vuelta a la página del trabajo con mensaje de éxito
            mensaje_exito = f"Se subieron {len(archivos_subidos)} archivo(s) correctamente"
            print(f"🔗 Redirigiendo a: /trabajos/ver/{trabajo_id}?success={quote(mensaje_exito)}")
            
            return RedirectResponse(
                url=f"/trabajos/ver/{trabajo_id}?success={quote(mensaje_exito)}",
                status_code=303
            )
        else:
            print(f"❌ No se pudo subir ningún archivo")
            mensaje_error = "No se pudo subir ningún archivo"
            if errores:
                mensaje_error += f". Errores: {', '.join(errores[:3])}"
            
            # Redirigir con mensaje de error
            return RedirectResponse(
                url=f"/trabajos/ver/{trabajo_id}?error={quote(mensaje_error)}",
                status_code=303
            )
            
    except HTTPException as he:
        print(f"❌ HTTPException: {he.detail}")
        raise he
        
    except Exception as e:
        print(f"❌ ERROR GENERAL inesperado: {e}")
        import traceback
        traceback.print_exc()
        
        # Redirigir con mensaje de error genérico
        return RedirectResponse(
            url=f"/trabajos/ver/{trabajo_id}?error={quote(f'Error interno: {str(e)}')}",
            status_code=303
        )
    
@app.post("/trabajos/{trabajo_id}/registrar-pago")
async def registrar_pago(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    monto_usd = float(form_data.get("monto_usd", 0))
    tasa_actual = float(form_data.get("tasa_actual", 36.5))
    metodo_pago = form_data.get("metodo_pago", "efectivo_usd")
    fecha_pago_str = form_data.get("fecha_pago")
    
    if monto_usd <= 0:
        return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}?error=Monto+inválido", status_code=303)
    
    trabajo = await db.get(models.Trabajo, trabajo_id)
    if not trabajo:
        return RedirectResponse(url=f"/trabajos?error=Trabajo+no+encontrado", status_code=303)
    
    # ✅ CONVERTIR TODOS LOS VALORES A FLOAT
    monto_total_usd = float(trabajo.monto_total or 0)
    monto_pagado_actual_usd = float(trabajo.monto_pagado_usd or 0)
    monto_pagado_actual_bs = float(trabajo.monto_pagado_bs or 0)
    porcentaje_actual = trabajo.porcentaje_pagado or 0
    
    if monto_total_usd <= 0:
        return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}?error=Monto+total+inválido", status_code=303)
    
    # Calcular nuevo monto total pagado
    nuevo_monto_pagado_usd = monto_pagado_actual_usd + monto_usd
    nuevo_monto_pagado_bs = monto_pagado_actual_bs + (monto_usd * tasa_actual)
    
    # Determinar el nuevo porcentaje (0%, 50%, 100%)
    nuevo_porcentaje = 0
    if nuevo_monto_pagado_usd >= monto_total_usd:
        nuevo_porcentaje = 100
    elif nuevo_monto_pagado_usd >= (monto_total_usd * 0.5):
        nuevo_porcentaje = 50
    
    # Validaciones
    if porcentaje_actual >= 100:
        return RedirectResponse(
            url=f"/trabajos/ver/{trabajo_id}?error=Este+trabajo+ya+está+100%25+pagado", 
            status_code=303
        )
    
    if nuevo_monto_pagado_usd > monto_total_usd * 1.01:  # Margen del 1%
        return RedirectResponse(
            url=f"/trabajos/ver/{trabajo_id}?error=No+se+puede+pagar+más+del+total.+Máximo:+${monto_total_usd:.2f}+USD", 
            status_code=303
        )
    
    # ✅ ACTUALIZAR CON FLOAT (NO Decimal)
    trabajo.monto_pagado_usd = nuevo_monto_pagado_usd
    trabajo.monto_pagado_bs = nuevo_monto_pagado_bs
    trabajo.porcentaje_pagado = nuevo_porcentaje
    trabajo.metodo_pago = metodo_pago
    
    if fecha_pago_str:
        trabajo.fecha_pago = datetime.fromisoformat(fecha_pago_str)
    else:
        trabajo.fecha_pago = datetime.now()
    
    trabajo.tasa_cambio_actual = tasa_actual
    
    # Lógica de estados
    mensaje_exito = None
    
    if nuevo_porcentaje == 100:
        trabajo.estado = "completado"
        trabajo.fecha_finalizacion = datetime.now()
        
        if porcentaje_actual == 0:
            mensaje_exito = "🎉 ¡PAGO COMPLETO 100% EN UNA SOLA TRANSACCIÓN! Estado: COMPLETADO"
        elif porcentaje_actual == 50:
            mensaje_exito = "🎉 ¡SEGUNDA MITAD PAGADA! PAGO COMPLETO 100% - Estado: COMPLETADO"
            
    elif nuevo_porcentaje == 50:
        if porcentaje_actual == 0:
            mensaje_exito = "✅ ¡ABONO DEL 50% REGISTRADO!"
            if trabajo.estado == "pendiente":
                trabajo.estado = "en_ejecucion"
                mensaje_exito += " Trabajo INICIADO (en ejecución)"
        else:
            mensaje_exito = f"💰 Abono registrado. Total acumulado: ${nuevo_monto_pagado_usd:.2f} USD"
            
    else:
        mensaje_exito = f"💰 Abono parcial registrado. Total: ${nuevo_monto_pagado_usd:.2f} USD"
        if monto_pagado_actual_usd == 0 and monto_usd > 0 and trabajo.estado == "pendiente":
            trabajo.estado = "en_ejecucion"
            mensaje_exito += " Trabajo INICIADO (en ejecución)"
    
    await db.commit()
    
    return RedirectResponse(
        url=f"/trabajos/ver/{trabajo_id}?success={mensaje_exito}", 
        status_code=303
    )

@app.post("/trabajos/editar/{trabajo_id}", response_class=HTMLResponse)
async def actualizar_trabajo(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    # Cargar datos necesarios para el manejo de errores
    roles = (await db.execute(select(models.Rol).order_by(models.Rol.nombre))).scalars().all()
    empleados = (await db.execute(
        select(models.Empleado).where(models.Empleado.activo == True)
    )).scalars().all()
    
    form = await request.form()
    nombre = form.get("nombre_trabajo", "").strip()
    monto_total = form.get("monto_total")
    fecha_inicio_str = form.get("fecha_inicio")
    estado = form.get("estado", "pendiente")
    
    # Validación básica
    if not nombre or not monto_total or not fecha_inicio_str:
        # Recargar trabajos con cálculo de comisiones
        trabajos = (await db.execute(
            select(models.Trabajo).order_by(models.Trabajo.fecha_inicio.desc())
        )).scalars().all()
        
        trabajos_data = []
        for t in trabajos:
            calc = await calcular_comisiones_trabajo(t, async_db=db)
            # Obtener roles asignados
            asignaciones_trabajo = (await db.execute(
                select(models.Asignacion).where(models.Asignacion.trabajo_id == t.id)
            )).scalars().all()
            
            rol_ids = list(set(a.rol_id for a in asignaciones_trabajo if a.rol_id))
            nombres_roles = []
            for rol_id in rol_ids:
                rol_nombre = next((r.nombre for r in roles if r.id == rol_id), f"Rol {rol_id}")
                nombres_roles.append(rol_nombre.title())
            
            trabajos_data.append({
                "id": t.id,
                "nombre": t.nombre,
                "monto_total": float(t.monto_total),
                "total_comisiones": calc["total_comisiones"],
                "ingreso_neto": round(float(t.monto_total) - calc["total_comisiones"], 2),
                "fecha_inicio": t.fecha_inicio.isoformat(),
                "estado": t.estado,
                "roles": ", ".join(nombres_roles) if nombres_roles else "Sin roles"
            })
        
        roles_data = [{"id": r.id, "nombre": r.nombre} for r in roles]
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo, "rol_id": e.rol_id} for e in empleados]
        
        return templates.TemplateResponse("trabajos/trabajos.html", {
            "request": request,
            "trabajos": trabajos_data,
            "roles": roles_data,
            "empleados": empleados_data,
            "error": "Nombre, monto y fecha son requeridos"
        })
    
    try:
        trabajo = await db.get(models.Trabajo, trabajo_id)
        if not trabajo:
            return RedirectResponse(url="/trabajos")
        
        # ✅ Solo actualizar campos básicos
        # ❌ NO modificar metros_cuadrados, unidades ni asignaciones
        trabajo.nombre = nombre
        trabajo.monto_total = float(monto_total)
        trabajo.fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        trabajo.estado = estado
        
        await db.commit()
        
        # ✅ Redirección con mensaje de éxito
        return RedirectResponse(url="/trabajos?success_edit=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"ERROR al editar trabajo {trabajo_id}: {str(e)}")
        
        # Manejo de error: recargar datos y mostrar formulario con error
        trabajo = await db.get(models.Trabajo, trabajo_id)
        if not trabajo:
            return RedirectResponse(url="/trabajos")
        
        # Recargar trabajos con cálculo de comisiones
        trabajos = (await db.execute(
            select(models.Trabajo).order_by(models.Trabajo.fecha_inicio.desc())
        )).scalars().all()
        
        trabajos_data = []
        for t in trabajos:
            calc = await calcular_comisiones_trabajo(t, async_db=db)
            # Obtener roles asignados
            asignaciones_trabajo = (await db.execute(
                select(models.Asignacion).where(models.Asignacion.trabajo_id == t.id)
            )).scalars().all()
            
            rol_ids = list(set(a.rol_id for a in asignaciones_trabajo if a.rol_id))
            nombres_roles = []
            for rol_id in rol_ids:
                rol_nombre = next((r.nombre for r in roles if r.id == rol_id), f"Rol {rol_id}")
                nombres_roles.append(rol_nombre.title())
            
            trabajos_data.append({
                "id": t.id,
                "nombre": t.nombre,
                "monto_total": float(t.monto_total),
                "total_comisiones": calc["total_comisiones"],
                "ingreso_neto": round(float(t.monto_total) - calc["total_comisiones"], 2),
                "fecha_inicio": t.fecha_inicio.isoformat(),
                "estado": t.estado,
                "roles": ", ".join(nombres_roles) if nombres_roles else "Sin roles"
            })
        
        roles_data = [{"id": r.id, "nombre": r.nombre} for r in roles]
        empleados_data = [{"id": e.id, "nombre": e.nombre_completo, "rol_id": e.rol_id} for e in empleados]
        
        return templates.TemplateResponse("trabajos/trabajos.html", {
            "request": request,
            "trabajos": trabajos_data,
            "roles": roles_data,
            "empleados": empleados_data,
            "error": "Error al actualizar el trabajo"
        })
        
@app.post("/trabajos/eliminar/{trabajo_id}")
async def eliminar_trabajo(
    trabajo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener el trabajo para verificar estado
    trabajo = await db.get(models.Trabajo, trabajo_id)
    if not trabajo:
        return RedirectResponse(url="/trabajos", status_code=303)
    
    # 🔒 No permitir eliminar trabajos finalizados
    if trabajo.estado == "finalizado":
        return RedirectResponse(url="/trabajos", status_code=303)
    
    # Eliminar asignaciones relacionadas
    await db.execute(
        delete(models.Asignacion).where(models.Asignacion.trabajo_id == trabajo_id)
    )
    
    # Eliminar el trabajo
    await db.delete(trabajo)
    await db.commit()

    return RedirectResponse(url="/trabajos?mensaje=eliminado", status_code=303)
    
    return RedirectResponse(url="/trabajos", status_code=303)

# ======================
# LISTAR CLIENTES
# ======================
@app.get("/clientes", response_class=HTMLResponse)
async def listar_clientes(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db),
    q: str = None,
    tipo: str = None
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Lógica para obtener clientes
    query = select(Cliente)
    
    if q:
        query = query.where(
            (Cliente.nombre_razon_social.ilike(f"%{q}%")) |
            (Cliente.cedula.ilike(f"%{q}%")) |
            (Cliente.rif.ilike(f"%{q}%"))
        )
    
    if tipo:
        query = query.where(Cliente.tipo_cliente == tipo)
    
    query = query.order_by(Cliente.nombre_razon_social)
    result = await db.execute(query)
    clientes = result.scalars().all()
    
    # 🔥 Cargar empleados vinculados
    for cliente in clientes:
        if cliente.empleado_id:
            emp_result = await db.execute(
                select(Empleado).where(Empleado.id == cliente.empleado_id)
            )
            cliente.empleado = emp_result.scalar_one_or_none()
    
    return templates.TemplateResponse("clientes/listar.html", {
        "request": request,
        "user": user,
        "clientes": clientes,
        "busqueda": q,
        "tipo_filtro": tipo
    })

# ==================== GET - Mostrar formulario ====================
@app.get("/clientes/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_cliente(
    request: Request,
    error: Optional[str] = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Cargar empleados activos para el select
    result = await db.execute(
        select(Empleado).where(Empleado.activo == True).order_by(Empleado.nombre_completo)
    )
    empleados = result.scalars().all()
    
    return templates.TemplateResponse("clientes/nuevo.html", {
        "request": request,
        "user": user,
        "empleados": empleados,
        "error": error
    })


# ==================== POST - Guardar cliente ====================
@app.post("/clientes/nuevo", response_class=HTMLResponse)
async def crear_cliente(
    request: Request,
    tipo_cliente: str = Form(...),
    nombre_razon_social: str = Form(...),
    telefono: str = Form(...),
    email: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    notas: Optional[str] = Form(None),
    cedula: Optional[str] = Form(None),
    rif: Optional[str] = Form(None),
    representante_legal: Optional[str] = Form(None),
    empleado_id: Optional[int] = Form(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    errores = []
    
    # Validaciones básicas
    if not nombre_razon_social or len(nombre_razon_social.strip()) < 2:
        errores.append("Nombre/Razón social es obligatorio")
    
    if not telefono:
        errores.append("Teléfono es obligatorio")
    
    if tipo_cliente == "natural":
        if not cedula:
            errores.append("Cédula es obligatoria")
    else:
        if not rif:
            errores.append("RIF es obligatorio")
        if not representante_legal:
            errores.append("Representante legal es obligatorio")
    
    # Validar empleado
    empleado_vinculado = None
    if empleado_id:
        result = await db.execute(select(Empleado).where(Empleado.id == empleado_id))
        empleado_vinculado = result.scalar_one_or_none()
        if not empleado_vinculado:
            errores.append("El empleado seleccionado no existe")
    
    if errores:
        error_msg = " • ".join(errores)
        return RedirectResponse(url=f"/clientes/nuevo?error={error_msg}", status_code=303)
    
    try:
        # 🔥 VALIDACIÓN DE DUPLICADOS CORREGIDA
        if tipo_cliente == "natural" and cedula:
            result = await db.execute(
                select(Cliente).where(
                    Cliente.cedula == cedula.strip(),
                    Cliente.activo == True
                )
            )
            existing = result.scalars().first()  # ✅ CORREGIDO
            if existing:
                return RedirectResponse(url="/clientes/nuevo?error=Ya existe un cliente con esta cédula", status_code=303)
        
        elif tipo_cliente == "juridico" and rif:
            result = await db.execute(
                select(Cliente).where(
                    Cliente.rif == rif.strip(),
                    Cliente.activo == True
                )
            )
            existing = result.scalars().first()  # ✅ CORREGIDO
            if existing:
                return RedirectResponse(url="/clientes/nuevo?error=Ya existe un cliente con este RIF", status_code=303)
        
        # Crear cliente
        nuevo_cliente = Cliente(
            tipo_cliente=tipo_cliente,
            nombre_razon_social=nombre_razon_social.strip(),
            telefono=telefono,
            email=email,
            direccion=direccion,
            notas=notas,
            cedula=cedula.strip() if cedula else None,
            rif=rif.strip() if rif else None,
            representante_legal=representante_legal,
            empleado_id=empleado_id if empleado_vinculado else None,
            activo=True
        )
        
        db.add(nuevo_cliente)
        await db.commit()
        await db.refresh(nuevo_cliente)
        
        return RedirectResponse(url="/clientes?success=1", status_code=303)
        
    except Exception as e:
        await db.rollback()
        print(f"❌ Error: {e}")
        return RedirectResponse(url=f"/clientes/nuevo?error=Error%20interno%3A%20{str(e)}", status_code=303)
# ======================
# FORMULARIO EDITAR CLIENTE
# ======================
@app.get("/clientes/{cliente_id}/editar", response_class=HTMLResponse)
async def formulario_editar_cliente(
    cliente_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    cliente = await db.get(Cliente, cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    return templates.TemplateResponse("clientes/editar.html", {
        "request": request,
        "user": user,
        "cliente": cliente
    })

# ======================
# ACTUALIZAR CLIENTE
# ======================
@app.post("/clientes/{cliente_id}/editar")
async def actualizar_cliente(
    cliente_id: int,
    request: Request,
    tipo_cliente: str = Form(...),
    nombre_razon_social: str = Form(...),
    telefono: str = Form(...),
    email: str = Form(None),
    direccion: str = Form(None),
    notas: str = Form(None),
    activo: str = Form("true"),
    # Campos persona natural
    cedula: str = Form(None),
    primer_nombre: str = Form(None),
    segundo_nombre: str = Form(None),
    primer_apellido: str = Form(None),
    segundo_apellido: str = Form(None),
    # Campos persona jurídica
    rif: str = Form(None),
    representante_legal: str = Form(None),
    telefono_empresa: str = Form(None),
    sitio_web: str = Form(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    cliente = await db.get(Cliente, cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Actualizar campos comunes
    cliente.nombre_razon_social = nombre_razon_social
    cliente.telefono = telefono
    cliente.email = email
    cliente.direccion = direccion
    cliente.notas = notas
    cliente.activo = activo.lower() == "true"
    
    # Actualizar campos según tipo
    if tipo_cliente == "natural":
        cliente.cedula = cedula
        cliente.primer_nombre = primer_nombre
        cliente.segundo_nombre = segundo_nombre
        cliente.primer_apellido = primer_apellido
        cliente.segundo_apellido = segundo_apellido
        # Limpiar campos de empresa
        cliente.rif = None
        cliente.representante_legal = None
        cliente.telefono_empresa = None
        cliente.sitio_web = None
    else:
        cliente.rif = rif
        cliente.representante_legal = representante_legal
        cliente.telefono_empresa = telefono_empresa
        cliente.sitio_web = sitio_web
        # Limpiar campos de persona natural
        cliente.cedula = None
        cliente.primer_nombre = None
        cliente.segundo_nombre = None
        cliente.primer_apellido = None
        cliente.segundo_apellido = None
    
    await db.commit()
    
   # En la ruta de actualización
    return RedirectResponse(url="/clientes?edited=1", status_code=status.HTTP_303_SEE_OTHER)

# ======================
# ACTIVAR/DESACTIVAR CLIENTE
# ======================
@app.post("/clientes/{cliente_id}/activar")
async def activar_cliente(
    cliente_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    cliente = await db.get(Cliente, cliente_id)
    if cliente:
        cliente.activo = True
        await db.commit()
    
    return RedirectResponse(url="/clientes?activated=1", status_code=status.HTTP_303_SEE_OTHER)

@app.post("/clientes/{cliente_id}/desactivar")
async def desactivar_cliente(
    cliente_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    cliente = await db.get(Cliente, cliente_id)
    if cliente:
        cliente.activo = False
        await db.commit()
    
    return RedirectResponse(url="/clientes?deactivated=1", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/presupuestos")
async def listar_presupuestos(
    request: Request,
    cliente_id: str = Query(None),
    estado: str = Query(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Construir query base
        query = select(Presupuesto).options(
            selectinload(Presupuesto.cliente)
        ).order_by(Presupuesto.id.desc())
        
        # Aplicar filtro de cliente
        if cliente_id and cliente_id != "todos" and cliente_id != "null":
            try:
                cliente_id_int = int(cliente_id)
                query = query.where(Presupuesto.cliente_id == cliente_id_int)
            except (ValueError, TypeError):
                pass  # Ignorar valores inválidos
        
        # Aplicar filtro de estado
        if estado and estado != "todos":
            query = query.where(Presupuesto.estado == estado)
        
        result = await db.execute(query)
        presupuestos = result.scalars().all()
        
        # Obtener clientes para el filtro
        clientes = await db.execute(select(Cliente).where(Cliente.activo == True))
        clientes = clientes.scalars().all()
        
        return templates.TemplateResponse("presupuestos/listar.html", {
            "request": request,
            "user": user,
            "presupuestos": presupuestos,
            "clientes": clientes,
            "filtro_cliente": cliente_id or "todos",
            "filtro_estado": estado or "todos"
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR PRESUPUESTOS: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/presupuestos?error=Error+al+cargar+presupuestos", status_code=303)


@app.get("/presupuestos/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_presupuesto(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Obtener datos necesarios
    clientes_result = await db.execute(select(Cliente).where(Cliente.activo == True))
    clientes = clientes_result.scalars().all()
    
    tipos_result = await db.execute(select(TipoTrabajo).where(TipoTrabajo.activo == True))
    tipos_trabajo = tipos_result.scalars().all()

    # 🔥 OBTENER LA TASA ACTIVA
    service = CurrencyService(db)
    tasa_actual = await service.get_tasa()
    
    return templates.TemplateResponse("presupuestos/nuevo.html", {
        "request": request,
        "user": user,
        "clientes": clientes,
        "tipos_trabajo": tipos_trabajo,
        "tasa_activa": float(tasa_actual) 
    })

@app.post("/presupuestos/nuevo")
async def crear_presupuesto(
    request: Request,
    cliente_id: int = Form(...),
    nombre_trabajo: str = Form(...),
    tasa_aplicada: str = Form(None),
    observaciones: str = Form(None),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    try:
        # Obtener todos los datos del formulario
        form_data = await request.form()
        
        # Extraer listas de items
        conceptos = form_data.getlist('conceptos[]')
        precios_unitarios = form_data.getlist('precios_unitarios[]')
        cantidades = form_data.getlist('cantidades[]')
        unidades_medida = form_data.getlist('unidades_medida[]')
        
        # Validar campos obligatorios básicos
        if not cliente_id or not nombre_trabajo.strip():
            return RedirectResponse(
                url="/presupuestos/nuevo?error=Campos%20obligatorios%20vac%C3%ADos",
                status_code=status.HTTP_303_SEE_OTHER
            )
        
        # Validar y procesar items
        total_base = 0.0
        items_validos = 0
        
        for i in range(len(conceptos)):
            # Verificar que todos los campos del item existan
            if (i < len(conceptos) and 
                i < len(precios_unitarios) and 
                i < len(cantidades)):
                
                concepto = conceptos[i].strip() if conceptos[i] else ""
                precio_str = precios_unitarios[i] if i < len(precios_unitarios) else ""
                cantidad_str = cantidades[i] if i < len(cantidades) else ""
                
                # Validar que el concepto no esté vacío y los valores sean numéricos
                if concepto and precio_str and cantidad_str:
                    try:
                        precio = float(str(precio_str).replace(',', ''))
                        cantidad = float(str(cantidad_str).replace(',', ''))
                        
                        if precio >= 0 and cantidad > 0:
                            subtotal = precio * cantidad
                            total_base += subtotal
                            items_validos += 1
                    except (ValueError, TypeError):
                        continue
        
        # Verificar que haya al menos un item válido
        if items_validos == 0:
            return RedirectResponse(
                url="/presupuestos/nuevo?error=Debe%20agregar%20al%20menos%20un%20item%20v%C3%A1lido",
                status_code=status.HTTP_303_SEE_OTHER
            )
        
        # Calcular total_cliente
        if tasa_aplicada and tasa_aplicada.strip():
            try:
                tasa_valor = float(tasa_aplicada.replace(',', ''))
                total_cliente = total_base * tasa_valor
            except ValueError:
                tasa_valor = None
                total_cliente = total_base
        else:
            tasa_valor = None
            total_cliente = total_base
        
        # Crear presupuesto
        nuevo_presupuesto = Presupuesto(
            numero_presupuesto=await generar_numero_presupuesto(db),
            cliente_id=cliente_id,
            nombre_trabajo=nombre_trabajo.strip(),
            total_base=total_base,
            total_cliente=total_cliente,
            tasa_aplicada=tasa_valor,
            observaciones=observaciones,
            creado_por=user.id,
            estado="borrador"
        )
        
        db.add(nuevo_presupuesto)
        await db.commit()
        await db.refresh(nuevo_presupuesto)
        
        # Guardar los items
        for i in range(len(conceptos)):
            if (i < len(conceptos) and 
                i < len(precios_unitarios) and 
                i < len(cantidades)):
                
                concepto = conceptos[i].strip() if conceptos[i] else ""
                precio_str = precios_unitarios[i] if i < len(precios_unitarios) else ""
                cantidad_str = cantidades[i] if i < len(cantidades) else ""
                unidad_medida = unidades_medida[i] if i < len(unidades_medida) else "unidad"
                
                if concepto and precio_str and cantidad_str:
                    try:
                        precio = float(str(precio_str).replace(',', ''))
                        cantidad = float(str(cantidad_str).replace(',', ''))
                        
                        if precio >= 0 and cantidad > 0:
                            subtotal = precio * cantidad
                            
                            nuevo_item = ItemPresupuesto(
                                presupuesto_id=nuevo_presupuesto.id,
                                concepto=concepto,
                                precio_unitario=precio,
                                cantidad=cantidad,
                                subtotal=subtotal,
                                unidad_medida=unidad_medida
                            )
                            db.add(nuevo_item)
                    except (ValueError, TypeError):
                        continue
        
        await db.commit()
        
        return RedirectResponse(url="/presupuestos?success=1", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Error al crear presupuesto: {e}")
        import traceback
        traceback.print_exc()
        await db.rollback()
        return RedirectResponse(
            url="/presupuestos/nuevo?error=Error%20al%20crear%20el%20presupuesto",
            status_code=status.HTTP_303_SEE_OTHER
        )
@app.get("/presupuestos/{presupuesto_id}/ver", response_class=HTMLResponse)
async def ver_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Cargar presupuesto
    presupuesto = await db.get(Presupuesto, presupuesto_id)
    if not presupuesto:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    
    # Cargar cliente
    cliente = await db.get(Cliente, presupuesto.cliente_id)
    
    # Cargar items
    items_result = await db.execute(
        select(ItemPresupuesto)
        .where(ItemPresupuesto.presupuesto_id == presupuesto_id)
        .order_by(ItemPresupuesto.id)
    )
    items_presupuesto = items_result.scalars().all()
    
    return templates.TemplateResponse("presupuestos/ver.html", {
        "request": request,
        "user": user,
        "presupuesto": presupuesto,
        "cliente": cliente,
        "items_presupuesto": items_presupuesto
    })


@app.get("/presupuestos/{presupuesto_id}", response_class=HTMLResponse)
async def redirigir_ver_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Verificar que el presupuesto exista
    presupuesto = await db.get(Presupuesto, presupuesto_id)
    if not presupuesto:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    
    # Redirigir a la vista de ver
    return RedirectResponse(url=f"/presupuestos/{presupuesto_id}/ver", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/presupuestos/{presupuesto_id}/editar", response_class=HTMLResponse)
async def formulario_editar_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    # Cargar el presupuesto
    presupuesto = await db.get(Presupuesto, presupuesto_id)
    if not presupuesto:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    
    # Cargar los items del presupuesto
    items_result = await db.execute(
        select(ItemPresupuesto)
        .where(ItemPresupuesto.presupuesto_id == presupuesto_id)
        .order_by(ItemPresupuesto.id)
    )
    items_presupuesto = items_result.scalars().all()
    
    # Cargar clientes para el selector
    clientes_result = await db.execute(
        select(Cliente).where(Cliente.activo == True)
    )
    clientes = clientes_result.scalars().all()
    
    return templates.TemplateResponse("presupuestos/editar.html", {
        "request": request,
        "user": user,
        "presupuesto": presupuesto,
        "items_presupuesto": items_presupuesto,  # ← ¡ESTA ES LA CLAVE!
        "clientes": clientes
    })

@app.post("/presupuestos/{presupuesto_id}/editar")
async def actualizar_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    try:
        # Obtener todos los datos del formulario
        form_data = await request.form()
        
        cliente_id = int(form_data.get('cliente_id'))
        nombre_trabajo = form_data.get('nombre_trabajo', '').strip()
        tasa_aplicada = form_data.get('tasa_aplicada')
        observaciones = form_data.get('observaciones')
        
        # Validar campos obligatorios
        if not cliente_id or not nombre_trabajo:
            return RedirectResponse(
                url=f"/presupuestos/{presupuesto_id}/editar?error=Campos%20obligatorios%20vac%C3%ADos",
                status_code=status.HTTP_303_SEE_OTHER
            )
        
        # Extraer items
        conceptos = form_data.getlist('conceptos[]')
        precios_unitarios = form_data.getlist('precios_unitarios[]')
        cantidades = form_data.getlist('cantidades[]')
        unidades_medida = form_data.getlist('unidades_medida[]')
        
        # Calcular total_base desde los items editados
        total_base = 0.0
        items_validos = 0
        
        for i in range(len(conceptos)):
            if (i < len(conceptos) and 
                i < len(precios_unitarios) and 
                i < len(cantidades)):
                
                concepto = conceptos[i].strip() if conceptos[i] else ""
                precio_str = precios_unitarios[i] if i < len(precios_unitarios) else ""
                cantidad_str = cantidades[i] if i < len(cantidades) else ""
                
                if concepto and precio_str and cantidad_str:
                    try:
                        precio = float(str(precio_str).replace(',', ''))
                        cantidad = float(str(cantidad_str).replace(',', ''))
                        
                        if precio >= 0 and cantidad > 0:
                            subtotal = precio * cantidad
                            total_base += subtotal
                            items_validos += 1
                    except (ValueError, TypeError):
                        continue
        
        if items_validos == 0:
            return RedirectResponse(
                url=f"/presupuestos/{presupuesto_id}/editar?error=Debe%20agregar%20al%20menos%20un%20item%20v%C3%A1lido",
                status_code=status.HTTP_303_SEE_OTHER
            )
        
        # Calcular total_cliente
        if tasa_aplicada and tasa_aplicada.strip():
            try:
                tasa_valor = float(tasa_aplicada.replace(',', ''))
                total_cliente = total_base * tasa_valor
            except ValueError:
                tasa_valor = None
                total_cliente = total_base
        else:
            tasa_valor = None
            total_cliente = total_base
        
        # Actualizar presupuesto
        presupuesto = await db.get(Presupuesto, presupuesto_id)
        if not presupuesto:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
        
        presupuesto.cliente_id = cliente_id
        presupuesto.nombre_trabajo = nombre_trabajo
        presupuesto.total_base = total_base
        presupuesto.total_cliente = total_cliente
        presupuesto.tasa_aplicada = tasa_valor
        presupuesto.observaciones = observaciones
        
        # Eliminar items existentes
        await db.execute(
            delete(ItemPresupuesto).where(ItemPresupuesto.presupuesto_id == presupuesto_id)
        )
        
        # Guardar nuevos items
        for i in range(len(conceptos)):
            if (i < len(conceptos) and 
                i < len(precios_unitarios) and 
                i < len(cantidades)):
                
                concepto = conceptos[i].strip() if conceptos[i] else ""
                precio_str = precios_unitarios[i] if i < len(precios_unitarios) else ""
                cantidad_str = cantidades[i] if i < len(cantidades) else ""
                unidad_medida = unidades_medida[i] if i < len(unidades_medida) else "unidad"
                
                if concepto and precio_str and cantidad_str:
                    try:
                        precio = float(str(precio_str).replace(',', ''))
                        cantidad = float(str(cantidad_str).replace(',', ''))
                        
                        if precio >= 0 and cantidad > 0:
                            subtotal = precio * cantidad
                            
                            nuevo_item = ItemPresupuesto(
                                presupuesto_id=presupuesto_id,
                                concepto=concepto,
                                precio_unitario=precio,
                                cantidad=cantidad,
                                subtotal=subtotal,
                                unidad_medida=unidad_medida
                            )
                            db.add(nuevo_item)
                    except (ValueError, TypeError):
                        continue
        
        await db.commit()
        return RedirectResponse(url="/presupuestos?edited=1", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Error al actualizar presupuesto: {e}")
        import traceback
        traceback.print_exc()
        await db.rollback()
        return RedirectResponse(
            url=f"/presupuestos/{presupuesto_id}/editar?error=Error%20al%20actualizar",
            status_code=status.HTTP_303_SEE_OTHER
        )

@app.get("/presupuestos/eliminar/{presupuesto_id}")
async def eliminar_presupuesto(
    presupuesto_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener el presupuesto
        presupuesto = await db.get(Presupuesto, presupuesto_id)
        if not presupuesto:
            return RedirectResponse(url="/presupuestos?error=Presupuesto+no+encontrado", status_code=303)
        
        # Verificar que sea borrador
        if presupuesto.estado != "borrador":
            return RedirectResponse(url="/presupuestos?error=No+se+pueden+eliminar+presupuestos+aprobados", status_code=303)
        
        # Eliminar los items del presupuesto primero
        await db.execute(
            delete(ItemPresupuesto).where(ItemPresupuesto.presupuesto_id == presupuesto_id)
        )
        
        # Eliminar el presupuesto
        await db.delete(presupuesto)
        await db.commit()
        
        return RedirectResponse(url="/presupuestos?success=Presupuesto+eliminado+exitosamente", status_code=303)
        
    except Exception as e:
        print(f"ERROR AL ELIMINAR PRESUPUESTO: {e}")
        await db.rollback()
        return RedirectResponse(url="/presupuestos?error=Error+al+eliminar+presupuesto", status_code=303)

@app.post("/presupuestos/eliminar/{presupuesto_id}")
async def eliminar_presupuesto(
    presupuesto_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener el presupuesto
        presupuesto = await db.get(Presupuesto, presupuesto_id)
        if not presupuesto:
            return RedirectResponse(url="/presupuestos?error=Presupuesto+no+encontrado", status_code=303)
        
        # Verificar que sea borrador
        if presupuesto.estado != "borrador":
            return RedirectResponse(url="/presupuestos?error=No+se+pueden+eliminar+presupuestos+aprobados", status_code=303)
        
        # Eliminar los items del presupuesto primero
        await db.execute(
            delete(ItemPresupuesto).where(ItemPresupuesto.presupuesto_id == presupuesto_id)
        )
        
        # Eliminar el presupuesto
        await db.delete(presupuesto)
        await db.commit()
        
        return RedirectResponse(url="/presupuestos?success=Presupuesto+eliminado+exitosamente", status_code=303)
        
    except Exception as e:
        print(f"ERROR AL ELIMINAR PRESUPUESTO: {e}")
        await db.rollback()
        return RedirectResponse(url="/presupuestos?error=Error+al+eliminar+presupuesto", status_code=303)



import base64
import os
import io


def get_logo_base64():
    """Retorna el logo en base64 para incrustar en PDFs"""
    try:
        # Obtener la ruta absoluta del directorio actual
        current_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(current_dir, 'static', 'logor.png')
        
        if not os.path.exists(logo_path):
            print(f"❌ Logo no encontrado en: {logo_path}")
            return None
        
        with open(logo_path, 'rb') as f:
            logo_data = f.read()
            logo_b64 = base64.b64encode(logo_data).decode('utf-8')
            print(f"✅ Logo cargado correctamente. Tamaño: {len(logo_data)} bytes")
            return logo_b64
    except Exception as e:
        print(f"❌ Error al cargar logo: {e}")
        return None

@app.get("/presupuestos/{presupuesto_id}/pdf", name="generar_pdf_presupuesto")
async def generar_pdf_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    try:
        # Cargar datos del presupuesto
        presupuesto = await db.get(Presupuesto, presupuesto_id)
        if not presupuesto:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
        
        cliente = await db.get(Cliente, presupuesto.cliente_id)
        
        items_result = await db.execute(
            select(ItemPresupuesto)
            .where(ItemPresupuesto.presupuesto_id == presupuesto_id)
            .order_by(ItemPresupuesto.id)
        )
        items_presupuesto = items_result.scalars().all()
        
        # Obtener logo en base64
        logo_b64 = get_logo_base64()
        
        # Renderizar el template HTML
        html_content = templates.TemplateResponse("presupuestos/pdf_template.html", {
            "request": request,
            "presupuesto": presupuesto,
            "cliente": cliente,
            "items_presupuesto": items_presupuesto,
            "logo_b64": logo_b64
        }).body.decode('utf-8')
        
        # Convertir a PDF
        pdf_buffer = io.BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=presupuesto_{presupuesto.numero_presupuesto}.pdf"}
        )
        
    except Exception as e:
        print(f"❌ Error al generar PDF: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(
            url=f"/presupuestos/{presupuesto_id}/ver?error=Error%20al%20generar%20PDF",
            status_code=status.HTTP_303_SEE_OTHER
        )


@app.post("/presupuestos/{presupuesto_id}/enviar")
async def enviar_presupuesto(
    presupuesto_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    presupuesto = await db.get(Presupuesto, presupuesto_id)
    if presupuesto:
        presupuesto.estado = "enviado"
        presupuesto.fecha_envio = func.now()
        await db.commit()
    
    return RedirectResponse(url="/presupuestos?sent=1", status_code=status.HTTP_303_SEE_OTHER)

@app.post("/presupuestos/{presupuesto_id}/aprobar")
async def aprobar_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        presupuesto = await db.get(models.Presupuesto, presupuesto_id)
        if not presupuesto or presupuesto.estado != 'enviado':
            return RedirectResponse(
                url=f"/presupuestos/{presupuesto_id}/ver?error=Presupuesto%20no%20v%C3%A1lido",
                status_code=303
            )
        
        # Cambiar estado a aprobado
        presupuesto.estado = "aprobado"
        await db.commit()
        
        # ✅ REDIRECCIÓN CORRECTA con parámetro de consulta
        return RedirectResponse(
            url=f"/trabajos/nuevo?presupuesto_id={presupuesto_id}",
            status_code=303
        )
        
    except Exception as e:
        print(f"Error al aprobar presupuesto: {e}")
        await db.rollback()
        return RedirectResponse(
            url=f"/presupuestos/{presupuesto_id}/ver?error=Error%20al%20aprobar",
            status_code=303
        )

@app.post("/presupuestos/{presupuesto_id}/rechazar")
async def rechazar_presupuesto(
    presupuesto_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    
    try:
        presupuesto = await db.get(Presupuesto, presupuesto_id)
        if not presupuesto or presupuesto.estado != 'enviado':
            return RedirectResponse(
                url=f"/presupuestos/{presupuesto_id}/ver?error=Presupuesto%20no%20v%C3%A1lido%20para%20rechazar",
                status_code=status.HTTP_303_SEE_OTHER
            )
        
        presupuesto.estado = "rechazado"
        await db.commit()
        
        return RedirectResponse(
            url=f"/presupuestos/{presupuesto_id}/ver?success=Presupuesto%20rechazado",
            status_code=status.HTTP_303_SEE_OTHER
        )
        
    except Exception as e:
        print(f"Error al rechazar presupuesto: {e}")
        await db.rollback()
        return RedirectResponse(
            url=f"/presupuestos/{presupuesto_id}/ver?error=Error%20al%20rechazar",
            status_code=status.HTTP_303_SEE_OTHER
        )

# Registrar pago
@app.post("/trabajos/{trabajo_id}/registrar-pago")
async def registrar_pago(trabajo_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    form_data = await request.form()
    monto = float(form_data.get('monto', 0))
    
    trabajo = await db.get(Trabajo, trabajo_id)
    trabajo.monto_pagado += monto
    trabajo.porcentaje_pagado = int((trabajo.monto_pagado / trabajo.total_presupuestado) * 100)
    await db.commit()
    
    return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}", status_code=303)

# Marcar como entregado
@app.post("/trabajos/{trabajo_id}/entregar")
async def entregar_trabajo(trabajo_id: int, db: AsyncSession = Depends(get_db)):
    trabajo = await db.get(Trabajo, trabajo_id)
    if trabajo.porcentaje_pagado >= 100:
        trabajo.entregado = True
        trabajo.fecha_entrega = datetime.now()
        await db.commit()
    
    return RedirectResponse(url=f"/trabajos/ver/{trabajo_id}", status_code=303)
    



from app.forecasting import generar_alerta_stock
from datetime import datetime, timedelta

@app.get("/inventario/materiales/", response_class=HTMLResponse)
async def listar_materiales(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener materiales y categorías
        query = select(MaterialInventario).join(CategoriaInventario).where(MaterialInventario.activo == True)
        result = await db.execute(query)
        materiales = result.scalars().all()
        
        categorias_query = select(CategoriaInventario).where(CategoriaInventario.activo == True)
        categorias_result = await db.execute(categorias_query)
        categorias = categorias_result.scalars().all()
        
        # ✅ Conversión segura de Decimal a float y preparar alertas
        materiales_lista = []
        alertas_dict = {}
        
        for material in materiales:
            # Conversión de precios y stock
            if material.precio_compra is not None:
                material.precio_compra = float(material.precio_compra)
            if material.precio_venta is not None:
                material.precio_venta = float(material.precio_venta)
            if material.stock_actual is not None:
                material.stock_actual = float(material.stock_actual)
            if material.stock_minimo is not None:
                material.stock_minimo = float(material.stock_minimo)
            
            # Agregar material a la lista principal
            materiales_lista.append(material)
            
            # Obtener historial de salidas del material (últimos 30 días)
            query_salidas = select(MovimientoInventario).where(
                MovimientoInventario.material_id == material.id,
                MovimientoInventario.tipo == "salida",
                MovimientoInventario.fecha >= datetime.utcnow() - timedelta(days=30)
            ).order_by(MovimientoInventario.fecha.desc())
            
            result_salidas = await db.execute(query_salidas)
            historial_salidas = result_salidas.scalars().all()
            
            # Generar alertas
            alerta_tradicional = False
            if material.stock_actual is not None and material.stock_minimo is not None:
                alerta_tradicional = material.stock_actual <= material.stock_minimo
            
            alerta_predictiva = None
            if historial_salidas and material.stock_actual is not None:
                alerta_predictiva = generar_alerta_stock(
                    material_id=material.id,
                    stock_actual=material.stock_actual,
                    historial_consumo=historial_salidas,
                    dias_anticipacion=3
                )
            
            # Guardar alertas en diccionario
            alertas_dict[material.id] = {
                "tradicional": alerta_tradicional,
                "predictiva": alerta_predictiva
            }
        
        return templates.TemplateResponse("inventario/lista.html", {
            "request": request,
            "user": user,
            "materiales": materiales_lista,  # ← Lista de objetos directos
            "alertas": alertas_dict,         # ← Diccionario separado
            "categorias": categorias
        })
        
    except Exception as e:
        import traceback
        print("❌ ERROR EN LISTAR MATERIALES:")
        print(f"Error: {str(e)}")
        print("Traceback:")
        traceback.print_exc()
        return HTMLResponse(f"<h1>Error: {str(e)}</h1>")

@app.get("/inventario/materiales/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_material(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    mensaje_error = request.query_params.get("error")
    mensaje_exito = request.query_params.get("mensaje")
    
    # Obtener categorías y unidades disponibles
    categorias = (await db.execute(
        select(CategoriaInventario).where(CategoriaInventario.activo == True)
    )).scalars().all()
    
    # Unidades comunes
    unidades = ["m²", "unidades", "litros", "rollos", "kg", "cm", "metros"]
    
    return templates.TemplateResponse("inventario/nuevo.html", {
        "request": request,
        "user": user,
        "categorias": categorias,
        "unidades": unidades,
        "material": None,
        "error": mensaje_error,      # ← Pasar el error
        "mensaje": mensaje_exito
    })

@app.post("/inventario/materiales/crear")
async def crear_material(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        # Validar datos requeridos
        codigo = form_data.get("codigo", "").strip()
        nombre = form_data.get("nombre", "").strip()
        categoria_id = int(form_data.get("categoria_id", 0))
        unidad_medida = form_data.get("unidad_medida", "").strip()
        
        if not codigo or not nombre or categoria_id <= 0 or not unidad_medida:
            raise ValueError("Datos incompletos")
        
        # Verificar código único
        existe_codigo = await db.execute(
            select(MaterialInventario).where(MaterialInventario.codigo == codigo)
        )
        if existe_codigo.scalar_one_or_none():
            raise ValueError("El código ya existe")
        
        # Determinar m2_por_unidad
        m2_por_unidad = 1.0
        if unidad_medida in ['rollos', 'unidades']:
            m2_por_unidad = float(form_data.get("m2_por_unidad", 1.0) or 1.0)
        
        # Crear nuevo material
        nuevo_material = MaterialInventario(
            codigo=codigo,
            nombre=nombre,
            descripcion=form_data.get("descripcion", ""),
            categoria_id=categoria_id,
            unidad_medida=unidad_medida,
            m2_por_unidad=m2_por_unidad,
            stock_actual=0.0,  # Empezamos con 0
            stock_minimo=float(form_data.get("stock_minimo", 0) or 0),
            precio_compra=float(form_data.get("precio_compra", 0)),
            precio_venta=parse_decimal(form_data.get("precio_venta", 0)),                 
            ubicacion=form_data.get("ubicacion", ""),
            observaciones=form_data.get("observaciones", ""),
            activo=True
        )
        
        db.add(nuevo_material)
        await db.commit()
        
        return RedirectResponse(
            url="/inventario/materiales/?mensaje=Material+creado+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/inventario/materiales/nuevo/?error={error_msg}",
            status_code=303
        )
        
    
@app.get("/inventario/materiales/editar/{material_id}", response_class=HTMLResponse)
async def formulario_editar_material(
    request: Request,
    material_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener material
    material = await db.get(MaterialInventario, material_id)
    if not material or not material.activo:
        return RedirectResponse(url="/inventario/?error=Material+no+encontrado", status_code=303)
    
    # Obtener categorías
    categorias = (await db.execute(
        select(CategoriaInventario).where(CategoriaInventario.activo == True)
    )).scalars().all()
    
    unidades = ["m²", "unidades", "litros", "rollos", "kg", "cm", "metros"]
    
    return templates.TemplateResponse("inventario/editar.html", {
        "request": request,
        "user": user,
        "material": material,
        "categorias": categorias,
        "unidades": unidades
    })

@app.post("/inventario/materiales/actualizar/{material_id}")
async def actualizar_material(
    request: Request,
    material_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        # Obtener material existente
        material = await db.get(MaterialInventario, material_id)
        if not material:
            raise ValueError("Material no encontrado")
        
        # Actualizar campos
        material.codigo = form_data.get("codigo", material.codigo)
        material.nombre = form_data.get("nombre", material.nombre)
        material.descripcion = form_data.get("descripcion", material.descripcion)
        material.categoria_id = int(form_data.get("categoria_id", material.categoria_id))
        material.unidad_medida = form_data.get("unidad_medida", material.unidad_medida)
        material.stock_minimo = float(form_data.get("stock_minimo", material.stock_minimo))
        material.precio_compra = float(form_data.get("precio_compra", material.precio_compra))
        material.ubicacion = form_data.get("ubicacion", material.ubicacion)
        material.observaciones = form_data.get("observaciones", material.observaciones)
        
        await db.commit()
        
        # ✅ CORRECCIÓN: URL correcta con "materiales/"
        return RedirectResponse(
            url="/inventario/materiales/?mensaje=Material+actualizado+exitosamente",
            status_code=303
        )
    
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/inventario/materiales/editar/{material_id}?error={error_msg}",
            status_code=303
        )
    
@app.get("/inventario/materiales/eliminar/{material_id}")
async def eliminar_material(
    material_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        material = await db.get(MaterialInventario, material_id)
        if material:
            # No eliminamos físicamente, solo desactivamos
            material.activo = False
            await db.commit()
        
        # ✅ CORRECCIÓN: URL correcta con "materiales/"
        return RedirectResponse(
            url="/inventario/materiales/?mensaje=Material+dado+de+baja",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        return RedirectResponse(
            url="/inventario/materiales/?error=Error+al+eliminar+material",
            status_code=303
        )


@app.get("/inventario/categorias/", response_class=HTMLResponse)
async def listar_categorias(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    categorias = (await db.execute(
        select(CategoriaInventario).order_by(CategoriaInventario.nombre)
    )).scalars().all()
    
    return templates.TemplateResponse("inventario/categorias/lista.html", {
        "request": request,
        "user": user,
        "categorias": categorias
    })

@app.get("/inventario/categorias/nuevo", response_class=HTMLResponse)
async def formulario_nueva_categoria(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    return templates.TemplateResponse("inventario/categorias/nuevo.html", {
        "request": request,
        "user": user
    })

@app.post("/inventario/categorias/crear")
async def crear_categoria(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    nombre = form_data.get("nombre", "").strip()
    descripcion = form_data.get("descripcion", "").strip()
    
    if not nombre:
        return RedirectResponse(url="/inventario/categorias/nuevo?error=Nombre+requerido", status_code=303)
    
    # Verificar duplicados
    existe = await db.execute(
        select(CategoriaInventario).where(CategoriaInventario.nombre == nombre)
    )
    if existe.scalar_one_or_none():
        return RedirectResponse(url="/inventario/categorias/nuevo?error=Categoría+ya+existe", status_code=303)
    
    nueva_categoria = CategoriaInventario(nombre=nombre, descripcion=descripcion)
    db.add(nueva_categoria)
    await db.commit()
    
    return RedirectResponse(url="/inventario/categorias/?mensaje=Categoría+creada+exitosamente", status_code=303)

@app.get("/inventario/categorias/editar/{categoria_id}", response_class=HTMLResponse)
async def formulario_editar_categoria(
    request: Request,
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    categoria = await db.get(CategoriaInventario, categoria_id)
    if not categoria:
        return RedirectResponse(url="/inventario/categorias/?error=Categoría+no+encontrada", status_code=303)
    
    return templates.TemplateResponse("inventario/categorias/editar.html", {
        "request": request,
        "user": user,
        "categoria": categoria
    })

@app.post("/inventario/categorias/actualizar/{categoria_id}")
async def actualizar_categoria(
    request: Request,
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    categoria = await db.get(CategoriaInventario, categoria_id)
    if not categoria:
        return RedirectResponse(url="/inventario/categorias/?error=Categoría+no+encontrada", status_code=303)
    
    form_data = await request.form()
    categoria.nombre = form_data.get("nombre", categoria.nombre).strip()
    categoria.descripcion = form_data.get("descripcion", categoria.descripcion).strip()
    categoria.activo = form_data.get("activo") == "on"
    
    await db.commit()
    return RedirectResponse(url="/inventario/categorias/?mensaje=Categoría+actualizada+exitosamente", status_code=303)

@app.get("/inventario/categorias/eliminar/{categoria_id}")
async def eliminar_categoria(
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    categoria = await db.get(CategoriaInventario, categoria_id)
    if categoria:
        categoria.activo = False
        await db.commit()
    
    return RedirectResponse(url="/inventario/categorias/?mensaje=Categoría+dada+de+baja", status_code=303)

# En main.py

@app.get("/inventario/proveedores/", response_class=HTMLResponse)
async def listar_proveedores(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    proveedores = (await db.execute(
        select(Proveedor).where(Proveedor.activo == True).order_by(Proveedor.nombre)
    )).scalars().all()
    
    return templates.TemplateResponse("inventario/proveedores/lista.html", {
        "request": request,
        "user": user,
        "proveedores": proveedores
    })

@app.get("/inventario/proveedores/nuevo", response_class=HTMLResponse)
async def formulario_nuevo_proveedor(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    return templates.TemplateResponse("inventario/proveedores/nuevo.html", {
        "request": request,
        "user": user
    })

@app.post("/inventario/proveedores/crear")
async def crear_proveedor(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    nombre = form_data.get("nombre", "").strip()
    
    if not nombre:
        return RedirectResponse(url="/inventario/proveedores/nuevo?error=Nombre+requerido", status_code=303)
    
    nuevo_proveedor = Proveedor(
        nombre=nombre,
        contacto=form_data.get("contacto", "").strip(),
        telefono=form_data.get("telefono", "").strip(),
        email=form_data.get("email", "").strip(),
        direccion=form_data.get("direccion", "").strip(),
        ruc=form_data.get("ruc", "").strip()
    )
    
    db.add(nuevo_proveedor)
    await db.commit()
    
    return RedirectResponse(url="/inventario/proveedores/?mensaje=Proveedor+creado+exitosamente", status_code=303)

@app.get("/inventario/proveedores/editar/{proveedor_id}", response_class=HTMLResponse)
async def formulario_editar_proveedor(
    request: Request,
    proveedor_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    proveedor = await db.get(Proveedor, proveedor_id)
    if not proveedor or not proveedor.activo:
        return RedirectResponse(url="/inventario/proveedores/?error=Proveedor+no+encontrado", status_code=303)
    
    return templates.TemplateResponse("inventario/proveedores/editar.html", {
        "request": request,
        "user": user,
        "proveedor": proveedor
    })

@app.post("/inventario/proveedores/actualizar/{proveedor_id}")
async def actualizar_proveedor(
    request: Request,
    proveedor_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    proveedor = await db.get(Proveedor, proveedor_id)
    if not proveedor:
        return RedirectResponse(url="/inventario/proveedores/?error=Proveedor+no+encontrado", status_code=303)
    
    form_data = await request.form()
    proveedor.nombre = form_data.get("nombre", proveedor.nombre).strip()
    proveedor.contacto = form_data.get("contacto", proveedor.contacto).strip()
    proveedor.telefono = form_data.get("telefono", proveedor.telefono).strip()
    proveedor.email = form_data.get("email", proveedor.email).strip()
    proveedor.direccion = form_data.get("direccion", proveedor.direccion).strip()
    proveedor.ruc = form_data.get("ruc", proveedor.ruc).strip()
    
    await db.commit()
    return RedirectResponse(url="/inventario/proveedores/?mensaje=Proveedor+actualizado+exitosamente", status_code=303)

@app.get("/inventario/proveedores/eliminar/{proveedor_id}")
async def eliminar_proveedor(
    proveedor_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    proveedor = await db.get(Proveedor, proveedor_id)
    if proveedor:
        proveedor.activo = False
        await db.commit()
    
    return RedirectResponse(url="/inventario/proveedores/?mensaje=Proveedor+dado+de+baja", status_code=303)



# ===== ENTRADAS DE INVENTARIO =====

@app.get("/inventario/entradas/", response_class=HTMLResponse)
async def listar_entradas(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        query = select(EntradaInventario).options(
            selectinload(EntradaInventario.material),
            selectinload(EntradaInventario.proveedor)
        ).order_by(EntradaInventario.fecha_entrada.desc())
        
        result = await db.execute(query)
        entradas = result.scalars().all()
        
        # ✅ Conversión segura de Decimal a float
        for entrada in entradas:
            if entrada.precio_compra is not None:
                entrada.precio_compra = float(entrada.precio_compra)
            if entrada.material and entrada.material.precio_compra is not None:
                entrada.material.precio_compra = float(entrada.material.precio_compra)
            if entrada.material and entrada.material.stock_actual is not None:
                entrada.material.stock_actual = float(entrada.material.stock_actual)
        
        return templates.TemplateResponse("inventario/entradas/lista.html", {
            "request": request,
            "user": user,
            "entradas": entradas
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR ENTRADAS: {e}")
        return HTMLResponse(f"<h1>Error al cargar entradas</h1><p>{str(e)}</p>")
    
@app.get("/inventario/entradas/nuevo", response_class=HTMLResponse)
async def formulario_nueva_entrada(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener materiales y proveedores activos
    materiales = (await db.execute(
        select(MaterialInventario).where(MaterialInventario.activo == True).order_by(MaterialInventario.nombre)
    )).scalars().all()
    
    proveedores = (await db.execute(
        select(Proveedor).where(Proveedor.activo == True).order_by(Proveedor.nombre)
    )).scalars().all()
    
    return templates.TemplateResponse("inventario/entradas/nuevo.html", {
        "request": request,
        "user": user,
        "materiales": materiales,
        "proveedores": proveedores
    })

@app.post("/inventario/entradas/crear")
async def crear_entrada(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        material_id = int(form_data.get("material_id", 0))
        cantidad = float(form_data.get("cantidad", 0))
        precio_compra = float(form_data.get("precio_compra", 0))
        
        if material_id <= 0 or cantidad <= 0 or precio_compra <= 0:
            raise ValueError("Datos inválidos")
        
        # Crear entrada
        nueva_entrada = EntradaInventario(
            material_id=material_id,
            proveedor_id=int(form_data.get("proveedor_id")) if form_data.get("proveedor_id") else None,
            cantidad=cantidad,
            precio_compra=precio_compra,
            numero_factura=form_data.get("numero_factura", "").strip(),
            observaciones=form_data.get("observaciones", "").strip(),
            usuario_id=user.id
        )
        
        db.add(nueva_entrada)
        
        # ✅ Actualización simple: solo stock, no precio promedio
        material = await db.get(MaterialInventario, material_id)
        if material:
            material.stock_actual += cantidad
            # No modificamos el precio_unitario del material
        
        # Registrar movimiento
        movimiento = MovimientoInventario(
            material_id=material_id,
            tipo="entrada",
            cantidad=cantidad,
            motivo="Compra a proveedor",
            referencia=form_data.get("numero_factura", ""),
            usuario_id=user.id,
            observaciones=form_data.get("observaciones", "")
        )
        db.add(movimiento)
        
        await db.commit()
        
        return RedirectResponse(
            url="/inventario/entradas/?mensaje=Entrada+registrada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/inventario/entradas/nuevo?error={error_msg}",
            status_code=303
        )
    
# ===== SALIDAS DE INVENTARIO =====

from sqlalchemy.orm import selectinload

@app.get("/inventario/salidas/", response_class=HTMLResponse)
async def listar_salidas(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        from datetime import timedelta
        import pytz
        
        # Función helper para extraer ID del trabajo desde la referencia
        def extraer_trabajo_id(referencia):
            """Extrae el ID numérico de una referencia tipo 'TRABAJO_123'"""
            if referencia and isinstance(referencia, str) and referencia.startswith('TRABAJO_'):
                try:
                    id_str = referencia.split('_')[-1]
                    return int(id_str)
                except (ValueError, IndexError):
                    return None
            return None
        
        # ✅ Cargar relaciones explícitamente con selectinload
        query = select(MovimientoInventario).options(
            selectinload(MovimientoInventario.material),
            selectinload(MovimientoInventario.usuario)
        ).where(MovimientoInventario.tipo == "salida").order_by(MovimientoInventario.fecha.desc())
        
        result = await db.execute(query)
        salidas = result.scalars().all()
        
        # ✅ Zona horaria de Venezuela (UTC-4)
        venezuela_tz = pytz.timezone('America/Caracas')
        
        # ✅ Conversión segura de Decimal a float y ajuste de hora a Venezuela
        for salida in salidas:
            if salida.material and salida.material.precio_compra is not None:
                salida.material.precio_unitario = float(salida.material.precio_compra)
            if salida.material and salida.material.stock_actual is not None:
                salida.material.stock_actual = float(salida.material.stock_actual)
            
            # 🔥 CONVERTIR FECHA UTC A HORA DE VENEZUELA (UTC-4)
            if salida.fecha:
                # Si la fecha es naive (sin zona horaria), asumimos que está en UTC y convertimos
                if salida.fecha.tzinfo is None:
                    # Asignar zona horaria UTC y luego convertir a Venezuela
                    fecha_utc = pytz.UTC.localize(salida.fecha)
                    salida.fecha_local = fecha_utc.astimezone(venezuela_tz)
                    print(f"🕐 Fecha UTC naive: {salida.fecha} -> Venezuela: {salida.fecha_local}")  # DEBUG
                else:
                    # Si ya tiene zona horaria, convertir directamente
                    salida.fecha_local = salida.fecha.astimezone(venezuela_tz)
                    print(f"🕐 Fecha con zona: {salida.fecha} -> Venezuela: {salida.fecha_local}")  # DEBUG
            else:
                salida.fecha_local = None
        
        return templates.TemplateResponse("inventario/salidas/lista.html", {
            "request": request,
            "user": user,
            "salidas": salidas,
            "extraer_trabajo_id": extraer_trabajo_id,
            "timedelta": timedelta,
            "pytz": pytz  # 🔥 Pasamos pytz por si se necesita
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR SALIDAS: {e}")
        import traceback
        traceback.print_exc()
        return HTMLResponse(f"<h1>Error al cargar salidas</h1><p>{str(e)}</p>")
    
@app.get("/inventario/salidas/nuevo", response_class=HTMLResponse)
async def formulario_nueva_salida(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener materiales con stock disponible
    materiales = (await db.execute(
        select(MaterialInventario)
        .where(MaterialInventario.activo == True)
        .where(MaterialInventario.stock_actual > 0)
        .order_by(MaterialInventario.nombre)
    )).scalars().all()
    
    return templates.TemplateResponse("inventario/salidas/nuevo.html", {
        "request": request,
        "user": user,
        "materiales": materiales
    })
@app.post("/inventario/salidas/crear")
async def crear_salida(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        material_id = int(form_data.get("material_id", 0))
        cantidad_str = form_data.get("cantidad", "0")
        motivo = form_data.get("motivo", "").strip()
        
        if material_id <= 0 or not cantidad_str or not motivo:
            raise ValueError("Datos inválidos: faltan campos requeridos")
        
        try:
            cantidad = float(cantidad_str)
        except ValueError:
            raise ValueError("Cantidad debe ser un número válido")
            
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser mayor que cero")
        
        # Verificar que el material existe y tiene suficiente stock
        material = await db.get(MaterialInventario, material_id)
        if not material:
            raise ValueError("Material no encontrado")
        
        # ✅ VALIDACIÓN DE STOCK
        if material.stock_actual < cantidad:
            raise ValueError(f"Stock insuficiente. Disponible: {material.stock_actual} {material.unidad_medida}")
        
        # Registrar movimiento de salida
        movimiento = MovimientoInventario(
            material_id=material_id,
            tipo="salida",
            cantidad=cantidad,
            motivo=motivo,
            referencia=form_data.get("referencia", "").strip(),
            usuario_id=user.id,
            observaciones=form_data.get("observaciones", "")
        )
        db.add(movimiento)
        
        # Actualizar stock del material
        material.stock_actual -= cantidad
        
        await db.commit()
        
        return RedirectResponse(
            url="/inventario/salidas/?mensaje=Salida+registrada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        print(f"ERROR EN SALIDA: {error_msg}")
        return RedirectResponse(
            url=f"/inventario/salidas/nuevo?error={error_msg}",
            status_code=303
        )

@app.get("/inventario/materiales/ver/{material_id}", response_class=HTMLResponse)
async def ver_detalle_material(
    request: Request,
    material_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener material
        material = await db.get(models.MaterialInventario, material_id)
        if not material:
            return RedirectResponse(url="/inventario/materiales/?error=Material+no+encontrado", status_code=303)
        
        # ✅ Conversión segura
        if material.precio_compra is not None:
            material.precio_compra = float(material.precio_compra)
        if material.stock_actual is not None:
            material.stock_actual = float(material.stock_actual)
        if material.stock_minimo is not None:
            material.stock_minimo = float(material.stock_minimo)
        
        # Obtener categoría
        categoria = await db.get(models.CategoriaInventario, material.categoria_id)
        
        # ✅ CORRECCIÓN PRINCIPAL: Obtener salidas desde SalidaMaterial
        salidas_query = select(models.SalidaMaterial).options(
            selectinload(models.SalidaMaterial.material)
        ).where(
            models.SalidaMaterial.material_id == material_id
        ).order_by(models.SalidaMaterial.fecha_salida.desc()).limit(10)
        
        salidas_result = await db.execute(salidas_query)
        movimientos = salidas_result.scalars().all()  # Renombramos a "movimientos" para no cambiar la plantilla
        
        # Convertir precios si es necesario (aunque en salidas no hay precio directo)
        for movimiento in movimientos:
            # Si necesitas el precio unitario del material en cada salida:
            if hasattr(movimiento, 'material') and movimiento.material:
                if movimiento.material.precio_compra is not None:
                    movimiento.material.precio_compra = float(movimiento.material.precio_compra)
        
        # Obtener proveedores que han suministrado este material
        proveedores_query = select(models.Proveedor).join(models.EntradaInventario).where(
            models.EntradaInventario.material_id == material_id
        ).distinct()
        proveedores_result = await db.execute(proveedores_query)
        proveedores = proveedores_result.scalars().all()
        
        return templates.TemplateResponse("inventario/detalle.html", {
            "request": request,
            "user": user,
            "material": material,
            "categoria": categoria,
            "movimientos": movimientos,  # Ahora son salidas reales
            "proveedores": proveedores
        })
        
    except Exception as e:
        print(f"ERROR EN VER DETALLE MATERIAL: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/inventario/materiales/?error=Error+al+cargar+detalle", status_code=303)

@app.get("/inventario/entradas/ver/{entrada_id}", response_class=HTMLResponse)
async def ver_detalle_entrada(
    request: Request,
    entrada_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # ✅ CARGA EXPLÍCITA DE TODAS LAS RELACIONES NECESARIAS
        entrada = await db.execute(
            select(EntradaInventario)
            .options(
                selectinload(EntradaInventario.material).selectinload(MaterialInventario.categoria),
                selectinload(EntradaInventario.proveedor)
            )
            .where(EntradaInventario.id == entrada_id)
        )
        entrada = entrada.scalar_one_or_none()
        
        if not entrada:
            return RedirectResponse(url="/inventario/entradas/?error=Entrada+no+encontrada", status_code=303)
        
        # ✅ CONVERSIÓN SEGURA SIN MODIFICAR OBJETOS ORM
        from decimal import Decimal
        
        def safe_float(value):
            if value is None:
                return 0.0
            if isinstance(value, Decimal):
                return float(value)
            return float(value)
        
        # Crear diccionario seguro para la plantilla
        entrada_dict = {
            "id": entrada.id,
            "cantidad": safe_float(entrada.cantidad),
            "precio_compra": safe_float(entrada.precio_compra),  # ← CORREGIDO
            "numero_factura": entrada.numero_factura or "",
            "fecha_entrada": entrada.fecha_entrada,
            "observaciones": entrada.observaciones or "",
            "usuario_id": entrada.usuario_id,
            "material": {
                "codigo": entrada.material.codigo,
                "nombre": entrada.material.nombre,
                "unidad_medida": entrada.material.unidad_medida,
                "precio_compra": safe_float(entrada.material.precio_compra),  # ← CORREGIDO
                "precio_venta": safe_float(entrada.material.precio_venta),   # ← AGREGADO
                "categoria": {
                "nombre": entrada.material.categoria.nombre
                } if entrada.material.categoria else None
            } if entrada.material else None,
                "proveedor": {
                "nombre": entrada.proveedor.nombre,
                "contacto": entrada.proveedor.contacto or ""
            } if entrada.proveedor else None
        }
        
        return templates.TemplateResponse("inventario/entradas/detalle.html", {
            "request": request,
            "user": user,
            "entrada": entrada_dict
        })
        
    except Exception as e:
        print(f"❌ ERROR EN DETALLE ENTRADA: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/inventario/entradas/?error=Error+al+cargar+detalle", status_code=303)

@app.get("/inventario/entradas/editar/{entrada_id}", response_class=HTMLResponse)
async def editar_entrada_form(
    request: Request,
    entrada_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        entrada = await db.get(EntradaInventario, entrada_id)
        if not entrada:
            return RedirectResponse(url="/inventario/entradas/?error=Entrada+no+encontrada", status_code=303)
        
        # Obtener materiales y proveedores
        materiales = await db.execute(select(MaterialInventario).where(MaterialInventario.activo == True))
        materiales = materiales.scalars().all()
        
        proveedores = await db.execute(select(Proveedor).where(Proveedor.activo == True))
        proveedores = proveedores.scalars().all()
        
        return templates.TemplateResponse("inventario/entradas/editar.html", {
            "request": request,
            "user": user,
            "entrada": entrada,
            "materiales": materiales,
            "proveedores": proveedores
        })
        
    except Exception as e:
        print(f"ERROR EN EDITAR ENTRADA FORM: {e}")
        return RedirectResponse(url="/inventario/entradas/?error=Error+al+cargar+formulario", status_code=303)

@app.post("/inventario/entradas/editar/{entrada_id}")
async def actualizar_entrada(
    request: Request,
    entrada_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        form_data = await request.form()
        
        # Obtener la entrada existente
        entrada = await db.get(EntradaInventario, entrada_id)
        if not entrada:
            return RedirectResponse(url="/inventario/entradas/?error=Entrada+no+encontrada", status_code=303)
        
        # Guardar valores anteriores para ajustar el stock si cambia
        cantidad_anterior = float(entrada.cantidad)
        material_id_anterior = entrada.material_id
        
        # Actualizar campos de la entrada
        entrada.material_id = int(form_data.get("material_id"))
        entrada.proveedor_id = int(form_data.get("proveedor_id")) if form_data.get("proveedor_id") else None
        entrada.cantidad = parse_decimal(form_data.get("cantidad", 0))
        entrada.precio_compra = parse_decimal(form_data.get("precio_compra", 0))
        entrada.numero_factura = form_data.get("numero_factura", "").strip()
        entrada.observaciones = form_data.get("observaciones", "").strip()
        
        # Actualizar el material asociado
        material = await db.get(MaterialInventario, entrada.material_id)
        if material:
            # Actualizar precios del material
            material.precio_compra = entrada.precio_compra
            material.precio_venta = parse_decimal(form_data.get("precio_venta", 0))  # ← ¡ESTO ES LO QUE FALTABA!
            
            # Ajustar stock: restar cantidad anterior y sumar nueva cantidad
            if material_id_anterior == entrada.material_id:
                # Mismo material: ajustar diferencia
                diferencia = float(entrada.cantidad) - cantidad_anterior
                material.stock_actual = float(material.stock_actual) + diferencia
            else:
                # Material diferente: restar del anterior y sumar al nuevo
                if material_id_anterior != entrada.material_id:
                    # Restar del material anterior
                    material_anterior = await db.get(MaterialInventario, material_id_anterior)
                    if material_anterior:
                        material_anterior.stock_actual = float(material_anterior.stock_actual) - cantidad_anterior
                    
                    # Sumar al nuevo material
                    material.stock_actual = float(material.stock_actual) + float(entrada.cantidad)
            
            material.fecha_actualizacion = datetime.utcnow()
        
        await db.commit()
        return RedirectResponse(url="/inventario/entradas/?success=1", status_code=303)
        
    except Exception as e:
        print(f"ERROR AL ACTUALIZAR ENTRADA: {e}")
        await db.rollback()
        return RedirectResponse(url=f"/inventario/entradas/editar/{entrada_id}?error=Error+al+actualizar", status_code=303)

@app.get("/inventario/entradas/eliminar/{entrada_id}")
async def eliminar_entrada(
    entrada_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        entrada = await db.get(EntradaInventario, entrada_id)
        if not entrada:
            raise ValueError("Entrada no encontrada")
        
        # Restaurar stock del material
        material = await db.get(MaterialInventario, entrada.material_id)
        if material:
            material.stock_actual -= entrada.cantidad
        
        # Eliminar la entrada
        await db.delete(entrada)
        await db.commit()
        
        return RedirectResponse(
            url="/inventario/entradas/?mensaje=Entrada+eliminada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/inventario/entradas/?error={error_msg}",
            status_code=303
        )

@app.get("/trabajos/{trabajo_id}/salidas/nuevo", response_class=HTMLResponse)
async def nueva_salida_trabajo(
    request: Request,
    trabajo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Verificar que el trabajo exista
        trabajo = await db.get(Trabajo, trabajo_id)
        if not trabajo:
            return RedirectResponse(url="/trabajos/?error=Trabajo+no+encontrado", status_code=303)
        
        # Obtener materiales disponibles
        materiales = await db.execute(
            select(MaterialInventario).where(MaterialInventario.activo == True)
        )
        materiales = materiales.scalars().all()
        
        return templates.TemplateResponse("inventario/salidas/nuevo_trabajo.html", {
            "request": request,
            "user": user,
            "trabajo": trabajo,
            "materiales": materiales
        })
        
    except Exception as e:
        print(f"ERROR EN NUEVA SALIDA TRABAJO: {e}")
        return RedirectResponse(url=f"/trabajos/{trabajo_id}?error=Error+al+cargar+formulario", status_code=303)

@app.post("/trabajos/{trabajo_id}/salidas/crear")
async def crear_salida_trabajo(
    request: Request,
    trabajo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        material_id = int(form_data.get("material_id", 0))
        cantidad_str = form_data.get("cantidad", "0")
        motivo = form_data.get("motivo", "").strip()
        
        if material_id <= 0 or not cantidad_str or not motivo:
            raise ValueError("Datos inválidos: faltan campos requeridos")
        
        try:
            cantidad = float(cantidad_str)
        except ValueError:
            raise ValueError("Cantidad debe ser un número válido")
            
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser mayor que cero")
        
        # Verificar que el trabajo exista
        trabajo = await db.get(Trabajo, trabajo_id)
        if not trabajo:
            raise ValueError("Trabajo no encontrado")
        
        # Verificar que el material existe y tiene suficiente stock
        material = await db.get(MaterialInventario, material_id)
        if not material:
            raise ValueError("Material no encontrado")
        
        if material.stock_actual < cantidad:
            raise ValueError(f"Stock insuficiente. Disponible: {material.stock_actual} {material.unidad_medida}")
        
        # Registrar movimiento de salida asociado al trabajo
        movimiento = MovimientoInventario(
            material_id=material_id,
            tipo="salida",
            cantidad=cantidad,
            motivo=motivo,
            referencia=f"TRABAJO_{trabajo_id}",
            usuario_id=user.id,
            observaciones=form_data.get("observaciones", ""),
            trabajo_id=trabajo_id  # ← ¡Aquí está la conexión!
        )
        db.add(movimiento)
        
        # Actualizar stock del material
        material.stock_actual -= cantidad
        
        await db.commit()
        
        return RedirectResponse(
            url=f"/trabajos/ver/{trabajo_id}?mensaje=Salida+registrada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        print(f"ERROR EN SALIDA TRABAJO: {error_msg}")
        return RedirectResponse(
            url=f"/trabajos/{trabajo_id}/salidas/nuevo?error={error_msg}",
            status_code=303
        )

@app.get("/activos-fijos/", response_class=HTMLResponse)
async def listar_activos_fijos(
    request: Request,
    categoria: Optional[int] = None,
    estado: Optional[str] = None,
    buscar: Optional[str] = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Consulta base
        query = select(ActivoFijo).options(
            selectinload(ActivoFijo.categoria),
            selectinload(ActivoFijo.empleado_asignado)
        )
        
        # Aplicar filtros
        if categoria:
            query = query.where(ActivoFijo.categoria_id == categoria)
        if estado:
            query = query.where(ActivoFijo.estado == estado)
        if buscar:
            query = query.where(
                or_(
                    ActivoFijo.codigo.ilike(f"%{buscar}%"),
                    ActivoFijo.nombre.ilike(f"%{buscar}%"),
                    ActivoFijo.numero_serie.ilike(f"%{buscar}%")
                )
            )
        
        query = query.order_by(ActivoFijo.fecha_adquisicion.desc())
        
        result = await db.execute(query)
        activos = result.scalars().all()
        
        # Obtener categorías para el filtro
        categorias_query = select(CategoriaActivoFijo).where(CategoriaActivoFijo.activo == True)
        categorias_result = await db.execute(categorias_query)
        categorias = categorias_result.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/lista.html", {
            "request": request,
            "user": user,
            "activos": activos,
            "categorias": categorias,
            "filtro_categoria": categoria,
            "filtro_estado": estado,
            "filtro_buscar": buscar
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR ACTIVOS FIJOS: {e}")
        return RedirectResponse(url="/dashboard?error=Error+al+cargar+activos", status_code=303)

@app.get("/activos-fijos/nuevo", response_class=HTMLResponse)
async def nuevo_activo_fijo_form(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        categorias = await db.execute(select(CategoriaActivoFijo).where(CategoriaActivoFijo.activo == True))
        categorias = categorias.scalars().all()
        
        empleados = await db.execute(select(Empleado).where(Empleado.activo == True))
        empleados = empleados.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/nuevo.html", {
            "request": request,
            "user": user,
            "categorias": categorias,
            "empleados": empleados
        })
        
    except Exception as e:
        print(f"ERROR EN NUEVO ACTIVO FORM: {e}")
        return RedirectResponse(url="/activos-fijos/?error=Error+al+cargar+formulario", status_code=303)

@app.post("/activos-fijos/crear")
async def crear_activo_fijo(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        # Validar datos
        codigo = form_data.get("codigo", "").strip()
        nombre = form_data.get("nombre", "").strip()
        categoria_id = int(form_data.get("categoria_id", 0))
        costo_inicial = float(form_data.get("costo_inicial", 0))
        fecha_adquisicion = form_data.get("fecha_adquisicion", "")
        
        if not codigo or not nombre or categoria_id <= 0 or costo_inicial <= 0 or not fecha_adquisicion:
            raise ValueError("Datos incompletos")
        
        # Verificar código único
        existe_codigo = await db.execute(select(ActivoFijo).where(ActivoFijo.codigo == codigo))
        if existe_codigo.scalar_one_or_none():
            raise ValueError("El código ya existe")
        
        # Verificar número de serie único
        numero_serie = form_data.get("numero_serie", "").strip()
        if numero_serie:
            existe_serie = await db.execute(select(ActivoFijo).where(ActivoFijo.numero_serie == numero_serie))
            if existe_serie.scalar_one_or_none():
                raise ValueError("El número de serie ya existe")
        
        # Crear activo
        nuevo_activo = ActivoFijo(
            codigo=codigo,
            nombre=nombre,
            descripcion=form_data.get("descripcion", ""),
            categoria_id=categoria_id,
            marca=form_data.get("marca", ""),
            modelo=form_data.get("modelo", ""),
            numero_serie=numero_serie,
            fecha_adquisicion=datetime.strptime(fecha_adquisicion, "%Y-%m-%d").date(),
            costo_inicial=costo_inicial,
            valor_residual=float(form_data.get("valor_residual", 0)),
            ubicacion=form_data.get("ubicacion", ""),
            estado=form_data.get("estado", "activo"),
            empleado_asignado_id=int(form_data.get("empleado_asignado_id")) if form_data.get("empleado_asignado_id") else None,
            observaciones=form_data.get("observaciones", "")
        )
        
        db.add(nuevo_activo)
        await db.commit()
        
        return RedirectResponse(
            url="/activos-fijos/?mensaje=Activo+fijo+creado+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/nuevo?error={error_msg}",
            status_code=303
        )


@app.get("/activos-fijos/eliminar/{activo_id}")
async def eliminar_activo_fijo(
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Verificar si el activo existe
        activo = await db.get(ActivoFijo, activo_id)
        if not activo:
            raise ValueError("Activo fijo no encontrado")
        
        # Eliminar el activo
        await db.delete(activo)
        await db.commit()
        
        return RedirectResponse(
            url="/activos-fijos/?mensaje=Activo+fijo+eliminado+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/?error={error_msg}",
            status_code=303
        )

# Lista de categorías
@app.get("/activos-fijos/categorias/", response_class=HTMLResponse)
async def listar_categorias_activos(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        query = select(CategoriaActivoFijo).order_by(CategoriaActivoFijo.nombre)
        result = await db.execute(query)
        categorias = result.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/categorias/lista.html", {
            "request": request,
            "user": user,
            "categorias": categorias
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR CATEGORIAS: {e}")
        return RedirectResponse(url="/activos-fijos/?error=Error+al+cargar+categorias", status_code=303)

# Formulario nuevo
@app.get("/activos-fijos/categorias/nuevo", response_class=HTMLResponse)
async def nueva_categoria_form(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    return templates.TemplateResponse("activos_fijos/categorias/nuevo.html", {
        "request": request,
        "user": user
    })

# Crear categoría
@app.post("/activos-fijos/categorias/crear")
async def crear_categoria(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        nombre = form_data.get("nombre", "").strip()
        vida_util = int(form_data.get("vida_util_anios", 0))
        tasa_dep = float(form_data.get("tasa_depreciacion", 0))
        activo = form_data.get("activo", "true") == "true"
        
        if not nombre or vida_util <= 0 or tasa_dep <= 0:
            raise ValueError("Datos incompletos o inválidos")
        
        # Verificar nombre único
        existe = await db.execute(select(CategoriaActivoFijo).where(CategoriaActivoFijo.nombre == nombre))
        if existe.scalar_one_or_none():
            raise ValueError("El nombre de categoría ya existe")
        
        nueva_categoria = CategoriaActivoFijo(
            nombre=nombre,
            descripcion=form_data.get("descripcion", ""),
            vida_util_anios=vida_util,
            tasa_depreciacion=tasa_dep,
            activo=activo
        )
        
        db.add(nueva_categoria)
        await db.commit()
        
        return RedirectResponse(
            url="/activos-fijos/categorias/?mensaje=Categoría+creada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/categorias/nuevo?error={error_msg}",
            status_code=303
        )

# Editar categoría
@app.get("/activos-fijos/categorias/editar/{categoria_id}", response_class=HTMLResponse)
async def editar_categoria_form(
    request: Request,
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        categoria = await db.get(CategoriaActivoFijo, categoria_id)
        if not categoria:
            return RedirectResponse(url="/activos-fijos/categorias/?error=Categoría+no+encontrada", status_code=303)
        
        return templates.TemplateResponse("activos_fijos/categorias/editar.html", {
            "request": request,
            "user": user,
            "categoria": categoria
        })
        
    except Exception as e:
        print(f"ERROR EN EDITAR CATEGORIA FORM: {e}")
        return RedirectResponse(url="/activos-fijos/categorias/?error=Error+al+cargar+formulario", status_code=303)

# Actualizar categoría
@app.post("/activos-fijos/categorias/actualizar/{categoria_id}")
async def actualizar_categoria(
    request: Request,
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        categoria = await db.get(CategoriaActivoFijo, categoria_id)
        if not categoria:
            raise ValueError("Categoría no encontrada")
        
        nombre = form_data.get("nombre", "").strip()
        vida_util = int(form_data.get("vida_util_anios", 0))
        tasa_dep = float(form_data.get("tasa_depreciacion", 0))
        activo = form_data.get("activo", "true") == "true"
        
        if not nombre or vida_util <= 0 or tasa_dep <= 0:
            raise ValueError("Datos incompletos o inválidos")
        
        # Verificar nombre único (excluyendo la actual)
        existe = await db.execute(
            select(CategoriaActivoFijo)
            .where(CategoriaActivoFijo.nombre == nombre)
            .where(CategoriaActivoFijo.id != categoria_id)
        )
        if existe.scalar_one_or_none():
            raise ValueError("El nombre de categoría ya existe")
        
        categoria.nombre = nombre
        categoria.descripcion = form_data.get("descripcion", "")
        categoria.vida_util_anios = vida_util
        categoria.tasa_depreciacion = tasa_dep
        categoria.activo = activo
        
        await db.commit()
        
        return RedirectResponse(
            url="/activos-fijos/categorias/?mensaje=Categoría+actualizada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/categorias/editar/{categoria_id}?error={error_msg}",
            status_code=303
        )

# Eliminar categoría
@app.get("/activos-fijos/categorias/eliminar/{categoria_id}")
async def eliminar_categoria(
    categoria_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        categoria = await db.get(CategoriaActivoFijo, categoria_id)
        if not categoria:
            raise ValueError("Categoría no encontrada")
        
        # Verificar si hay activos asignados
        activos_asignados = await db.execute(
            select(ActivoFijo).where(ActivoFijo.categoria_id == categoria_id)
        )
        if activos_asignados.scalars().first():
            raise ValueError("No se puede eliminar: hay activos asignados a esta categoría")
        
        await db.delete(categoria)
        await db.commit()
        
        return RedirectResponse(
            url="/activos-fijos/categorias/?mensaje=Categoría+eliminada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/categorias/?error={error_msg}",
            status_code=303
        )

@app.get("/activos-fijos/ver/{activo_id}", response_class=HTMLResponse)
async def ver_activo_fijo(
    request: Request,
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener el activo con sus relaciones
        query = select(ActivoFijo).options(
            selectinload(ActivoFijo.categoria),
            selectinload(ActivoFijo.empleado_asignado)
        ).where(ActivoFijo.id == activo_id)
        
        result = await db.execute(query)
        activo = result.scalar_one_or_none()
        
        if not activo:
            return RedirectResponse(url="/activos-fijos/?error=Activo+fijo+no+encontrado", status_code=303)
        
        # Obtener movimientos del activo
        movimientos_query = select(MovimientoActivoFijo).options(
            selectinload(MovimientoActivoFijo.empleado)
        ).where(MovimientoActivoFijo.activo_id == activo_id).order_by(MovimientoActivoFijo.fecha.desc())
        
        movimientos_result = await db.execute(movimientos_query)
        movimientos = movimientos_result.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/ver.html", {
            "request": request,
            "user": user,
            "activo": activo,
            "movimientos": movimientos
        })
        
    except Exception as e:
        print(f"ERROR EN VER ACTIVO FIJO: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/activos-fijos/?error=Error+al+cargar+detalle", status_code=303)

@app.get("/activos-fijos/editar/{activo_id}", response_class=HTMLResponse)
async def editar_activo_fijo_form(
    request: Request,
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener el activo
        activo = await db.get(ActivoFijo, activo_id)
        if not activo:
            return RedirectResponse(url="/activos-fijos/?error=Activo+fijo+no+encontrado", status_code=303)
        
        # Obtener categorías y empleados para los selects
        categorias_query = select(CategoriaActivoFijo).where(CategoriaActivoFijo.activo == True)
        categorias_result = await db.execute(categorias_query)
        categorias = categorias_result.scalars().all()
        
        empleados_query = select(Empleado).where(Empleado.activo == True)
        empleados_result = await db.execute(empleados_query)
        empleados = empleados_result.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/editar.html", {
            "request": request,
            "user": user,
            "activo": activo,
            "categorias": categorias,
            "empleados": empleados
        })
        
    except Exception as e:
        print(f"ERROR EN EDITAR ACTIVO FORM: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/activos-fijos/?error=Error+al+cargar+formulario", status_code=303)


@app.post("/activos-fijos/actualizar/{activo_id}")
async def actualizar_activo_fijo(
    request: Request,
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        activo = await db.get(ActivoFijo, activo_id)
        if not activo:
            raise ValueError("Activo fijo no encontrado")
        
        # Validar y actualizar campos
        codigo = form_data.get("codigo", "").strip()
        nombre = form_data.get("nombre", "").strip()
        categoria_id = int(form_data.get("categoria_id", 0))
        costo_inicial = float(form_data.get("costo_inicial", 0))
        fecha_adquisicion = form_data.get("fecha_adquisicion", "")
        
        if not codigo or not nombre or categoria_id <= 0 or costo_inicial <= 0 or not fecha_adquisicion:
            raise ValueError("Datos incompletos")
        
        # Verificar código único (excluyendo el actual)
        existe_codigo = await db.execute(
            select(ActivoFijo)
            .where(ActivoFijo.codigo == codigo)
            .where(ActivoFijo.id != activo_id)
        )
        if existe_codigo.scalar_one_or_none():
            raise ValueError("El código ya existe")
        
        # Verificar número de serie único
        numero_serie = form_data.get("numero_serie", "").strip()
        if numero_serie:
            existe_serie = await db.execute(
                select(ActivoFijo)
                .where(ActivoFijo.numero_serie == numero_serie)
                .where(ActivoFijo.id != activo_id)
            )
            if existe_serie.scalar_one_or_none():
                raise ValueError("El número de serie ya existe")
        
        # Actualizar el activo
        activo.codigo = codigo
        activo.nombre = nombre
        activo.descripcion = form_data.get("descripcion", "")
        activo.categoria_id = categoria_id
        activo.marca = form_data.get("marca", "")
        activo.modelo = form_data.get("modelo", "")
        activo.numero_serie = numero_serie
        activo.fecha_adquisicion = datetime.strptime(fecha_adquisicion, "%Y-%m-%d").date()
        activo.costo_inicial = costo_inicial
        activo.valor_residual = float(form_data.get("valor_residual", 0))
        activo.ubicacion = form_data.get("ubicacion", "")
        activo.estado = form_data.get("estado", "activo")
        activo.empleado_asignado_id = int(form_data.get("empleado_asignado_id")) if form_data.get("empleado_asignado_id") else None
        activo.observaciones = form_data.get("observaciones", "")
        activo.fecha_actualizacion = datetime.utcnow()
        
        await db.commit()
        
        return RedirectResponse(
            url=f"/activos-fijos/ver/{activo_id}?mensaje=Activo+fijo+actualizado+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/editar/{activo_id}?error={error_msg}",
            status_code=303
        )  

# Formulario nuevo movimiento
@app.get("/activos-fijos/movimientos/nuevo/{activo_id}", response_class=HTMLResponse)
async def nuevo_movimiento_form(
    request: Request,
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Verificar que el activo exista
        activo = await db.get(ActivoFijo, activo_id)
        if not activo:
            return RedirectResponse(url="/activos-fijos/?error=Activo+fijo+no+encontrado", status_code=303)
        
        # Obtener empleados para el select
        empleados_query = select(Empleado).where(Empleado.activo == True)
        empleados_result = await db.execute(empleados_query)
        empleados = empleados_result.scalars().all()
        
        return templates.TemplateResponse("activos_fijos/movimientos/nuevo.html", {
            "request": request,
            "user": user,
            "activo": activo,
            "empleados": empleados
        })
        
    except Exception as e:
        print(f"ERROR EN NUEVO MOVIMIENTO FORM: {e}")
        return RedirectResponse(url=f"/activos-fijos/ver/{activo_id}?error=Error+al+cargar+formulario", status_code=303)

# Crear movimiento
@app.post("/activos-fijos/movimientos/crear/{activo_id}")
async def crear_movimiento(
    request: Request,
    activo_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        # Verificar que el activo exista
        activo = await db.get(ActivoFijo, activo_id)
        if not activo:
            raise ValueError("Activo fijo no encontrado")
        
        tipo = form_data.get("tipo", "").strip()
        motivo = form_data.get("motivo", "").strip()
        
        if not tipo or not motivo:
            raise ValueError("Tipo y motivo son requeridos")
        
        # Crear el movimiento
        movimiento = MovimientoActivoFijo(
            activo_id=activo_id,
            tipo=tipo,
            empleado_id=user.id,  # El usuario actual realiza el movimiento
            empleado_asignado_id=int(form_data.get("empleado_asignado_id")) if form_data.get("empleado_asignado_id") else None,
            ubicacion_anterior=activo.ubicacion,
            ubicacion_nueva=form_data.get("ubicacion_nueva", "").strip(),
            motivo=motivo,
            observaciones=formando.get("observaciones", "")
        )
        
        db.add(movimiento)
        
        # Actualizar el activo según el tipo de movimiento
        if tipo == 'asignacion' and movimiento.empleado_asignado_id:
            activo.empleado_asignado_id = movimiento.empleado_asignado_id
        elif tipo == 'desasignacion':
            activo.empleado_asignado_id = None
        elif tipo == 'traslado' and movimiento.ubicacion_nueva:
            activo.ubicacion = movimiento.ubicacion_nueva
        elif tipo == 'baja':
            activo.estado = 'dado_baja'
        elif tipo == 'reparacion':
            activo.estado = 'en_reparacion'
            if movimiento.ubicacion_nueva:
                activo.ubicacion = movimiento.ubicacion_nueva
        
        await db.commit()
        
        return RedirectResponse(
            url=f"/activos-fijos/ver/{activo_id}?mensaje=Movimiento+registrado+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/activos-fijos/movimientos/nuevo/{activo_id}?error={error_msg}",
            status_code=303
        )

@app.get("/deducciones/nuevo", response_class=HTMLResponse)
async def nueva_deduccion_form(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener SOLO empleados que tienen asignaciones y están activos
        empleados_query = (
            select(Empleado)
            .join(Asignacion, Empleado.id == Asignacion.empleado_id)
            .where(Empleado.activo == True)
            .distinct()
            .order_by(Empleado.nombre_completo)
        )
        empleados_result = await db.execute(empleados_query)
        empleados = empleados_result.scalars().all()
        
        # Si no hay empleados con asignaciones, mostrar mensaje
        if not empleados:
            return RedirectResponse(url="/deducciones?error=No+hay+empleados+con+trabajos+asignados", status_code=303)
        
        return templates.TemplateResponse("deducciones/nuevo.html", {
            "request": request,
            "user": user,
            "empleados": empleados
        })
        
    except Exception as e:
        print(f"ERROR EN NUEVA DEDUCCION FORM: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/deducciones?error=Error+al+cargar+formulario", status_code=303)
@app.get("/deducciones", response_class=HTMLResponse)
async def listar_deducciones(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Obtener todas las deducciones con relaciones
        query = select(DeduccionEmpleado).options(
            selectinload(DeduccionEmpleado.empleado),
            selectinload(DeduccionEmpleado.trabajo)
        ).order_by(DeduccionEmpleado.fecha_deduccion.desc())
        
        result = await db.execute(query)
        deducciones = result.scalars().all()
        
        return templates.TemplateResponse("deducciones/lista.html", {
            "request": request,
            "user": user,
            "deducciones": deducciones
        })
        
    except Exception as e:
        print(f"ERROR EN LISTAR DEDUCCIONES: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/dashboard?error=Error+al+cargar+deducciones", status_code=303)
    
@app.post("/deducciones/crear")
async def crear_deduccion(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        empleado_id = int(form_data.get("empleado_id", 0))
        trabajo_id = int(form_data.get("trabajo_id", 0))
        tipo_deduccion = form_data.get("tipo_deduccion", "").strip()
        monto = float(form_data.get("monto", 0))
        motivo = form_data.get("motivo", "").strip()
        descripcion = form_data.get("descripcion", "").strip()
        
        # Validaciones
        if not all([empleado_id, trabajo_id, tipo_deduccion, monto > 0, motivo, descripcion]):
            raise ValueError("Todos los campos son requeridos")
        
        if tipo_deduccion not in ['fijo', 'porcentaje', 'monto_total']:
            raise ValueError("Tipo de deducción inválido")
        
        # Verificar que el empleado estuvo asignado al trabajo
        asignacion = await db.execute(
            select(Asignacion)
            .where(Asignacion.empleado_id == empleado_id)
            .where(Asignacion.trabajo_id == trabajo_id)
        )
        if not asignacion.scalar_one_or_none():
            raise ValueError("El empleado no estuvo asignado a este trabajo")
        
        # Crear la deducción
        deduccion = DeduccionEmpleado(
            empleado_id=empleado_id,
            trabajo_id=trabajo_id,
            tipo_deduccion=tipo_deduccion,
            monto=monto,
            motivo=motivo,
            descripcion=descripcion,
            observaciones=form_data.get("observaciones", ""),
            creado_por=user.id
        )
        
        db.add(deduccion)
        await db.commit()
        
        return RedirectResponse(
            url="/deducciones?mensaje=Deducción+registrada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        print(f"ERROR AL CREAR DEDUCCIÓN: {error_msg}")
        return RedirectResponse(
            url=f"/deducciones/nuevo?error={error_msg}",
            status_code=303
        )

from sqlalchemy import text
@app.get("/api/trabajos/empleado/{empleado_id}")
async def get_trabajos_por_empleado(
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    
    try:
        query = """
            SELECT t.id, t.nombre_trabajo
            FROM trabajo t
            INNER JOIN asignacion a ON t.id = a.trabajo_id
            WHERE a.empleado_id = :empleado_id
            AND t.estado = 'completado'
            ORDER BY t.fecha_inicio DESC
        """
        
        result = await db.execute(text(query), {"empleado_id": empleado_id})
        rows = result.fetchall()
        
        trabajos = []
        for row in rows:
            trabajos.append({
                "id": row[0],
                "nombre_trabajo": row[1]
            })
        
        return trabajos
        
    except Exception as e:
        print(f"Error en API: {e}")
        return []
    
@app.get("/deducciones/ver/{deduccion_id}", response_class=HTMLResponse)
async def ver_deduccion(
    request: Request,
    deduccion_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        query = select(DeduccionEmpleado).options(
            selectinload(DeduccionEmpleado.empleado),
            selectinload(DeduccionEmpleado.trabajo),
            selectinload(DeduccionEmpleado.creador)  # ← Esta relación ahora debería funcionar
        ).where(DeduccionEmpleado.id == deduccion_id)
        
        result = await db.execute(query)
        deduccion = result.scalar_one_or_none()
        
        if not deduccion:
            return RedirectResponse(url="/deducciones?error=Deducción+no+encontrada", status_code=303)
        
        return templates.TemplateResponse("deducciones/ver.html", {
            "request": request,
            "user": user,
            "deduccion": deduccion
        })
        
    except Exception as e:
        print(f"ERROR EN VER DEDUCCION: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/deducciones?error=Error+al+cargar+detalle", status_code=303)

@app.get("/deducciones/editar/{deduccion_id}", response_class=HTMLResponse)
async def editar_deduccion_form(
    request: Request,
    deduccion_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Cargar la deducción CON todas sus relaciones
        query = select(DeduccionEmpleado).options(
            selectinload(DeduccionEmpleado.empleado),
            selectinload(DeduccionEmpleado.trabajo),
            selectinload(DeduccionEmpleado.creador)  # ← ¡Esta línea es crucial!
        ).where(DeduccionEmpleado.id == deduccion_id)
        
        result = await db.execute(query)
        deduccion = result.scalar_one_or_none()
        
        if not deduccion:
            return RedirectResponse(url="/deducciones?error=Deducción+no+encontrada", status_code=303)
        
        return templates.TemplateResponse("deducciones/editar.html", {
            "request": request,
            "user": user,
            "deduccion": deduccion
        })
        
    except Exception as e:
        print(f"ERROR EN EDITAR DEDUCCION FORM: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(url="/deducciones?error=Error+al+cargar+formulario", status_code=303)

# En actualizar_deduccion (versión simple)
@app.post("/deducciones/actualizar/{deduccion_id}")
async def actualizar_deduccion(
    request: Request,
    deduccion_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        deduccion = await db.get(DeduccionEmpleado, deduccion_id)
        if not deduccion:
            raise ValueError("Deducción no encontrada")
        
        # Actualizar todos los campos incluyendo estado
        deduccion.tipo_deduccion = form_data.get("tipo_deduccion", deduccion.tipo_deduccion)
        deduccion.monto = float(form_data.get("monto", deduccion.monto))
        deduccion.motivo = form_data.get("motivo", deduccion.motivo)
        deduccion.descripcion = form_data.get("descripcion", deduccion.descripcion)
        deduccion.observaciones = form_data.get("observaciones", deduccion.observaciones)
        deduccion.estado = form_data.get("estado", deduccion.estado)  # ← Simple y directo
        
        await db.commit()
        
        return RedirectResponse(
            url=f"/deducciones/ver/{deduccion_id}?mensaje=Deducción+actualizada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/deducciones/editar/{deduccion_id}?error={error_msg}",
            status_code=303
        )

@app.get("/deducciones/eliminar/{deduccion_id}")
async def eliminar_deduccion(
    deduccion_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        deduccion = await db.get(DeduccionEmpleado, deduccion_id)
        if not deduccion:
            raise ValueError("Deducción no encontrada")
        
        await db.delete(deduccion)
        await db.commit()
        
        return RedirectResponse(
            url="/deducciones?mensaje=Deducción+eliminada+exitosamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        error_msg = str(e).replace(" ", "+")
        return RedirectResponse(
            url=f"/deducciones?error={error_msg}",
            status_code=303
        )

async def aplicar_deducciones_pendientes(db: AsyncSession, fecha_inicio, fecha_fin):
    """
    Aplica automáticamente las deducciones pendientes en un rango de fechas
    """
    try:
        # Obtener deducciones pendientes en el rango de fechas
        query = select(DeduccionEmpleado).where(
            DeduccionEmpleado.estado == "pendiente",
            DeduccionEmpleado.fecha_deduccion >= fecha_inicio,
            DeduccionEmpleado.fecha_deduccion <= fecha_fin
        )
        
        result = await db.execute(query)
        deducciones_pendientes = result.scalars().all()
        
        # Aplicar cada deducción
        for deduccion in deducciones_pendientes:
            deduccion.estado = "aplicada"
            deduccion.fecha_aplicacion = datetime.utcnow()
        
        if deducciones_pendientes:
            await db.commit()
            print(f"Aplicadas {len(deducciones_pendientes)} deducciones automáticamente")
        
        return len(deducciones_pendientes)
        
    except Exception as e:
        await db.rollback()
        print(f"Error al aplicar deducciones automáticamente: {e}")
        return 0

from datetime import date
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from datetime import date
from sqlalchemy import select
from sqlalchemy.orm import joinedload

@app.get("/finanzas/caja-diaria/nuevo", response_class=HTMLResponse)
async def formulario_cierre_caja(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    hoy = date.today()
    
    # ✅ FILTRAR TRABAJOS CON PAGOS HOY
    trabajos_con_pagos_hoy = (await db.execute(
        select(Trabajo).where(
            cast(Trabajo.fecha_pago,SQLDate) == hoy,          # ← Comparación exacta de fecha
            Trabajo.porcentaje_pagado > 0       # ← Tiene al menos un pago
        )
    )).scalars().all()
    
    # ✅ AGRUPAR INGRESOS POR MÉTODO DE PAGO Y MONEDA CORRECTA
    ingresos_por_tipo_usd = {
        'efectivo_usd': 0.0,
        'efectivo_bs': 0.0, 
        'pago_movil': 0.0,
        'tarjeta': 0.0
    }
    ingresos_por_tipo_bs = {
        'efectivo_usd': 0.0,
        'efectivo_bs': 0.0, 
        'pago_movil': 0.0,
        'tarjeta': 0.0
    }
    ingreso_total_usd = 0.0
    ingreso_total_bs = 0.0
    
    for trabajo in trabajos_con_pagos_hoy:
        metodo = trabajo.metodo_pago or 'efectivo_usd'
        monto_usd = float(trabajo.monto_pagado_usd or 0.0)
        tasa = float(trabajo.tasa_cambio_actual or 36.50)
        monto_bs = monto_usd * tasa
        
        if metodo == 'efectivo_usd':
            ingresos_por_tipo_usd[metodo] += monto_usd
        else:
            # Para efectivo_bs, pago_movil, tarjeta → mostrar en Bs
            ingresos_por_tipo_bs[metodo] += monto_bs
        
        ingreso_total_usd += monto_usd
        ingreso_total_bs += monto_bs
    
    # === TRABAJOS PAGADOS HOY (para mostrar detalles) ===
    trabajos_info = []
    for trabajo in trabajos_con_pagos_hoy:
        asignaciones = (await db.execute(
            select(Asignacion).where(Asignacion.trabajo_id == trabajo.id)
        )).scalars().all()
        
        empleados = []
        for asignacion in asignaciones:
            empleado = (await db.execute(
                select(Empleado.nombre_completo).where(Empleado.id == asignacion.empleado_id)
            )).scalar_one_or_none()
            if empleado:
                empleados.append(empleado)
        
        trabajos_info.append({
            "nombre": trabajo.nombre_trabajo,
            "monto_total": float(trabajo.monto_total_usd),
            "monto_pagado": float(trabajo.monto_pagado_usd),
            "porcentaje_pagado": trabajo.porcentaje_pagado,
            "empleados": empleados,
            "metodo_pago": trabajo.metodo_pago or "efectivo_usd",
            "tasa_cambio_actual": float(trabajo.tasa_cambio_actual) if trabajo.tasa_cambio_actual else 36.50  
        })
    
    # === GASTOS DIARIOS ===
    # === GASTOS DIARIOS POR CATEGORÍA ===
    gastos = (await db.execute(
        select(GastoDiario)
        .options(joinedload(GastoDiario.categoria), joinedload(GastoDiario.subcategoria))
        .where(GastoDiario.fecha == hoy)
    )).scalars().all()

    gastos_por_categoria = {
        'materiales': 0.0,
        'personal': 0.0,
        'operacion': 0.0,
        'administracion': 0.0
    }
    gasto_total_bs = 0.0

    for gasto in gastos:
        cat_nombre = gasto.categoria.nombre if gasto.categoria else 'administracion'
        if cat_nombre in gastos_por_categoria:
            gastos_por_categoria[cat_nombre] += float(gasto.monto)
        gasto_total_bs += float(gasto.monto)
    
    # === PRÉSTAMOS DEL DÍA ===
    prestamos = (await db.execute(
        select(Prestamo).where(Prestamo.fecha == hoy)
    )).scalars().all()
    prestamo_total_bs = sum(float(p.monto) for p in prestamos)
    
    # === CÁLCULOS FINANCIEROS ===
    egresos_totales_bs = gasto_total_bs + prestamo_total_bs
    tasa_aproximada = 36.50
    egresos_usd = egresos_totales_bs / tasa_aproximada if tasa_aproximada > 0 else 0
    margen_ganancia = ((ingreso_total_usd - egresos_usd) / ingreso_total_usd * 100) if ingreso_total_usd > 0 else 0
    
    # === VALIDACIONES ===
    alertas = []
    if ingreso_total_usd == 0 and not trabajos_con_pagos_hoy:
        alertas.append("⚠️ No hay ingresos ni trabajos pagados hoy")
    elif ingreso_total_usd == 0:
        alertas.append("⚠️ Hay trabajos pagados pero no se registraron en ingresos")
    
    if margen_ganancia < 25 and ingreso_total_usd > 0:
        alertas.append(f"⚠️ Margen de ganancia bajo ({margen_ganancia:.1f}%)")
    if gasto_total_bs > ingreso_total_bs * 0.3 and ingreso_total_bs > 0:
        alertas.append("⚠️ Gastos operativos altos (>30% de ingresos)")
    
    return templates.TemplateResponse("finanzas/cierre_caja_mejorado.html", {
        "request": request,
        "user": user,
        "hoy": hoy,
        "ingresos_por_tipo_usd": ingresos_por_tipo_usd,
        "ingresos_por_tipo_bs": ingresos_por_tipo_bs,
        "ingreso_total_usd": round(ingreso_total_usd, 2),
        "ingreso_total_bs": round(ingreso_total_bs, 2),
        "trabajos_info": trabajos_info,
        "gastos_por_categoria": gastos_por_categoria,
        "gasto_total_bs": round(gasto_total_bs, 2),
        "prestamo_total_bs": round(prestamo_total_bs, 2),
        "egresos_totales_bs": round(egresos_totales_bs, 2),
        "margen_ganancia": round(margen_ganancia, 1),
        "alertas": alertas,
        "gastos_detalle": [{
            "descripcion": g.descripcion or "Sin descripción",
            "categoria": g.categoria.nombre if g.categoria else "general",
            "subcategoria": g.subcategoria.nombre if g.subcategoria else "",
            "monto": round(float(g.monto), 2)
        } for g in gastos],
        "ingresos_detalle": [{
            "concepto": t["nombre"],
            "monto_usd": round(t["monto_pagado"], 2),
            "monto_bs": round(t["monto_pagado"] * t["tasa_cambio_actual"], 2),
            "metodo_pago": t["metodo_pago"],
            "tasa": round(t["tasa_cambio_actual"], 2)
        } for t in trabajos_info]
    })




@app.post("/finanzas/caja-diaria/guardar")
async def guardar_cierre_caja(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    form_data = await request.form()
    
    try:
        # === DATOS DEL FORMULARIO ===
        caja_inicial_usd = float(form_data.get("caja_inicial_usd", 0))
        caja_inicial_bs = float(form_data.get("caja_inicial_bs", 0))
        caja_final_real_usd = float(form_data.get("caja_final_real_usd", 0))
        caja_final_real_bs = float(form_data.get("caja_final_real_bs", 0))
        tasa_cambio = float(form_data.get("tasa_cambio", 36.50))
        
        # ============================================
        # 📊 DATOS DE CONVERSIÓN BS → USD (desde el frontend)
        # ============================================
        total_pagos_usd_directos = float(form_data.get("total_pagos_usd_directos", 0))
        total_bs_convertido_usd = float(form_data.get("total_bs_convertido_usd", 0))
        pagos_bs_convertidos = form_data.get("pagos_bs_convertidos", "[]")
        
        print("=" * 60)
        print("📊 DATOS RECIBIDOS:")
        print(f"   total_pagos_usd_directos: {total_pagos_usd_directos}")
        print(f"   total_bs_convertido_usd: {total_bs_convertido_usd}")
        print("=" * 60)
        
        # ============================================
        # 🚨 IMPORTANTE: Usar los valores del frontend para el total de ingresos USD
        # ============================================
        total_ingresos_usd = total_pagos_usd_directos + total_bs_convertido_usd
        
        # Para los ingresos en Bs, puedes calcularlos desde la BD o usar 0
        total_ingresos_bs = 0  # O calcúlalo si es necesario
        
        # === GASTOS (EGRESOS) DEL DÍA ===
        hoy = date.today()
        gastos_hoy = (await db.execute(
            select(GastoDiario).where(GastoDiario.fecha == hoy)
        )).scalars().all()
        total_egresos_bs = sum(float(g.monto) for g in gastos_hoy)
        total_egresos_usd = 0.0
        
        # === CÁLCULO DE DIFERENCIAS ===
        # Diferencia USD = CajaFinalUSD - CajaInicialUSD - TotalIngresosUSD + TotalEgresosUSD
        diferencia_usd = caja_final_real_usd - caja_inicial_usd - total_ingresos_usd + total_egresos_usd
        
        # Diferencia Bs = CajaFinalBs - CajaInicialBs - TotalIngresosBs + TotalEgresosBs
        diferencia_bs = caja_final_real_bs - caja_inicial_bs - total_ingresos_bs + total_egresos_bs
        
        print(f"📊 TOTALES CALCULADOS:")
        print(f"   USD Directos: {total_pagos_usd_directos}")
        print(f"   USD Convertidos: {total_bs_convertido_usd}")
        print(f"   Total Ingresos USD: {total_ingresos_usd}")
        
        # === CREAR REGISTRO DE CIERRE ===
        nuevo_cierre = CierreCaja(
            fecha=hoy,
            usuario_id=user.id,
            caja_inicial_usd=caja_inicial_usd,
            caja_inicial_bs=caja_inicial_bs,
            caja_final_real_usd=caja_final_real_usd,
            caja_final_real_bs=caja_final_real_bs,
            tasa_cambio=tasa_cambio,
            pagos_usd_directos=total_pagos_usd_directos,
            pagos_bs_convertidos_usd=total_bs_convertido_usd,
            pagos_bs_convertidos_detalle=pagos_bs_convertidos,
            total_ingresos_usd=total_ingresos_usd,  # 👈 AHORA USA LA SUMA CORRECTA
            total_ingresos_bs=total_ingresos_bs,
            total_egresos_usd=total_egresos_usd,
            total_egresos_bs=total_egresos_bs,
            diferencia_usd=diferencia_usd,
            diferencia_bs=diferencia_bs,
            activo=True
        )
        
        db.add(nuevo_cierre)
        await db.commit()
        
        print(f"✅ CIERRE GUARDADO - ID: {nuevo_cierre.id}")
        print(f"   Total Ingresos USD guardado: {nuevo_cierre.total_ingresos_usd}")
        
        return RedirectResponse(
            url=f"/finanzas/caja-diaria/ver/{nuevo_cierre.id}?success=Cierre+guardado+correctamente",
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(
            url="/finanzas/caja-diaria/nuevo?error=Error+al+guardar+cierre",
            status_code=303
        )
    
@app.get("/finanzas/caja-diaria/ver/{cierre_id}")
async def ver_cierre_caja(
    cierre_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    cierre = (await db.execute(
        select(CierreCaja).where(CierreCaja.id == cierre_id)
    )).scalar_one_or_none()
    
    if not cierre:
        raise HTTPException(status_code=404, detail="Cierre de caja no encontrado")
    
    # Obtener valores de forma segura (usando getattr)
    pagos_usd_directos = getattr(cierre, 'pagos_usd_directos', 0) or 0
    pagos_bs_convertidos_usd = getattr(cierre, 'pagos_bs_convertidos_usd', 0) or 0
    
    # Parsear pagos convertidos
    pagos_convertidos = []
    if hasattr(cierre, 'pagos_bs_convertidos_detalle') and cierre.pagos_bs_convertidos_detalle:
        try:
            pagos_convertidos = json.loads(cierre.pagos_bs_convertidos_detalle)
        except:
            pagos_convertidos = []
    
    # Crear un objeto con valores por defecto para el template
    cierre_data = {
        'id': cierre.id,
        'fecha': cierre.fecha,
        'tasa_cambio': cierre.tasa_cambio,
        'caja_inicial_usd': cierre.caja_inicial_usd,
        'caja_inicial_bs': cierre.caja_inicial_bs,
        'caja_final_real_usd': cierre.caja_final_real_usd,
        'caja_final_real_bs': cierre.caja_final_real_bs,
        'pagos_usd_directos': pagos_usd_directos,
        'pagos_bs_convertidos_usd': pagos_bs_convertidos_usd,
        'total_ingresos_bs': getattr(cierre, 'total_ingresos_bs', 0) or 0,
        'total_egresos_usd': getattr(cierre, 'total_egresos_usd', 0) or 0,
        'total_egresos_bs': getattr(cierre, 'total_egresos_bs', 0) or 0,
        'diferencia_usd': getattr(cierre, 'diferencia_usd', 0) or 0,
        'diferencia_bs': getattr(cierre, 'diferencia_bs', 0) or 0,
    }
    
    return templates.TemplateResponse("finanzas/ver_cierre.html", {
        "request": request,
        "user": user,
        "cierre": cierre_data,  # Usar el dict con valores seguros
        "pagos_convertidos": pagos_convertidos
    })

from sqlalchemy import select, desc

@app.get("/finanzas/caja-diaria/historial", response_class=HTMLResponse)
async def historial_cierres(
    request: Request,
    fecha: Optional[str] = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Construir consulta
    query = select(CierreCaja).order_by(desc(CierreCaja.fecha))
    
    if fecha:
        query = query.where(CierreCaja.fecha == date.fromisoformat(fecha))
    
    # Ejecutar consulta
    cierres = (await db.execute(query)).scalars().all()
    
    # Obtener nombres de usuarios para mostrar
    usuarios = {}
    for cierre in cierres:
        if cierre.usuario_id not in usuarios:
            usuario = (await db.execute(
                select(Usuario.nombre).where(Usuario.id == cierre.usuario_id)
            )).scalar_one_or_none()
            usuarios[cierre.usuario_id] = usuario or "Desconocido"
    
    return templates.TemplateResponse("finanzas/historial_cierres.html", {
        "request": request,
        "user": user,
        "cierres": cierres,
        "usuarios": usuarios,
        "fecha_filtro": fecha
    })

from sqlalchemy import select
@app.get("/clientes/pagos-pendientes", response_class=HTMLResponse)
async def pagos_pendientes(
    request: Request,
    user=Depends(get_current_user_from_session)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # ✅ REDIRIGIR A CUENTAS POR COBRAR
    return RedirectResponse(url="/cuentas-por-cobrar", status_code=302)


@app.get("/finanzas/pagos_semanales", response_class=HTMLResponse)
async def pagos_semanales(
    request: Request,
    semana: Optional[str] = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # ✅ 1. DEFINIR FECHAS CORRECTAMENTE
    if semana:
        hoy = datetime.strptime(semana, "%Y-%m-%d").date()
    else:
        hoy = date.today()
    
    # Calcular lunes de la semana (inicio)
    lunes = hoy - timedelta(days=hoy.weekday())
    # Domingo de la semana (fin)
    domingo = lunes + timedelta(days=6)
    
    print(f"📅 Semana: {lunes} al {domingo}")
    
    # ✅ 2. OBTENER TASA DE CAMBIO
    from app.services.currency_service import CurrencyService
    currency_service = CurrencyService(db)
    tasa_actual = await currency_service.get_tasa()
    print(f"💰 Tasa de cambio: Bs. {tasa_actual:.2f} / USD")
    
    # ✅ 3. OBTENER EMPLEADOS ACTIVOS
    empleados = (await db.execute(
        select(Empleado).where(Empleado.activo == True)
    )).scalars().all()
    
    pagos_empleados = []
    total_general_neto = 0.0
    total_general_bs = 0.0
    
    for empleado in empleados:
        print(f"\n👤 Procesando: {empleado.nombre_completo}")
        
        # ============================================
        # 1. CALCULAR DÍAS TRABAJADOS (CORREGIDO)
        # ============================================
        fecha_inicio_filtro = datetime.combine(lunes, datetime.min.time())
        fecha_fin_filtro = datetime.combine(domingo, datetime.max.time())
        
        # Obtener asistencias de la semana
        asistencias = (await db.execute(
            select(Asistencia).where(
                Asistencia.empleado_id == empleado.id,
                Asistencia.fecha >= lunes,
                Asistencia.fecha <= domingo
            )
        )).scalars().all()
        
        DIAS_LABORALES = 6
        
        # 🔥 CONTAR DÍAS AUSENTES (los que están marcados)
        dias_ausentes = sum(1 for a in asistencias if a.estado == "ausente")
        
        # 🔥 EL RESTO DE DÍAS SON PRESENTES
        dias_presentes = DIAS_LABORALES - dias_ausentes
        
        # Tardanza (si existe)
        dias_tardanza = sum(1 for a in asistencias if a.estado == "tardanza")
        
        print(f"  📅 Asistencias: Presente={dias_presentes}, Ausente={dias_ausentes}, Tardanza={dias_tardanza}")
        
        # ============================================
        # 2. TRABAJOS Y COMISIONES
        # ============================================
        trabajos_detalle = []
        total_comisiones = 0.0
        total_monto_trabajos = 0.0
        total_ganancia_empresa = 0.0
        
        if empleado.tipo_contrato in ["comision", "mixto"]:
            trabajos_comision = (await db.execute(
                select(Trabajo)
                .join(Asignacion, Trabajo.id == Asignacion.trabajo_id)
                .where(
                    Asignacion.empleado_id == empleado.id,
                    Trabajo.estado == "completado",
                    Trabajo.fecha_inicio >= fecha_inicio_filtro,
                    Trabajo.fecha_inicio <= fecha_fin_filtro,
                    Trabajo.porcentaje_pagado > 0
                )
                .order_by(Trabajo.fecha_inicio.desc())
            )).scalars().all()
            
            print(f"  📋 Trabajos encontrados: {len(trabajos_comision)}")
            
            for trabajo in trabajos_comision:
                asignacion = (await db.execute(
                    select(Asignacion).where(
                        Asignacion.trabajo_id == trabajo.id,
                        Asignacion.empleado_id == empleado.id
                    )
                )).scalar_one_or_none()
                
                if asignacion:
                    monto_total = float(trabajo.monto_total_usd or 0.0)
                    comision = float(asignacion.valor_comision or 0.0)
                    ganancia_empresa = monto_total - comision
                    porcentaje_empleado = (comision / monto_total * 100) if monto_total > 0 else 0.0
                    
                    total_comisiones += comision
                    total_monto_trabajos += monto_total
                    total_ganancia_empresa += ganancia_empresa
                    
                    trabajos_detalle.append({
                        "nombre": trabajo.nombre_trabajo,
                        "descripcion": trabajo.descripcion or "",
                        "monto_total": round(monto_total, 2),
                        "comision": round(comision, 2),
                        "ganancia_empresa": round(ganancia_empresa, 2),
                        "porcentaje_empleado": round(porcentaje_empleado, 2),
                        "fecha": trabajo.fecha_inicio.strftime("%d/%m/%Y") if trabajo.fecha_inicio else ""
                    })
        
        # ============================================
        # 3. DEDUCCIONES
        # ============================================
        deducciones_db = (await db.execute(
            select(DeduccionEmpleado).where(
                DeduccionEmpleado.empleado_id == empleado.id,
                DeduccionEmpleado.fecha_deduccion >= fecha_inicio_filtro,
                DeduccionEmpleado.fecha_deduccion <= fecha_fin_filtro,
                DeduccionEmpleado.estado == "pendiente"
            )
        )).scalars().all()
        
        deducciones_detalle = []
        total_deducciones = 0.0
        for d in deducciones_db:
            monto = float(d.monto)
            total_deducciones += monto
            deducciones_detalle.append({
                "motivo": d.motivo,
                "descripcion": d.descripcion or "",
                "monto": round(monto, 2),
                "fecha": d.fecha_deduccion.strftime("%d/%m/%Y") if d.fecha_deduccion else ""
            })
        
        print(f"  🔧 Deducciones: {len(deducciones_detalle)} - Total: ${total_deducciones}")
        
        # ============================================
        # 4. PRÉSTAMOS
        # ============================================
        prestamos_db = (await db.execute(
            select(Prestamo).where(
                Prestamo.empleado_id == empleado.id,
                Prestamo.pagado == False
            )
        )).scalars().all()
        
        prestamos_detalle = []
        total_prestamos = 0.0
        for p in prestamos_db:
            monto = float(p.monto)
            total_prestamos += monto
            prestamos_detalle.append({
                "descripcion": p.descripcion or "Préstamo",
                "monto": round(monto, 2),
                "fecha": p.fecha.strftime("%d/%m/%Y") if p.fecha else "",
                "fecha_inicio": p.fecha_inicio.strftime("%d/%m/%Y") if p.fecha_inicio else "",
                "cuotas": p.cuotas or 1,
                "cuotas_pagadas": p.cuotas_pagadas or 0
            })
        
        print(f"  💳 Préstamos: {len(prestamos_detalle)} - Total: ${total_prestamos}")
        
        # ============================================
        # 5. COMPRAS
        # ============================================
        compras_db = (await db.execute(
            select(CompraEmpleado).where(
                CompraEmpleado.empleado_id == empleado.id,
                CompraEmpleado.estado_pago == 'pendiente'
            )
        )).scalars().all()
        
        compras_detalle = []
        total_compras = 0.0
        for c in compras_db:
            monto = float(c.total_a_descontar or 0.0)
            total_compras += monto
            
            trabajo_nombre = None
            if c.trabajo_id:
                trabajo = await db.get(Trabajo, c.trabajo_id)
                trabajo_nombre = trabajo.nombre_trabajo if trabajo else None
            
            compras_detalle.append({
                "descripcion": c.descripcion_producto or "Compra",
                "fecha": c.fecha_compra.strftime("%d/%m/%Y") if c.fecha_compra else "",
                "trabajo": trabajo_nombre,
                "monto_base": round(monto, 2),
                "monto": round(monto, 2),
            })
        
        print(f"  🛒 Compras: {len(compras_detalle)} - Total: ${total_compras}")
        
        # ============================================
        # 6. CÁLCULOS FINALES
        # ============================================
        descuento_ausencias = 0.0
        
        # 🔥 Calcular sueldo fijo con descuento
        if empleado.tipo_contrato in ["fijo", "mixto"]:
            sueldo_fijo_semanal = float(empleado.sueldo_fijo or 0.0)
            sueldo_diario = sueldo_fijo_semanal / DIAS_LABORALES
            
            # 🔥 Sueldo final = días presentes × sueldo diario
            sueldo_fijo = dias_presentes * sueldo_diario
            
            # 🔥 Descuento = días ausentes × sueldo diario (SOLO PARA MOSTRAR)
            descuento_ausencias = dias_ausentes * sueldo_diario
            
            print(f"  💰 Sueldo semanal: ${sueldo_fijo_semanal:.2f}")
            print(f"  💰 Sueldo diario: ${sueldo_diario:.2f}")
            print(f"  📅 Días presentes: {dias_presentes}")
            print(f"  📅 Días ausentes: {dias_ausentes}")
            print(f"  💵 Sueldo final: ${sueldo_fijo:.2f}")
            print(f"  ❌ Descuento: ${descuento_ausencias:.2f}")
        else:
            sueldo_fijo = 0.0
            print(f"  💰 Empleado por comisión - Sin sueldo fijo")

        # 🔥 Total de egresos = préstamos + deducciones + compras
        total_egresos = total_prestamos + total_deducciones + total_compras

        # 🔥 Total neto = sueldo_fijo + comisiones - egresos
        total_ingresos = sueldo_fijo + total_comisiones
        total_neto = total_ingresos - total_egresos

        # 🔥 Calcular porcentajes globales
        if total_monto_trabajos > 0:
            porcentaje_empleado_global = (total_comisiones / total_monto_trabajos) * 100
            porcentaje_empresa_global = 100.0 - porcentaje_empleado_global
        else:
            porcentaje_empleado_global = 0.0
            porcentaje_empresa_global = 100.0

        if trabajos_detalle:
            porcentaje_promedio = sum(t["porcentaje_empleado"] for t in trabajos_detalle) / len(trabajos_detalle)
        else:
            porcentaje_promedio = 0.0

        # 🔥 Calcular en Bs
        tasa_float = float(tasa_actual) if tasa_actual else 40.0
        total_neto_bs = total_neto * tasa_float

        print(f"  💵 Sueldo fijo final: ${sueldo_fijo:.2f}")
        print(f"  💵 Comisiones: ${total_comisiones:.2f}")
        print(f"  💵 Total ingresos: ${total_ingresos:.2f}")
        print(f"  💵 Total egresos: ${total_egresos:.2f}")
        print(f"  💵 Neto a pagar: ${total_neto:.2f}")
        
        # ============================================
        # 7. AGREGAR AL RESULTADO
        # ============================================
        if total_ingresos > 0 or total_egresos > 0 or trabajos_detalle:
            pago_data = {
                "empleado": {
                    "id": empleado.id,
                    "nombre_completo": empleado.nombre_completo,
                },
                "tipo_contrato": empleado.tipo_contrato,
                "sueldo_fijo": round(sueldo_fijo, 2),
                "comisiones": round(total_comisiones, 2),
                "prestamos": round(total_prestamos, 2),
                "deducciones": round(total_deducciones, 2),
                "compras": round(total_compras, 2),
                "descuento_ausencias": round(descuento_ausencias, 2),
                "dias_presentes": dias_presentes,
                "dias_ausentes": dias_ausentes,
                "dias_tardanza": dias_tardanza,
                "total_neto": round(max(total_neto, 0), 2),
                "total_neto_bs": round(max(total_neto_bs, 0), 2),
                "total_base": round(sueldo_fijo + total_monto_trabajos, 2),
                "total_empresa": round(total_ganancia_empresa, 2),
                "porcentaje_empleado": round(porcentaje_empleado_global, 2),
                "porcentaje_empresa": round(porcentaje_empresa_global, 2),
                "porcentaje_promedio_empleado": round(porcentaje_promedio, 2),
                "trabajos_detalle": trabajos_detalle,
                "deducciones_detalle": deducciones_detalle,
                "prestamos_detalle": prestamos_detalle,
                "compras_detalle": compras_detalle
            }
            
            pagos_empleados.append(pago_data)
            total_general_neto += pago_data["total_neto"]
            total_general_bs += pago_data["total_neto_bs"]
    
    # Ordenar por nombre
    pagos_empleados.sort(key=lambda x: x["empleado"]["nombre_completo"])
    
    print(f"\n📊 TOTAL: {len(pagos_empleados)} empleados con pagos")
    print(f"💰 TOTAL GENERAL NETO: ${total_general_neto:.2f}")
    print(f"💰 TOTAL GENERAL Bs: Bs. {total_general_bs:.2f}")
    
    # 🔥 DETERMINAR SI EL USUARIO ES OPERATIVO
    es_operativo = False
    if user.rol:
        rol_str = str(user.rol).lower()
        es_operativo = rol_str in ["operativo", "operador"]
    
    # 🔥 SELECCIONAR EL TEMPLATE SEGÚN EL ROL
    if es_operativo:
        template_name = "finanzas/pagos_semanales_operativo.html"
    else:
        template_name = "finanzas/pagos_semanales.html"
    
    return templates.TemplateResponse(template_name, {
        "request": request,
        "user": user,
        "pagos_empleados": pagos_empleados,
        "fecha_inicio": lunes,
        "fecha_fin": domingo,
        "tasa_actual": float(tasa_actual),
        "total_general_neto": total_general_neto,
        "total_general_bs": total_general_bs,
        "es_operativo": es_operativo,
        "now": datetime.now()
    })

# ============================================================
# POST CORREGIDO - USANDO @app EN VEZ DE @router
# ============================================================
@app.post("/finanzas/pagos_semanales/procesar", response_class=RedirectResponse)
async def procesar_pagos_semanales(
    request: Request,
    fecha_pago: date = Form(...),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    print(f"✅ POST recibido - Fecha: {fecha_pago}, User: {user.id if user else 'None'}")
    
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        # Calcular semana actual (lunes a domingo)
        dias_desde_lunes = fecha_pago.weekday()
        lunes_actual = fecha_pago - timedelta(days=dias_desde_lunes)
        
        print(f"📅 Semana a procesar: {lunes_actual} al {fecha_pago}")
        
        # Obtener empleados activos
        empleados = (await db.execute(
            select(Empleado).where(Empleado.activo == True)
        )).scalars().all()
        
        if not empleados:
            return RedirectResponse(
                url="/finanzas/pagos_semanales?error=No+hay+empleados+activos", 
                status_code=303
            )
        
        pagos_creados = 0
        
        for empleado in empleados:
            # 1. Sueldo fijo
            total_fijo = float(empleado.sueldo_fijo or 0.0) if empleado.tipo_contrato in ["fijo", "mixto"] else 0.0
            
            # 2. Comisiones por trabajos completados en la semana
            comisiones = 0.0
            if empleado.tipo_contrato in ["comision", "mixto"]:
                trabajos = (await db.execute(
                    select(Trabajo).join(Asignacion).where(
                        Asignacion.empleado_id == empleado.id,
                        Trabajo.estado == "completado",
                        Trabajo.fecha_inicio >= datetime.combine(lunes_actual, datetime.min.time()),
                        Trabajo.fecha_inicio <= datetime.combine(fecha_pago, datetime.max.time()),
                        Trabajo.porcentaje_pagado > 0
                    )
                )).scalars().all()
                
                for t in trabajos:
                    a = (await db.execute(
                        select(Asignacion).where(
                            Asignacion.trabajo_id == t.id, 
                            Asignacion.empleado_id == empleado.id
                        )
                    )).scalar_one_or_none()
                    if a:
                        comisiones += float(a.valor_comision or 0.0)
            
            # 3. Préstamos pendientes
            prestamos_pendientes = (await db.execute(
                select(Prestamo).where(
                    Prestamo.empleado_id == empleado.id, 
                    Prestamo.pagado == False
                )
            )).scalars().all()
            total_prestamos = sum(float(p.monto) for p in prestamos_pendientes)
            
            # 4. Deducciones aplicadas en la semana
            deducciones_aplicadas = (await db.execute(
                select(DeduccionEmpleado).where(
                    DeduccionEmpleado.empleado_id == empleado.id,
                    DeduccionEmpleado.fecha_deduccion >= datetime.combine(lunes_actual, datetime.min.time()),
                    DeduccionEmpleado.fecha_deduccion <= datetime.combine(fecha_pago, datetime.max.time()),
                    DeduccionEmpleado.estado == "pendiente"
                )
            )).scalars().all()
            total_deducciones = sum(float(d.monto) for d in deducciones_aplicadas)
            
            # 5. Calcular neto a pagar
            total_ingresos = total_fijo + comisiones
            total_egresos = total_prestamos + total_deducciones
            total_neto = max(total_ingresos - total_egresos, 0)
            
            # 6. Guardar pago de la semana actual
            if total_ingresos > 0 or total_egresos > 0:
                nuevo_pago = PagoSemanal(
                    empleado_id=empleado.id,
                    fecha_pago=fecha_pago,
                    semana_inicio=lunes_actual,
                    semana_fin=fecha_pago,
                    sueldo_fijo=total_fijo,
                    total_comisiones=comisiones,
                    total_prestamos=total_prestamos,
                    total_deducciones=total_deducciones,
                    total_neto=total_neto,
                    tasa_cambio=0.0,
                    procesado_por=user.id
                )
                db.add(nuevo_pago)
                pagos_creados += 1
                
                # Marcar préstamos como pagados
                for p in prestamos_pendientes:
                    p.pagado = True
                
                # Marcar deducciones como procesadas
                for d in deducciones_aplicadas:
                    d.estado = "procesada"
        
        # 7. Calcular siguiente semana para redirigir
        siguiente_lunes = lunes_actual + timedelta(days=7)
        siguiente_domingo = siguiente_lunes + timedelta(days=6)
        
        # 8. Crear/esqueletos para la siguiente semana (opcional)
        # Verificar si ya existen pagos para la siguiente semana
        pagos_existentes = await db.execute(
            select(PagoSemanal).where(
                PagoSemanal.semana_inicio == siguiente_lunes
            ).limit(1)
        )
        existe = pagos_existentes.first()
        
        if not existe:
            # Crear registros preliminares para la siguiente semana
            for empleado in empleados:
                pago_futuro = PagoSemanal(
                    empleado_id=empleado.id,
                    fecha_pago=siguiente_domingo,
                    semana_inicio=siguiente_lunes,
                    semana_fin=siguiente_domingo,
                    sueldo_fijo=0.0,
                    total_comisiones=0.0,
                    total_prestamos=0.0,
                    total_deducciones=0.0,
                    total_neto=0.0,
                    tasa_cambio=0.0,
                    procesado_por=None
                )
                db.add(pago_futuro)
        
        await db.commit()
        
        print(f"✅ Pagos procesados: {pagos_creados} empleados")
        print(f"📅 Siguiente semana: {siguiente_lunes} al {siguiente_domingo}")
        
        # ✅ Redirigir a la siguiente semana automáticamente
        return RedirectResponse(
            url=f"/finanzas/pagos_semanales?semana={siguiente_domingo.isoformat()}&success=✅+Pagos+procesados+correctamente.+{pagos_creados}+empleados+pagados", 
            status_code=303
        )
        
    except Exception as e:
        await db.rollback()
        print(f"❌ Error al procesar: {e}")
        import traceback
        traceback.print_exc()
        return RedirectResponse(
            url="/finanzas/pagos_semanales?error=Error+interno+al+procesar+los+pagos", 
            status_code=303
        )




# Asegúrate de que el directorio exista
PDF_DIR = "static/reportes/semanales"
os.makedirs(PDF_DIR, exist_ok=True)


# app/main.py

from fastapi import Request, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import insert
from datetime import datetime, timedelta, date
import os
from weasyprint import HTML
from app.models import Empleado, Trabajo, Asignacion, DeduccionEmpleado, Prestamo, ReporteSemanal, ItemFactura 
from app.database import get_db


@app.get("/finanzas/pagos_semanales/pdf")
async def generar_pdf_reporte_semanal(
    request: Request,
    fecha: str = Query(None),
    tasa_cambio: float = Query(None), # 🔹 La tasa viene por query parameter
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    # 1. Validar tasa de cambio
    if tasa_cambio is None or tasa_cambio <= 0:
        raise HTTPException(status_code=400, detail="Debe proporcionar una tasa de cambio válida en el parámetro 'tasa_cambio'.")
    
    # 2. Determinar rango de semana
    if fecha:
        try:
            fecha_ref = datetime.strptime(fecha, "%Y-%m-%d").date()
        except ValueError:
            fecha_ref = date.today()
    else:
        fecha_ref = date.today()
    
    dias_desde_lunes = fecha_ref.weekday()
    fecha_inicio = fecha_ref - timedelta(days=dias_desde_lunes)
    fecha_fin = fecha_inicio + timedelta(days=6)
    
    # 3. Obtener empleados activos
    empleados = (await db.execute(
        select(Empleado).where(Empleado.activo == True)
    )).scalars().all()
    
    pagos_empleados = []
    total_general_neto = 0.0
    total_general_bs = 0.0
    
    for emp in empleados:
        # Inicializar variables
        total_fijo = 0.0
        comisiones = 0.0
        total_monto_trabajos = 0.0
        total_ganancia_empresa = 0.0
        trabajos_detalle = []
        
        # Sueldo Fijo
        if emp.tipo_contrato in ["fijo", "mixto"]:
            total_fijo = float(emp.sueldo_fijo or 0.0)
        
        # Trabajos y Comisiones
        if emp.tipo_contrato in ["comision", "mixto"]:
            trabajos_comision = (await db.execute(
                select(Trabajo)
                .join(Asignacion, Trabajo.id == Asignacion.trabajo_id)
                .where(
                    Asignacion.empleado_id == emp.id,
                    Trabajo.estado == "completado",
                    Trabajo.fecha_inicio >= datetime.combine(fecha_inicio, datetime.min.time()),
                    Trabajo.fecha_inicio <= datetime.combine(fecha_fin, datetime.max.time()),
                    Trabajo.porcentaje_pagado > 0
                )
            )).scalars().all()
            
            for trabajo in trabajos_comision:
                asignacion = (await db.execute(
                    select(Asignacion).where(
                        Asignacion.trabajo_id == trabajo.id,
                        Asignacion.empleado_id == emp.id
                    )
                )).scalar_one_or_none()
                
                if asignacion:
                    monto_total = float(trabajo.monto_total or 0.0)
                    comision = float(asignacion.valor_comision or 0.0)
                    ganancia = monto_total - comision
                    
                    porcentaje_emp = (comision / monto_total * 100) if monto_total > 0 else 0.0
                    
                    comisiones += comision
                    total_monto_trabajos += monto_total
                    total_ganancia_empresa += ganancia
                    
                    trabajos_detalle.append({
                        "nombre": trabajo.nombre_trabajo,
                        "descripcion": trabajo.descripcion,
                        "monto_total": round(monto_total, 2),
                        "comision": round(comision, 2),
                        "ganancia_empresa": round(ganancia, 2),
                        "porcentaje_empleado": round(porcentaje_emp, 2)
                    })
        
        # Deducciones
        deducciones_db = (await db.execute(
            select(DeduccionEmpleado).where(
                DeduccionEmpleado.empleado_id == emp.id,
                DeduccionEmpleado.fecha_deduccion >= datetime.combine(fecha_inicio, datetime.min.time()),
                DeduccionEmpleado.fecha_deduccion <= datetime.combine(fecha_fin, datetime.max.time()),
                DeduccionEmpleado.estado == "pendiente"
            )
        )).scalars().all()
        
        deducciones_detalle = []
        total_deducciones = 0.0
        for d in deducciones_db:
            monto = float(d.monto)
            total_deducciones += monto
            deducciones_detalle.append({
                "motivo": d.motivo,
                "descripcion": d.descripcion,
                "monto": round(monto, 2)
            })
        
        # Préstamos
        prestamos_db = (await db.execute(
            select(Prestamo).where(
                Prestamo.empleado_id == emp.id,
                Prestamo.pagado == False
            )
        )).scalars().all()
        
        prestamos_detalle = []
        total_prestamos = 0.0
        for p in prestamos_db:
            monto = float(p.monto)
            total_prestamos += monto
            prestamos_detalle.append({
                "descripcion": p.descripcion,
                "monto": round(monto, 2),
                "fecha": p.fecha
            })
        
        # Cálculos Finales
        total_base = total_fijo + total_monto_trabajos
        total_bruto = total_fijo + comisiones
        total_descuentos = total_prestamos + total_deducciones
        total_neto = max(total_bruto - total_descuentos, 0)
        
        # 🔹 Calcular monto en Bs usando la tasa manual
        monto_bs = round(total_neto * tasa_cambio, 2)
        
        if total_monto_trabajos > 0:
            porcentaje_empleado_global = (comisiones / total_monto_trabajos) * 100
        else:
            porcentaje_empleado_global = 0.0
        
        porcentaje_empresa_global = 100.0 - porcentaje_empleado_global
        
        if trabajos_detalle:
            porcentaje_promedio = sum(t["porcentaje_empleado"] for t in trabajos_detalle) / len(trabajos_detalle)
        else:
            porcentaje_promedio = 0.0
        
        # Agregar al resultado si hay movimiento
        if total_bruto > 0 or total_descuentos > 0:
            total_general_neto += total_neto
            total_general_bs += monto_bs
            
            pagos_empleados.append({
                "empleado": emp,
                "tipo_contrato": emp.tipo_contrato,
                "sueldo_fijo": round(total_fijo, 2),
                "comisiones": round(comisiones, 2),
                "prestamos": round(total_prestamos, 2),
                "deducciones": round(total_deducciones, 2),
                "total_neto": round(total_neto, 2),
                "monto_bs": monto_bs,  # 🔹 Monto en Bs calculado
                "total_base": round(total_base, 2),
                "total_empresa": round(total_ganancia_empresa, 2),
                "porcentaje_empleado": round(porcentaje_empleado_global, 2),
                "porcentaje_empresa": round(porcentaje_empresa_global, 2),
                "porcentaje_promedio_empleado": round(porcentaje_promedio, 2),
                "trabajos_detalle": trabajos_detalle,
                "deducciones_detalle": deducciones_detalle,
                "prestamos_detalle": prestamos_detalle
            })
    
    # 4. Renderizar template
    try:
        html_content = templates.TemplateResponse("finanzas/pagos_semanales_pdf.html", {
            "request": request,
            "pagos_empleados": pagos_empleados,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "user": user,
            "tasa_cambio": tasa_cambio,         # 🔹 Pasar tasa al template
            "total_general_neto": round(total_general_neto, 2),
            "total_general_bs": round(total_general_bs, 2)
        }).body.decode("utf-8")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al renderizar template: {str(e)}")
    
    # 5. Generar PDF
    try:
        pdf_bytes = HTML(string=html_content, base_url=str(request.base_url)).write_pdf()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")
    
    # 6. Guardar archivo
    filename = f"reporte_semanal_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
    filepath = os.path.join(PDF_DIR, filename)
    
    try:
        with open(filepath, "wb") as f:
            f.write(pdf_bytes)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar archivo: {str(e)}")
    
    # 7. Guardar registro en BD (sin Configuracion)
    reporte_data = {
        "semana_inicio": fecha_inicio,
        "semana_fin": fecha_fin,
        "generado_en": datetime.utcnow(),
        "archivo_pdf_path": filepath,
        "datos": {
            "total_empleados": len(pagos_empleados),
            "total_neto": round(total_general_neto, 2),
            "monto_bs": round(total_general_bs, 2),
            "tasa_cambio": tasa_cambio,  # 🔹 Se guarda la tasa usada
            "generado_por": user.id
        }
    }
    
    await db.execute(insert(ReporteSemanal).values(**reporte_data))
    await db.commit()
    
    # 8. Retornar PDF
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/trabajos/{trabajo_id}", response_class=HTMLResponse)
async def ver_trabajo(
    trabajo_id: int,
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener el trabajo
    trabajo = (await db.execute(
        select(Trabajo).where(Trabajo.id == trabajo_id)
    )).scalar_one_or_none()
    
    if not trabajo:
        raise HTTPException(status_code=404, detail="Trabajo no encontrado")
    
    # Obtener cliente
    cliente = (await db.execute(
        select(Cliente.nombre_razon_social, Cliente.telefono, Cliente.email)
        .where(Cliente.id == trabajo.cliente_id)
    )).first()
    
    # Obtener empleados asignados
    asignaciones = (await db.execute(
        select(Asignacion).where(Asignacion.trabajo_id == trabajo_id)
    )).scalars().all()
    
    empleados = []
    for asignacion in asignaciones:
        empleado = (await db.execute(
            select(Empleado.nombre_completo).where(Empleado.id == asignacion.empleado_id)
        )).scalar_one_or_none()
        if empleado:
            empleados.append(empleado)
    
    return templates.TemplateResponse("trabajos/ver_trabajo.html", {
        "request": request,
        "user": user,
        "trabajo": trabajo,
        "cliente": cliente,
        "empleados": empleados
    })

# app/main.py
from fastapi import Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from decimal import Decimal
from app.models import CompraEmpleado, Empleado, Trabajo, Usuario
from app.database import get_db


# --- RUTA GET: Mostrar Formulario de Compra ---
@app.get("/empleados/{empleado_id}/compras/crear", response_class=HTMLResponse)
async def mostrar_formulario_compra(
    request: Request,
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    # 1. Buscar Empleado
    emp_res = await db.execute(select(Empleado).where(Empleado.id == empleado_id))
    empleado = emp_res.scalar_one_or_none()
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    # 2. CARGAR TRABAJOS (Esto llena el select en el HTML)
    trab_res = await db.execute(select(Trabajo).order_by(Trabajo.nombre_trabajo))
    trabajos = trab_res.scalars().all()

    return templates.TemplateResponse("empleados/compras/crear.html", {
        "request": request,
        "empleado": empleado,
        "trabajos": trabajos, # <--- SE PASAN LOS TRABAJOS
        "user": user
    })


# --- RUTA POST: Guardar Compra ---
@app.post("/empleados/{empleado_id}/compras", response_class=RedirectResponse)
async def crear_compra_empleado(
    empleado_id: int,
    trabajo_id: int = Form(None),
    tipo_producto: str = Form(...),
    descripcion_producto: str = Form(...),
    cantidad: int = Form(...),
    precio_unitario: float = Form(...),
    descuento_aplicado: float = Form(0.00),
    observaciones: str = Form(""),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    # Validaciones básicas
    if precio_unitario <= 0:
        raise HTTPException(status_code=400, detail="El precio debe ser mayor a 0")
    
    subtotal = Decimal(str(precio_unitario)) * Decimal(str(cantidad))
    total = subtotal - Decimal(str(descuento_aplicado))
    
    nueva_compra = CompraEmpleado(
        empleado_id=empleado_id,
        trabajo_id=trabajo_id if trabajo_id else None,
        tipo_producto=tipo_producto,
        descripcion_producto=descripcion_producto,
        cantidad=cantidad,
        precio_unitario=precio_unitario,
        subtotal=float(subtotal),
        descuento_aplicado=descuento_aplicado,
        total_a_descontar=float(total),
        estado_pago="pendiente", # Queda pendiente para nómina
        creado_por=user.id,
        observaciones=observaciones
    )
    
    db.add(nueva_compra)
    await db.commit()
    
    return RedirectResponse(url=f"/empleados/{empleado_id}/compras", status_code=303)
# app/main.py

from fastapi import Request, Depends, HTTPException
from fastapi.responses import HTMLResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from app.models import CompraEmpleado, Empleado, Trabajo
from app.database import get_db




# ⚠️ IMPORTANTE: Esta ruta debe ir ANTES de /empleados/{empleado_id}/compras
@app.get("/empleados/compras", response_class=HTMLResponse)
async def listar_todas_las_compras(
    request: Request,
    skip: int = 0,
    limit: int = 100,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    # Obtener todas las compras con información del empleado y trabajo
    compras_result = await db.execute(
        select(CompraEmpleado)
        .join(CompraEmpleado.empleado)
        .outerjoin(CompraEmpleado.trabajo)
        .order_by(CompraEmpleado.fecha_compra.desc())
        .offset(skip).limit(limit)
    )
    compras = compras_result.scalars().all()

    # Obtener listas para los filtros
    empleados_result = await db.execute(select(Empleado))
    empleados = empleados_result.scalars().all()

    trabajos_result = await db.execute(select(Trabajo))
    trabajos = trabajos_result.scalars().all()

    return templates.TemplateResponse("empleados/compras/listar_todas.html", {
        "request": request,
        "compras": compras,
        "empleados": empleados,
        "trabajos": trabajos,
        "user": user
    })


# --- RUTA LISTAR COMPRAS DEL EMPLEADO ---
@app.get("/empleados/{empleado_id}/compras", response_class=HTMLResponse)
async def listar_compras_empleado(
    request: Request,
    empleado_id: int,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    emp_res = await db.execute(select(Empleado).where(Empleado.id == empleado_id))
    empleado = emp_res.scalar_one_or_none()
    
    compras_res = await db.execute(
        select(CompraEmpleado)
        .where(CompraEmpleado.empleado_id == empleado_id)
        .order_by(CompraEmpleado.fecha_compra.desc())
    )
    compras = compras_res.scalars().all()
    
    return templates.TemplateResponse("empleados/compras/listar.html", {
        "request": request,
        "empleado": empleado,
        "compras": compras,
        "user": user
    })


from fastapi import FastAPI, UploadFile, File, Depends
from sqlalchemy.orm import Session
import pandas as pd
import shutil
import os
from datetime import datetime

# ... tus imports existentes ...

@app.post("/importar-excel-empleados")
async def importar_empleados_excel(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        # Guardar archivo temporalmente
        os.makedirs("uploads", exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ruta_temporal = f"uploads/{timestamp}_{archivo.filename}"
        
        with open(ruta_temporal, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)
        
        # Leer Excel
        df = pd.read_excel(ruta_temporal)
        
        resultados = {
            "importados": 0,
            "duplicados": 0,
            "errores": []
        }
        
        # Procesar cada fila
        for index, fila in df.iterrows():
            try:
                # Verificar si ya existe (por nombre)
                existe = db.query(Empleado).filter(
                    Empleado.nombre_completo == fila['nombre_completo']
                ).first()
                
                if existe:
                    resultados["duplicados"] += 1
                    continue
                
                # Crear nuevo empleado
                nuevo_empleado = Empleado(
                    nombre_completo=fila['nombre_completo'],
                    tipo_contrato=fila['tipo_contrato'],
                    sueldo_fijo=float(fila.get('sueldo_fijo', 0)),
                    activo=fila.get('activo', True),
                    rol_id=int(fila['rol_id']) if 'rol_id' in fila and pd.notna(fila['rol_id']) else None
                )
                
                db.add(nuevo_empleado)
                db.commit()
                resultados["importados"] += 1
                
            except Exception as e:
                resultados["errores"].append(f"Fila {index + 2}: {str(e)}")
                db.rollback()
        
        # Eliminar archivo temporal
        os.remove(ruta_temporal)
        
        return resultados
        
    except Exception as e:
        return {"error": str(e)}

from fastapi import UploadFile, File, Depends
from sqlalchemy.orm import Session
import pandas as pd
import shutil
import os
from datetime import datetime
@app.post("/importar-excel-clientes")
async def importar_clientes_excel(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)  # Nota: AsyncSession
):
    import pandas as pd
    import os
    import shutil
    from datetime import datetime
    
    resultados = {
        "importados": 0,
        "duplicados": 0,
        "errores": []
    }
    
    try:
        # Crear carpeta uploads
        os.makedirs("uploads", exist_ok=True)
        
        # Guardar archivo temporal
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ruta_temporal = f"uploads/{timestamp}_{archivo.filename}"
        
        with open(ruta_temporal, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)
        
        # Leer Excel
        df = pd.read_excel(ruta_temporal)
        
        print("=" * 50)
        print(f"📊 Columnas encontradas: {df.columns.tolist()}")
        print(f"📊 Filas: {len(df)}")
        print("=" * 50)
        
        if len(df) == 0:
            resultados["mensaje"] = "El archivo Excel está vacío"
            return resultados
        
        # Procesar cada fila
        for index, fila in df.iterrows():
            try:
                # Limpiar valores NaN
                fila_limpia = {}
                for col in df.columns:
                    valor = fila[col]
                    if pd.isna(valor):
                        fila_limpia[col] = None
                    else:
                        valor_str = str(valor).strip()
                        if valor_str == '' or valor_str == 'nan':
                            fila_limpia[col] = None
                        else:
                            fila_limpia[col] = valor_str
                
                # Determinar tipo de cliente
                tipo_cliente = None
                if fila_limpia.get('tipo_cliente'):
                    tipo_cliente = str(fila_limpia['tipo_cliente']).lower()
                elif fila_limpia.get('cedula'):
                    tipo_cliente = 'natural'
                elif fila_limpia.get('rif'):
                    tipo_cliente = 'juridico'
                
                if not tipo_cliente:
                    resultados["errores"].append(f"Fila {index + 2}: No se pudo determinar tipo de cliente")
                    continue
                
                # Obtener nombre
                nombre = None
                if fila_limpia.get('nombre_razon_social'):
                    nombre = fila_limpia['nombre_razon_social']
                elif tipo_cliente == 'natural':
                    partes = []
                    if fila_limpia.get('primer_nombre'):
                        partes.append(fila_limpia['primer_nombre'])
                    if fila_limpia.get('segundo_nombre'):
                        partes.append(fila_limpia['segundo_nombre'])
                    if fila_limpia.get('primer_apellido'):
                        partes.append(fila_limpia['primer_apellido'])
                    if fila_limpia.get('segundo_apellido'):
                        partes.append(fila_limpia['segundo_apellido'])
                    nombre = ' '.join(partes) if partes else None
                
                if not nombre:
                    resultados["errores"].append(f"Fila {index + 2}: Nombre o razón social requerido")
                    continue
                
                # Verificar duplicados (usando SELECT asíncrono)
                existe = None
                from sqlalchemy import select
                
                if tipo_cliente == 'natural' and fila_limpia.get('cedula'):
                    stmt = select(Cliente).where(
                        Cliente.tipo_cliente == 'natural',
                        Cliente.cedula == fila_limpia['cedula']
                    )
                    result = await db.execute(stmt)
                    existe = result.scalar_one_or_none()
                    
                elif tipo_cliente == 'juridico' and fila_limpia.get('rif'):
                    stmt = select(Cliente).where(
                        Cliente.tipo_cliente == 'juridico',
                        Cliente.rif == fila_limpia['rif']
                    )
                    result = await db.execute(stmt)
                    existe = result.scalar_one_or_none()
                
                if existe:
                    resultados["duplicados"] += 1
                    continue
                
                # Crear cliente
                nuevo_cliente = Cliente(
                    tipo_cliente=tipo_cliente,
                    nombre_razon_social=nombre,
                    telefono=fila_limpia.get('telefono'),
                    email=fila_limpia.get('email'),
                    direccion=fila_limpia.get('direccion'),
                    notas=fila_limpia.get('notas'),
                    activo=True,
                    cedula=fila_limpia.get('cedula'),
                    primer_nombre=fila_limpia.get('primer_nombre'),
                    segundo_nombre=fila_limpia.get('segundo_nombre'),
                    primer_apellido=fila_limpia.get('primer_apellido'),
                    segundo_apellido=fila_limpia.get('segundo_apellido'),
                    rif=fila_limpia.get('rif'),
                    representante_legal=fila_limpia.get('representante_legal'),
                    telefono_empresa=fila_limpia.get('telefono_empresa'),
                    sitio_web=fila_limpia.get('sitio_web')
                )
                
                db.add(nuevo_cliente)
                await db.commit()  # ¡Importante: await!
                resultados["importados"] += 1
                print(f"✅ Importado: {nombre}")
                
            except Exception as e:
                error_msg = f"Fila {index + 2}: {str(e)}"
                resultados["errores"].append(error_msg)
                print(f"❌ {error_msg}")
                await db.rollback()  # ¡Importante: await!
        
        # Eliminar archivo temporal
        os.remove(ruta_temporal)
        
        print(f"\n📊 RESULTADO: {resultados['importados']} importados, {resultados['duplicados']} duplicados")
        
        return {
            "importados": resultados["importados"],
            "duplicados": resultados["duplicados"],
            "errores": resultados["errores"][:10]
        }
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR: {traceback.format_exc()}")
        return {
            "error": str(e),
            "importados": 0,
            "duplicados": 0,
            "errores": [str(e)]
        }

## ============================================
# CUENTAS POR COBRAR - RUTAS
# Agrega esto al final de tu main.py
# ============================================

# ---------------------------------------------------
# Página principal del módulo
# ---------------------------------------------------
@app.get("/cuentas-cobrar", response_class=HTMLResponse)
async def cuentas_cobrar_index(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    return templates.TemplateResponse("cuentas_cobrar/index.html", {
        "request": request,
        "user": user
    })


# ---------------------------------------------------
# API: Listar cuentas con filtros
# ---------------------------------------------------
@app.get("/api/cuentas-cobrar/listar")
async def api_listar_cuentas(
    request: Request,
    cliente: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    desde: Optional[str] = Query(None),
    hasta: Optional[str] = Query(None),
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    sql = """
        SELECT 
            f.id as factura_id,
            f.numero_factura,
            TO_CHAR(f.fecha_emision, 'YYYY-MM-DD') as fecha_emision,
            TO_CHAR(f.fecha_vencimiento, 'YYYY-MM-DD') as fecha_vencimiento,
            f.cliente_id,
            c.nombre_razon_social as cliente_nombre,
            f.total as monto_total_usd,
            f.total_bs as monto_total_bs,
            COALESCE((
                SELECT SUM(a.monto_usd) FROM abonos_factura a WHERE a.factura_id = f.id
            ), 0) as total_abonado_usd,
            COALESCE((
                SELECT SUM(a.monto_bs) FROM abonos_factura a WHERE a.factura_id = f.id
            ), 0) as total_abonado_bs,
            (f.total - COALESCE((SELECT SUM(a.monto_usd) FROM abonos_factura a WHERE a.factura_id = f.id), 0)) as saldo_pendiente_usd,
            f.estado,
            f.metodo_pago as tipo_pago,
            (CURRENT_DATE - f.fecha_emision) as dias_vencido,
            CASE 
                WHEN f.estado = 'pagada' THEN 'Pagado'
                WHEN f.estado = 'anulada' THEN 'Anulada'
                ELSE 'Pendiente'
            END as estado_texto
        FROM facturas f
        INNER JOIN clientes c ON f.cliente_id = c.id
        WHERE f.estado IN ('pendiente', 'pagada')
    """
    params = {}
    
    if cliente:
        sql += " AND c.nombre_razon_social ILIKE :cliente"
        params["cliente"] = f'%{cliente}%'
    
    if estado:
        sql += " AND f.estado = :estado"
        params["estado"] = estado
    
    if desde:
        sql += " AND f.fecha_emision >= :desde"
        params["desde"] = desde
    
    if hasta:
        sql += " AND f.fecha_emision <= :hasta"
        params["hasta"] = hasta
    
    sql += " ORDER BY f.fecha_emision DESC"
    
    result = await db.execute(text(sql), params)
    rows = result.fetchall()
    
    facturas = []
    total_facturado_usd = 0
    total_facturado_bs = 0
    total_abonado_usd = 0
    total_abonado_bs = 0
    total_pendiente_usd = 0
    
    for row in rows:
        monto_usd = float(row[6]) if row[6] else 0
        monto_bs = float(row[7]) if row[7] else 0
        abonado_usd = float(row[8]) if row[8] else 0
        abonado_bs = float(row[9]) if row[9] else 0
        saldo_usd = float(row[10]) if row[10] else 0
        
        total_facturado_usd += monto_usd
        total_facturado_bs += monto_bs
        total_abonado_usd += abonado_usd
        total_abonado_bs += abonado_bs
        total_pendiente_usd += saldo_usd
        
        facturas.append({
            "id": row[0],
            "numero_documento": row[1],
            "fecha_emision": row[2],
            "fecha_vencimiento": row[3],
            "cliente_id": row[4],
            "cliente_nombre": row[5],
            "monto_total_usd": round(monto_usd, 2),
            "monto_total_bs": round(monto_bs, 2),
            "total_abonado_usd": round(abonado_usd, 2),
            "total_abonado_bs": round(abonado_bs, 2),
            "saldo_pendiente_usd": round(saldo_usd, 2),
            "estado": row[11],
            "tipo_pago": row[12] or "Transferencia",
            "dias_vencido": row[13].days if row[13] else 0,
            "estado_texto": row[14]
        })
    
    return JSONResponse({
        "cuentas": facturas,
        "resumen": {
            "total_facturado_usd": round(total_facturado_usd, 2),
            "total_facturado_bs": round(total_facturado_bs, 2),
            "total_abonado_usd": round(total_abonado_usd, 2),
            "total_abonado_bs": round(total_abonado_bs, 2),
            "total_pendiente_usd": round(total_pendiente_usd, 2)
        }
    })
# ---------------------------------------------------
# API: Crear nueva cuenta por cobrar
# ---------------------------------------------------
@app.post("/api/cuentas-cobrar/crear")
async def api_crear_cuenta(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    data = await request.json()
    
    cliente_id = data.get('cliente_id')
    monto_usd = data.get('monto_usd')
    dias_credito = data.get('dias_credito', 30)
    descripcion = data.get('descripcion', '')
    
    if not cliente_id or not monto_usd:
        return JSONResponse({"error": "Cliente y monto son requeridos"}, status_code=400)
    
    try:
        # Verificar que el cliente existe
        cliente_check = await db.execute(
            text("SELECT id FROM clientes WHERE id = :id AND activo = true"),
            {"id": cliente_id}
        )
        if not cliente_check.fetchone():
            return JSONResponse({"error": "Cliente no encontrado o inactivo"}, status_code=400)
        
        # Tasa por defecto
        tasa_valor = 36.50
        
        monto_bs = float(monto_usd) * tasa_valor
        
        # Generar número de documento
        anio = datetime.now().year
        num_result = await db.execute(
            text("""
                SELECT COALESCE(MAX(CAST(SPLIT_PART(numero_documento, '-', 3) AS INTEGER)), 0) + 1 
                FROM cuentas_cobrar 
                WHERE numero_documento LIKE :pattern
            """),
            {"pattern": f'CXC-{anio}-%'}
        )
        next_num = num_result.fetchone()[0] or 1
        numero_documento = f'CXC-{anio}-{str(next_num).zfill(4)}'
        
        # Fechas
        fecha_emision = date.today()
        fecha_vencimiento = fecha_emision + timedelta(days=dias_credito)
        
        # Insertar cuenta
        await db.execute(
            text("""
                INSERT INTO cuentas_cobrar (
                    numero_documento, fecha_emision, fecha_vencimiento, cliente_id,
                    monto_total_bs, monto_total_usd, saldo_pendiente_bs, saldo_pendiente_usd,
                    tasa_cambio_original, descripcion, estado
                ) VALUES (
                    :numero_documento, :fecha_emision, :fecha_vencimiento, :cliente_id,
                    :monto_total_bs, :monto_total_usd, :saldo_pendiente_bs, :saldo_pendiente_usd,
                    :tasa_cambio_original, :descripcion, 'pendiente'
                )
            """),
            {
                "numero_documento": numero_documento,
                "fecha_emision": fecha_emision,
                "fecha_vencimiento": fecha_vencimiento,
                "cliente_id": cliente_id,
                "monto_total_bs": monto_bs,
                "monto_total_usd": monto_usd,
                "saldo_pendiente_bs": monto_bs,
                "saldo_pendiente_usd": monto_usd,
                "tasa_cambio_original": tasa_valor,
                "descripcion": descripcion
            }
        )
        
        await db.commit()
        
        return JSONResponse({"success": True, "numero": numero_documento})
    
    except Exception as e:
        await db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)


# ---------------------------------------------------
# API: Registrar abono
# ---------------------------------------------------
@app.post("/api/abonos/registrar")
async def registrar_abono(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    data = await request.json()
    
    factura_id = data.get('factura_id')
    monto_usd = data.get('monto_usd')
    forma_pago = data.get('forma_pago')
    referencia = data.get('referencia')
    tasa = data.get('tasa', 36.50)
    
    if not factura_id or not monto_usd:
        return JSONResponse({"error": "Factura y monto son requeridos"}, status_code=400)
    
    try:
        monto_bs = float(monto_usd) * float(tasa)
        
        # Insertar abono
        await db.execute(
            text("""
                INSERT INTO abonos_factura (factura_id, monto_usd, monto_bs, tasa_cambio, forma_pago, referencia, creado_por)
                VALUES (:factura_id, :monto_usd, :monto_bs, :tasa, :forma_pago, :referencia, :creado_por)
            """),
            {
                "factura_id": factura_id,
                "monto_usd": monto_usd,
                "monto_bs": monto_bs,
                "tasa": tasa,
                "forma_pago": forma_pago,
                "referencia": referencia,
                "creado_por": user.id
            }
        )
        
        # Calcular nuevos saldos usando los valores de la factura (no tasa fija)
        result = await db.execute(
            text("""
                SELECT 
                    f.total as total_usd,
                    f.total_bs as total_bs,
                    COALESCE(SUM(a.monto_usd), 0) as total_abonado_usd,
                    COALESCE(SUM(a.monto_bs), 0) as total_abonado_bs
                FROM facturas f
                LEFT JOIN abonos_factura a ON f.id = a.factura_id
                WHERE f.id = :id
                GROUP BY f.id, f.total, f.total_bs
            """),
            {"id": factura_id}
        )
        row = result.fetchone()
        
        total_usd = float(row[0]) if row[0] else 0
        total_bs = float(row[1]) if row[1] else 0
        total_abonado_usd = float(row[2]) if row[2] else 0
        total_abonado_bs = float(row[3]) if row[3] else 0
        
        nuevo_saldo_usd = total_usd - total_abonado_usd
        nuevo_saldo_bs = total_bs - total_abonado_bs
        
        # Determinar nuevo estado
        if nuevo_saldo_usd <= 0.01:
            await db.execute(
                text("UPDATE facturas SET estado = 'pagada', fecha_pago = NOW() WHERE id = :id"),
                {"id": factura_id}
            )
            factura_pagada = True
        else:
            factura_pagada = False
        
        await db.commit()
        
        return JSONResponse({
            "success": True,
            "mensaje": "Abono registrado correctamente",
            "factura_pagada": factura_pagada,
            "nuevo_saldo_usd": nuevo_saldo_usd,
            "nuevo_saldo_bs": nuevo_saldo_bs
        })
    
    except Exception as e:
        await db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)

# ---------------------------------------------------
# API: Listar clientes (para el select)
# ---------------------------------------------------
@app.get("/api/cuentas-cobrar/clientes")
async def api_clientes_cxc(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    result = await db.execute(
        text("""
            SELECT id, nombre_razon_social as nombre, telefono 
            FROM clientes 
            WHERE activo = true 
            ORDER BY nombre_razon_social
        """)
    )
    rows = result.fetchall()
    clientes = [{"id": row[0], "nombre": row[1], "telefono": row[2]} for row in rows]
    
    return JSONResponse(clientes)


# ---------------------------------------------------
# API: Resumen para dashboard
# ---------------------------------------------------
@app.get("/api/cuentas-cobrar/resumen")
async def api_resumen_cxc(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    result = await db.execute(
        text("""
            SELECT 
                COALESCE(SUM(CAST(saldo_pendiente_usd AS DECIMAL(12,2))), 0) as total_pendiente,
                COUNT(DISTINCT CASE WHEN CAST(saldo_pendiente_usd AS DECIMAL(12,2)) > 0 THEN cliente_id END) as clientes_deudores,
                COALESCE(SUM(CASE 
                    WHEN EXTRACT(MONTH FROM created_at) = EXTRACT(MONTH FROM CURRENT_DATE)
                    AND EXTRACT(YEAR FROM created_at) = EXTRACT(YEAR FROM CURRENT_DATE)
                    AND estado = 'pagada' 
                    THEN CAST(monto_total_usd AS DECIMAL(12,2)) ELSE 0 END), 0) as cobrado_mes
            FROM cuentas_cobrar
        """)
    )
    row = result.fetchone()
    
    return JSONResponse({
        "total_pendiente": float(row[0]) if row[0] else 0,
        "clientes_deudores": int(row[1]) if row[1] else 0,
        "cobrado_mes": float(row[2]) if row[2] else 0
    })


# ---------------------------------------------------
# Historial de abonos de una cuenta
# ---------------------------------------------------
@app.get("/api/facturas/{factura_id}/abonos")
async def get_historial_abonos(
    factura_id: int,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    # Obtener datos de la factura
    factura_result = await db.execute(
        text("""
            SELECT f.numero_factura, f.total, f.cliente_id, c.nombre_razon_social
            FROM facturas f
            INNER JOIN clientes c ON f.cliente_id = c.id
            WHERE f.id = :id
        """),
        {"id": factura_id}
    )
    factura = factura_result.fetchone()
    
    if not factura:
        return JSONResponse({"error": "Factura no encontrada"}, status_code=404)
    
    # Obtener abonos
    abonos_result = await db.execute(
        text("""
            SELECT 
                TO_CHAR(a.fecha_abono, 'DD/MM/YYYY') as fecha,
                a.monto_usd,
                a.monto_bs,
                a.forma_pago,
                a.referencia,
                a.tasa_cambio
            FROM abonos_factura a
            WHERE a.factura_id = :id
            ORDER BY a.fecha_abono DESC
        """),
        {"id": factura_id}
    )
    
    abonos = []
    total_abonado_usd = 0
    for row in abonos_result.fetchall():
        monto_usd = float(row[1]) if row[1] else 0
        total_abonado_usd += monto_usd
        abonos.append({
            "fecha": row[0],
            "monto_usd": monto_usd,
            "monto_bs": float(row[2]) if row[2] else 0,
            "forma_pago": row[3],
            "referencia": row[4],
            "tasa": float(row[5]) if row[5] else 0
        })
    
    return JSONResponse({
        "factura_numero": factura[0],
        "total_usd": float(factura[1]) if factura[1] else 0,
        "cliente_nombre": factura[3],
        "total_abonado_usd": total_abonado_usd,
        "saldo_pendiente_usd": (float(factura[1]) if factura[1] else 0) - total_abonado_usd,
        "abonos": abonos
    })

@app.get("/api/facturas/pendientes")
async def get_facturas_pendientes(
    request: Request,
    user: Usuario = Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    
    result = await db.execute(
        text("""
            SELECT 
                f.id,
                f.numero_factura,
                TO_CHAR(f.fecha_emision, 'YYYY-MM-DD') as fecha_emision,
                f.cliente_id,
                c.nombre_razon_social as cliente_nombre,
                f.total as monto_usd,
                f.total_bs as monto_bs,
                f.estado,
                COALESCE((
                    SELECT COALESCE(SUM(a.monto_usd), 0) FROM abonos_factura a WHERE a.factura_id = f.id
                ), 0) as abonado_usd,
                COALESCE((
                    SELECT COALESCE(SUM(a.monto_bs), 0) FROM abonos_factura a WHERE a.factura_id = f.id
                ), 0) as abonado_bs
            FROM facturas f
            INNER JOIN clientes c ON f.cliente_id = c.id
            WHERE f.estado = 'pendiente'
            ORDER BY f.fecha_emision DESC
        """)
    )
    
    facturas = []
    for row in result.fetchall():
        monto_usd = float(row[5]) if row[5] else 0
        monto_bs = float(row[6]) if row[6] else 0
        abonado_usd = float(row[8]) if row[8] else 0
        abonado_bs = float(row[9]) if row[9] else 0
        saldo_usd = monto_usd - abonado_usd
        saldo_bs = monto_bs - abonado_bs  # ← Esto es clave: usar el monto_bs original de la factura
        
        facturas.append({
            "id": row[0],
            "numero_factura": row[1],
            "fecha_emision": row[2],
            "cliente_id": row[3],
            "cliente_nombre": row[4],
            "monto_usd": round(monto_usd, 2),
            "monto_bs": round(monto_bs, 2),
            "abonado_usd": round(abonado_usd, 2),
            "abonado_bs": round(abonado_bs, 2),
            "saldo_usd": round(saldo_usd, 2),
            "saldo_bs": round(saldo_bs, 2),
            "tasa_conversion": round(monto_bs / monto_usd, 4) if monto_usd > 0 else 36.50
        })
    
    return JSONResponse({"facturas": facturas})
    


@app.get("/reportes/mensual", response_class=HTMLResponse)
async def reporte_mensual(
    request: Request,
    mes: int = None,
    anio: int = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    from datetime import datetime, date
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    import calendar
    import json
    
    # Fechas por defecto: mes actual
    hoy = datetime.now()
    if not mes:
        mes = hoy.month
    if not anio:
        anio = hoy.year
    
    # Calcular primer y último día del mes
    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, calendar.monthrange(anio, mes)[1])
    nombre_mes = calendar.month_name[mes]
    
    fecha_ini = primer_dia
    fecha_fin_obj = ultimo_dia
    
    print(f"📅 Consultando del {fecha_ini} al {fecha_fin_obj}")
    
    # ============================================
    # 1. INGRESOS DESDE CIERRES DE CAJA
    # ============================================
    cierres = (await db.execute(
        select(models.CierreCaja).where(
            models.CierreCaja.fecha >= fecha_ini,
            models.CierreCaja.fecha <= fecha_fin_obj
        )
    )).scalars().all()
    
    ingreso_total_usd = 0.0
    tasa_promedio = 0.0
    diferencia_promedio = 0.0
    suma_tasas = 0.0
    suma_diferencias = 0.0
    cantidad_cierres = 0
    ingresos_detalle = []
    
    for cierre in cierres:
        total_ingreso = float(cierre.total_ingresos_usd) if cierre.total_ingresos_usd is not None else 0
        ingreso_total_usd += total_ingreso
        
        tasa = float(cierre.tasa_cambio) if cierre.tasa_cambio is not None else 0
        diferencia = float(cierre.diferencia_usd) if cierre.diferencia_usd is not None else 0
        
        if tasa > 0:
            suma_tasas += tasa
            suma_diferencias += diferencia
            cantidad_cierres += 1
        
        ingresos_detalle.append({
            "fecha": cierre.fecha.isoformat(),
            "descripcion": f"Cierre #{cierre.id}",
            "total_ingreso": round(total_ingreso, 2),
            "tasa": round(tasa, 2),
            "diferencia": round(diferencia, 2)
        })
    
    tasa_promedio = suma_tasas / cantidad_cierres if cantidad_cierres > 0 else 0
    diferencia_promedio = suma_diferencias / cantidad_cierres if cantidad_cierres > 0 else 0
    
    # ============================================
    # 2. GASTOS OPERATIVOS
    # ============================================
    gastos = (await db.execute(
        select(models.GastoDiario).where(
            models.GastoDiario.fecha >= fecha_ini,
            models.GastoDiario.fecha <= fecha_fin_obj
        )
    )).scalars().all()
    
    gasto_total_usd = 0.0
    gastos_detalle = []
    for g in gastos:
        monto = float(g.monto) if g.monto is not None else 0
        monto_usd = monto / 36.50
        gasto_total_usd += monto_usd
        gastos_detalle.append({
            "fecha": g.fecha.isoformat(),
            "descripcion": g.descripcion or "Gasto",
            "monto_usd": round(monto_usd, 2)
        })
    
    # ============================================
    # 3. PAGOS A EMPLEADOS
    # ============================================
    pagos_empleados = (await db.execute(
        select(models.PagoSemanal)
        .where(
            models.PagoSemanal.fecha_pago >= fecha_ini,
            models.PagoSemanal.fecha_pago <= fecha_fin_obj
        )
        .options(selectinload(models.PagoSemanal.empleado))
    )).scalars().all()
    
    pago_empleados_usd = 0.0
    empleados_detalle = []
    
    for pago in pagos_empleados:
        total_neto = float(pago.total_neto) if pago.total_neto is not None else 0
        pago_empleados_usd += total_neto
        
        nombre_empleado = pago.empleado.nombre_completo if pago.empleado else f"Empleado ID {pago.empleado_id}"
        
        empleados_detalle.append({
            "nombre": nombre_empleado,
            "sueldo_fijo": float(pago.sueldo_fijo) if pago.sueldo_fijo is not None else 0,
            "comisiones": float(pago.total_comisiones) if pago.total_comisiones is not None else 0,
            "prestamos": float(pago.total_prestamos) if pago.total_prestamos is not None else 0,
            "deducciones": float(pago.total_deducciones) if pago.total_deducciones is not None else 0,
            "pago_neto": total_neto
        })
    
    # ============================================
    # 4. COMPRAS DE MATERIALES
    # ============================================
    compras_detalle = []
    total_compras_usd = 0.0
    
    try:
        entradas = (await db.execute(
            select(models.EntradaInventario)
            .where(
                models.EntradaInventario.fecha_entrada >= fecha_ini,
                models.EntradaInventario.fecha_entrada <= fecha_fin_obj
            )
            .options(selectinload(models.EntradaInventario.material))
        )).scalars().all()
        
        for entrada in entradas:
            if entrada.material:
                precio = float(entrada.precio_unitario) if entrada.precio_unitario is not None else 0
                cantidad = float(entrada.cantidad) if entrada.cantidad is not None else 0
                monto_usd = precio * cantidad
                total_compras_usd += monto_usd
                compras_detalle.append({
                    "nombre": entrada.material.nombre,
                    "cantidad": cantidad,
                    "unidad": entrada.material.unidad_medida,
                    "precio_unitario": precio,
                    "costo_total": round(monto_usd, 2),
                    "fecha": entrada.fecha_entrada.isoformat(),
                    "proveedor": entrada.proveedor_nombre or "N/A"
                })
    except Exception as e:
        print(f"Error en compras de materiales: {e}")
    
    # ============================================
    # 5. MATERIALES UTILIZADOS
    # ============================================
    materiales_detalle = []
    total_materiales_usd = 0.0
    
    try:
        salidas = (await db.execute(
            select(models.SalidaMaterial)
            .where(
                models.SalidaMaterial.fecha_salida >= fecha_ini,
                models.SalidaMaterial.fecha_salida <= fecha_fin_obj
            )
        )).scalars().all()
        
        for s in salidas:
            material = await db.execute(
                select(models.MaterialInventario).where(models.MaterialInventario.id == s.material_id)
            )
            material_obj = material.scalar_one_or_none()
            
            if material_obj:
                precio = float(material_obj.precio_compra) if material_obj.precio_compra is not None else 0
                cantidad = float(s.cantidad) if s.cantidad is not None else 0
                monto_usd = precio * cantidad
                total_materiales_usd += monto_usd
                materiales_detalle.append({
                    "nombre": material_obj.nombre,
                    "cantidad": cantidad,
                    "unidad": material_obj.unidad_medida,
                    "costo_total": round(monto_usd, 2),
                    "motivo": s.motivo or "Uso en trabajo",
                    "fecha": s.fecha_salida.isoformat()
                })
    except Exception as e:
        print(f"Error en salidas de materiales: {e}")
    
    # ============================================
    # 6. ACTIVOS FIJOS (Depreciación Mensual)
    # ============================================
    activos_detalle = []
    total_depreciacion_usd = 0.0
    
    try:
        activos = (await db.execute(
            select(models.ActivoFijo).where(models.ActivoFijo.estado == "activo")
        )).scalars().all()
        
        for a in activos:
            vida_util_anios = 5
            if a.categoria_id:
                cat_result = await db.execute(
                    select(models.CategoriaActivoFijo.vida_util_anios)
                    .where(models.CategoriaActivoFijo.id == a.categoria_id)
                )
                vida = cat_result.scalar()
                if vida:
                    vida_util_anios = float(vida)
            
            costo_inicial = float(a.costo_inicial) if a.costo_inicial is not None else 0
            valor_residual = float(a.valor_residual) if a.valor_residual is not None else 0
            
            if vida_util_anios > 0 and costo_inicial > 0:
                depreciacion_anual = (costo_inicial - valor_residual) / vida_util_anios
                depreciacion_mensual = depreciacion_anual / 12
                total_depreciacion_usd += depreciacion_mensual
                
                activos_detalle.append({
                    "nombre": a.nombre,
                    "valor_original": round(costo_inicial, 2),
                    "depreciacion_mensual": round(depreciacion_mensual, 2),
                    "valor_actual": round(costo_inicial - depreciacion_mensual, 2)
                })
    except Exception as e:
        print(f"Error en activos fijos: {e}")
    
    # ============================================
    # 7. CÁLCULOS FINALES
    # ============================================
    total_egresos_usd = gasto_total_usd + pago_empleados_usd + total_compras_usd + total_materiales_usd + total_depreciacion_usd
    saldo_final_usd = ingreso_total_usd - total_egresos_usd
    margen_utilidad = (saldo_final_usd / ingreso_total_usd * 100) if ingreso_total_usd > 0 else 0
    
    # ============================================
    # 8. DATOS PARA EL TEMPLATE
    # ============================================
    datos = {
        "periodo": {"inicio": fecha_ini.isoformat(), "fin": fecha_fin_obj.isoformat()},
        "resumen": {
            "ingreso_total_usd": round(ingreso_total_usd, 2),
            "gasto_total_usd": round(gasto_total_usd, 2),
            "pago_empleados_usd": round(pago_empleados_usd, 2),
            "total_compras_usd": round(total_compras_usd, 2),
            "total_materiales_usd": round(total_materiales_usd, 2),
            "total_depreciacion_usd": round(total_depreciacion_usd, 2),
            "saldo_final_usd": round(saldo_final_usd, 2),
            "margen_utilidad": round(margen_utilidad, 2),
            "tasa_promedio": round(tasa_promedio, 2),
            "diferencia_promedio": round(diferencia_promedio, 2)
        },
        "ingresos_detalle": ingresos_detalle,
        "gastos_detalle": gastos_detalle,
        "empleados_detalle": empleados_detalle,
        "compras_detalle": compras_detalle,
        "materiales_detalle": materiales_detalle,
        "activos_detalle": activos_detalle
    }
    
    return templates.TemplateResponse("reportes/mensual.html", {
        "request": request,
        "datos": datos,
        "mes": mes,
        "anio": anio,
        "nombre_mes": nombre_mes
    })

# ============================================
# REPORTE TRIMESTRAL
# ============================================
@app.get("/reportes/trimestral", response_class=HTMLResponse)
async def reporte_trimestral(
    request: Request,
    trimestre: int = None,
    anio: int = None,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login")
    
    from datetime import datetime, date
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    from collections import defaultdict
    import calendar
    import json
    
    # Fechas por defecto: trimestre actual
    hoy = datetime.now()
    if not anio:
        anio = hoy.year
    if not trimestre:
        trimestre = (hoy.month - 1) // 3 + 1
    
    # Calcular fechas del trimestre
    if trimestre == 1:
        primer_dia = date(anio, 1, 1)
        ultimo_dia = date(anio, 3, calendar.monthrange(anio, 3)[1])
        nombre_trimestre = "Primer Trimestre (Ene - Mar)"
    elif trimestre == 2:
        primer_dia = date(anio, 4, 1)
        ultimo_dia = date(anio, 6, calendar.monthrange(anio, 6)[1])
        nombre_trimestre = "Segundo Trimestre (Abr - Jun)"
    elif trimestre == 3:
        primer_dia = date(anio, 7, 1)
        ultimo_dia = date(anio, 9, calendar.monthrange(anio, 9)[1])
        nombre_trimestre = "Tercer Trimestre (Jul - Sep)"
    else:
        primer_dia = date(anio, 10, 1)
        ultimo_dia = date(anio, 12, calendar.monthrange(anio, 12)[1])
        nombre_trimestre = "Cuarto Trimestre (Oct - Dic)"
    
    fecha_ini = primer_dia
    fecha_fin_obj = ultimo_dia
    
    print(f"📅 Consultando del {fecha_ini} al {fecha_fin_obj}")
    
    try:
        # ============================================
        # 1. INGRESOS DESDE CIERRES DE CAJA
        # ============================================
        cierres = (await db.execute(
            select(models.CierreCaja).where(
                models.CierreCaja.fecha >= fecha_ini,
                models.CierreCaja.fecha <= fecha_fin_obj
            )
        )).scalars().all()
        
        ingreso_total_usd = 0.0
        tasa_promedio = 0.0
        diferencia_promedio = 0.0
        suma_tasas = 0.0
        suma_diferencias = 0.0
        cantidad_cierres = 0
        ingresos_detalle = []
        
        for cierre in cierres:
            total_ingreso = float(cierre.total_ingresos_usd) if cierre.total_ingresos_usd is not None else 0
            ingreso_total_usd += total_ingreso
            
            tasa = float(cierre.tasa_cambio) if cierre.tasa_cambio is not None else 0
            diferencia = float(cierre.diferencia_usd) if cierre.diferencia_usd is not None else 0
            
            if tasa > 0:
                suma_tasas += tasa
                suma_diferencias += diferencia
                cantidad_cierres += 1
            
            ingresos_detalle.append({
                "fecha": cierre.fecha.isoformat(),
                "descripcion": f"Cierre #{cierre.id}",
                "total_ingreso": round(total_ingreso, 2),
                "tasa": round(tasa, 2),
                "diferencia": round(diferencia, 2)
            })
        
        tasa_promedio = suma_tasas / cantidad_cierres if cantidad_cierres > 0 else 0
        diferencia_promedio = suma_diferencias / cantidad_cierres if cantidad_cierres > 0 else 0
        
        # ============================================
        # 2. GASTOS OPERATIVOS
        # ============================================
        gastos = (await db.execute(
            select(models.GastoDiario).where(
                models.GastoDiario.fecha >= fecha_ini,
                models.GastoDiario.fecha <= fecha_fin_obj
            )
        )).scalars().all()
        
        gasto_total_usd = 0.0
        gastos_detalle = []
        for g in gastos:
            monto = float(g.monto) if g.monto is not None else 0
            monto_usd = monto / 36.50
            gasto_total_usd += monto_usd
            gastos_detalle.append({
                "fecha": g.fecha.isoformat(),
                "descripcion": g.descripcion or "Gasto",
                "monto_usd": round(monto_usd, 2)
            })
        
        # ============================================
        # 3. PAGOS A EMPLEADOS (CALCULADO)
        # ============================================
        empleados = (await db.execute(
            select(models.Empleado).where(models.Empleado.activo == True)
        )).scalars().all()
        
        trabajos = (await db.execute(
            select(models.Trabajo).where(
                models.Trabajo.estado == "completado",
                models.Trabajo.fecha_inicio >= fecha_ini,
                models.Trabajo.fecha_inicio <= fecha_fin_obj
            )
        )).scalars().all()
        
        comisiones_por_empleado = defaultdict(float)
        
        for trabajo in trabajos:
            asignaciones = (await db.execute(
                select(models.Asignacion)
                .where(models.Asignacion.trabajo_id == trabajo.id)
                .options(selectinload(models.Asignacion.empleado))
            )).scalars().all()
            
            monto_total = float(trabajo.monto_total or 0)
            
            asignaciones_por_rol = defaultdict(list)
            for asign in asignaciones:
                asignaciones_por_rol[asign.rol_id].append(asign)
            
            for rol_id, asignaciones_rol in asignaciones_por_rol.items():
                if not asignaciones_rol:
                    continue
                
                primera_asign = asignaciones_rol[0]
                tipo_comision = primera_asign.tipo_comision
                valor_comision = float(primera_asign.valor_comision or 0)
                
                comision_total_rol = 0.0
                if tipo_comision == "porcentaje":
                    comision_total_rol = monto_total * (valor_comision / 100)
                elif tipo_comision == "fijo":
                    comision_total_rol = valor_comision
                elif tipo_comision == "por_m2":
                    metros = float(trabajo.metros_cuadrados or 0)
                    comision_total_rol = metros * valor_comision
                elif tipo_comision == "por_unidad":
                    unidades = int(trabajo.unidades or 0)
                    comision_total_rol = unidades * valor_comision
                
                comision_por_empleado = comision_total_rol / len(asignaciones_rol)
                
                for asign in asignaciones_rol:
                    if asign.empleado:
                        comisiones_por_empleado[asign.empleado.id] += comision_por_empleado
        
        prestamos = (await db.execute(
            select(models.Prestamo).where(models.Prestamo.pagado == False)
        )).scalars().all()
        
        prestamos_por_empleado = defaultdict(float)
        for p in prestamos:
            if p.empleado_id:
                prestamos_por_empleado[p.empleado_id] += float(p.monto or 0)
        
        deducciones = (await db.execute(
            select(models.DeduccionEmpleado).where(
                models.DeduccionEmpleado.fecha_deduccion >= fecha_ini,
                models.DeduccionEmpleado.fecha_deduccion <= fecha_fin_obj,
                models.DeduccionEmpleado.estado == "pendiente"
            )
        )).scalars().all()
        
        deducciones_por_empleado = defaultdict(float)
        for d in deducciones:
            if d.empleado_id:
                deducciones_por_empleado[d.empleado_id] += float(d.monto or 0)
        
        pago_empleados_usd = 0.0
        empleados_detalle = []
        
        for emp in empleados:
            if emp.tipo_contrato in ["fijo", "mixto"]:
                sueldo_fijo = float(emp.salario_fijo or 0)
            else:
                sueldo_fijo = 0.0
            
            comisiones = comisiones_por_empleado.get(emp.id, 0.0)
            prestamos_emp = prestamos_por_empleado.get(emp.id, 0.0)
            deducciones_emp = deducciones_por_empleado.get(emp.id, 0.0)
            
            pago_neto = sueldo_fijo + comisiones - prestamos_emp - deducciones_emp
            
            if pago_neto != 0 or sueldo_fijo != 0 or comisiones != 0:
                pago_empleados_usd += pago_neto
                empleados_detalle.append({
                    "nombre": emp.nombre_completo,
                    "sueldo_fijo": round(sueldo_fijo, 2),
                    "comisiones": round(comisiones, 2),
                    "prestamos": round(prestamos_emp, 2),
                    "deducciones": round(deducciones_emp, 2),
                    "pago_neto": round(pago_neto, 2)
                })
        
        # ============================================
        # 4. COMPRAS DE MATERIALES
        # ============================================
        compras_detalle = []
        total_compras_usd = 0.0
        
        try:
            entradas = (await db.execute(
                select(models.EntradaInventario)
                .where(
                    models.EntradaInventario.fecha_entrada >= fecha_ini,
                    models.EntradaInventario.fecha_entrada <= fecha_fin_obj
                )
                .options(selectinload(models.EntradaInventario.material))
            )).scalars().all()
            
            for entrada in entradas:
                if entrada.material:
                    precio = float(entrada.precio_unitario) if entrada.precio_unitario is not None else 0
                    cantidad = float(entrada.cantidad) if entrada.cantidad is not None else 0
                    monto_usd = precio * cantidad
                    total_compras_usd += monto_usd
                    compras_detalle.append({
                        "nombre": entrada.material.nombre,
                        "cantidad": cantidad,
                        "unidad": entrada.material.unidad_medida,
                        "precio_unitario": precio,
                        "costo_total": round(monto_usd, 2),
                        "fecha": entrada.fecha_entrada.isoformat(),
                        "proveedor": entrada.proveedor_nombre or "N/A"
                    })
        except Exception as e:
            print(f"Error en compras: {e}")
        
        # ============================================
        # 5. MATERIALES UTILIZADOS
        # ============================================
        materiales_detalle = []
        total_materiales_usd = 0.0
        
        try:
            salidas = (await db.execute(
                select(models.SalidaMaterial)
                .where(
                    models.SalidaMaterial.fecha_salida >= fecha_ini,
                    models.SalidaMaterial.fecha_salida <= fecha_fin_obj
                )
            )).scalars().all()
            
            for s in salidas:
                material = await db.execute(
                    select(models.MaterialInventario).where(models.MaterialInventario.id == s.material_id)
                )
                material_obj = material.scalar_one_or_none()
                
                if material_obj:
                    precio = float(material_obj.precio_compra) if material_obj.precio_compra is not None else 0
                    cantidad = float(s.cantidad) if s.cantidad is not None else 0
                    monto_usd = precio * cantidad
                    total_materiales_usd += monto_usd
                    materiales_detalle.append({
                        "nombre": material_obj.nombre,
                        "cantidad": cantidad,
                        "unidad": material_obj.unidad_medida,
                        "costo_total": round(monto_usd, 2),
                        "motivo": s.motivo or "Uso en trabajo",
                        "fecha": s.fecha_salida.isoformat()
                    })
        except Exception as e:
            print(f"Error en materiales: {e}")
        
        # ============================================
        # 6. ACTIVOS FIJOS (Depreciación Trimestral)
        # ============================================
        activos_detalle = []
        total_depreciacion_usd = 0.0
        
        try:
            activos = (await db.execute(
                select(models.ActivoFijo).where(models.ActivoFijo.estado == "activo")
            )).scalars().all()
            
            for a in activos:
                vida_util_anios = 5
                if a.categoria_id:
                    cat_result = await db.execute(
                        select(models.CategoriaActivoFijo.vida_util_anios)
                        .where(models.CategoriaActivoFijo.id == a.categoria_id)
                    )
                    vida = cat_result.scalar()
                    if vida:
                        vida_util_anios = float(vida)
                
                costo_inicial = float(a.costo_inicial) if a.costo_inicial is not None else 0
                valor_residual = float(a.valor_residual) if a.valor_residual is not None else 0
                
                if vida_util_anios > 0 and costo_inicial > 0:
                    depreciacion_anual = (costo_inicial - valor_residual) / vida_util_anios
                    depreciacion_trimestral = depreciacion_anual / 4
                    total_depreciacion_usd += depreciacion_trimestral
                    
                    activos_detalle.append({
                        "nombre": a.nombre,
                        "valor_original": round(costo_inicial, 2),
                        "depreciacion_trimestral": round(depreciacion_trimestral, 2),
                        "valor_actual": round(costo_inicial - depreciacion_trimestral, 2)
                    })
        except Exception as e:
            print(f"Error en activos: {e}")
        
        # ============================================
        # 7. CÁLCULOS FINALES
        # ============================================
        total_egresos_usd = gasto_total_usd + pago_empleados_usd + total_compras_usd + total_materiales_usd + total_depreciacion_usd
        saldo_final_usd = ingreso_total_usd - total_egresos_usd
        margen_utilidad = (saldo_final_usd / ingreso_total_usd * 100) if ingreso_total_usd > 0 else 0
        
        # ============================================
        # 8. DATOS PARA EL TEMPLATE
        # ============================================
        datos = {
            "periodo": {"inicio": fecha_ini.isoformat(), "fin": fecha_fin_obj.isoformat()},
            "resumen": {
                "ingreso_total_usd": round(ingreso_total_usd, 2),
                "gasto_total_usd": round(gasto_total_usd, 2),
                "pago_empleados_usd": round(pago_empleados_usd, 2),
                "total_compras_usd": round(total_compras_usd, 2),
                "total_materiales_usd": round(total_materiales_usd, 2),
                "total_depreciacion_usd": round(total_depreciacion_usd, 2),
                "saldo_final_usd": round(saldo_final_usd, 2),
                "margen_utilidad": round(margen_utilidad, 2),
                "tasa_promedio": round(tasa_promedio, 2),
                "diferencia_promedio": round(diferencia_promedio, 2)
            },
            "ingresos_detalle": ingresos_detalle,
            "gastos_detalle": gastos_detalle,
            "empleados_detalle": empleados_detalle,
            "compras_detalle": compras_detalle,
            "materiales_detalle": materiales_detalle,
            "activos_detalle": activos_detalle
        }
        
        return templates.TemplateResponse("reportes/trimestral.html", {
            "request": request,
            "datos": datos,
            "trimestre": trimestre,
            "anio": anio,
            "nombre_trimestre": nombre_trimestre
        })
        
    except Exception as e:
        print(f"❌ Error en reporte trimestral: {e}")
        import traceback
        traceback.print_exc()
        
        datos_vacios = {
            "periodo": {"inicio": fecha_ini.isoformat(), "fin": fecha_fin_obj.isoformat()},
            "resumen": {
                "ingreso_total_usd": 0, "gasto_total_usd": 0, "pago_empleados_usd": 0,
                "total_compras_usd": 0, "total_materiales_usd": 0, "total_depreciacion_usd": 0,
                "saldo_final_usd": 0, "margen_utilidad": 0, "tasa_promedio": 0, "diferencia_promedio": 0
            },
            "ingresos_detalle": [], "gastos_detalle": [], 
            "empleados_detalle": [], "compras_detalle": [],
            "materiales_detalle": [], "activos_detalle": []
        }
        
        return templates.TemplateResponse("reportes/trimestral.html", {
            "request": request,
            "datos": datos_vacios,
            "trimestre": trimestre,
            "anio": anio,
            "nombre_trimestre": nombre_trimestre
        })
@app.get("/compras/empleados", response_class=HTMLResponse)
async def todas_las_compras(
    request: Request,
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # 🔥 DEPURACIÓN
    print("=== CARGANDO COMPRAS ===")
    
    # Obtener todas las compras
    result = await db.execute(
        select(CompraEmpleado)
        .order_by(CompraEmpleado.fecha_compra.desc())
    )
    compras = result.scalars().all()
    
    # 🔥 PRINT para ver cuántas compras hay
    print(f"📌 Compras encontradas: {len(compras)}")
    
    # Cargar relaciones
    for compra in compras:
        print(f"   - Compra ID: {compra.id}, Empleado ID: {compra.empleado_id}")
        if compra.empleado_id:
            emp_result = await db.execute(select(Empleado).where(Empleado.id == compra.empleado_id))
            compra.empleado = emp_result.scalar_one_or_none()
            print(f"     Empleado: {compra.empleado.nombre_completo if compra.empleado else 'No encontrado'}")
        if compra.trabajo_id:
            trabajo_result = await db.execute(select(Trabajo).where(Trabajo.id == compra.trabajo_id))
            compra.trabajo = trabajo_result.scalar_one_or_none()
    
    # Estadísticas
    pendientes = sum(1 for c in compras if c.estado_pago == 'pendiente')
    descontados = sum(1 for c in compras if c.estado_pago == 'descontado')
    total_adeudado = sum(float(c.total_a_descontar) for c in compras if c.estado_pago == 'pendiente')
    
    print(f"📌 Pendientes: {pendientes}, Descontados: {descontados}, Adeudado: {total_adeudado}")
    print("=== FIN ===")
    
    return templates.TemplateResponse("empleados/compras/lista_general.html", {
        "request": request,
        "user": user,
        "compras": compras,
        "pendientes": pendientes,
        "descontados": descontados,
        "total_adeudado": total_adeudado
    })

# main.py - Agrega estos endpoints al final

# === ENDPOINTS PARA CUENTAS POR COBRAR ===

@app.get("/api/v1/clientes/activos")
async def get_clientes_activos(db: Session = Depends(get_db)):
    """Obtener clientes activos"""
    try:
        clientes = db.query(Cliente).filter(Cliente.activo == True).all()
        return [
            {
                "id": c.id,
                "nombre_razon_social": c.nombre_razon_social,
                "nombre": c.nombre_razon_social
            }
            for c in clientes
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/deudas/pendientes")
async def get_deudas_pendientes(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    search: str = None,
    origen: str = None
):
    """Obtener deudas pendientes"""
    try:
        from sqlalchemy import literal, func
        
        # 1. Trabajos con saldo pendiente (0% o 50%)
        trabajos_query = db.query(
            literal('trabajo').label('origen'),
            Trabajo.id.label('trabajo_id'),
            literal(None).label('deuda_manual_id'),
            Trabajo.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            Trabajo.nombre_trabajo.label('concepto'),
            Trabajo.monto_total.label('monto_total'),
            func.coalesce(Trabajo.monto_pagado_usd, 0).label('monto_pagado'),
            (Trabajo.monto_total - func.coalesce(Trabajo.monto_pagado_usd, 0)).label('saldo_pendiente'),
            Trabajo.fecha_creacion.label('fecha'),
            Trabajo.porcentaje_pagado.label('porcentaje_pagado'),
            literal(None).label('estado_deuda')
        ).join(Cliente, Trabajo.cliente_id == Cliente.id).filter(
            Trabajo.porcentaje_pagado.in_([0, 50])
        )
        
        # 2. Deudas manuales
        deudas_manuales_query = db.query(
            literal('manual').label('origen'),
            literal(None).label('trabajo_id'),
            DeudaManual.id.label('deuda_manual_id'),
            DeudaManual.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            DeudaManual.nombre_trabajo.label('concepto'),
            DeudaManual.monto_deuda.label('monto_total'),
            DeudaManual.monto_pagado.label('monto_pagado'),
            DeudaManual.saldo_pendiente.label('saldo_pendiente'),
            DeudaManual.fecha_registro.label('fecha'),
            literal(None).label('porcentaje_pagado'),
            DeudaManual.estado.label('estado_deuda')
        ).join(Cliente, DeudaManual.cliente_id == Cliente.id).filter(
            DeudaManual.estado.in_(['pendiente', 'parcial'])
        )
        
        # Aplicar filtros
        if search:
            search_filter = f"%{search}%"
            trabajos_query = trabajos_query.filter(
                Cliente.nombre_razon_social.ilike(search_filter)
            )
            deudas_manuales_query = deudas_manuales_query.filter(
                Cliente.nombre_razon_social.ilike(search_filter)
            )
        
        if origen == 'trabajo':
            deudas_manuales_query = deudas_manuales_query.filter(False)
        elif origen == 'manual':
            trabajos_query = trabajos_query.filter(False)
        
        # Unir y ordenar
        resultados = trabajos_query.union_all(deudas_manuales_query).order_by(
            'cliente_nombre', 'fecha'
        ).offset(skip).limit(limit).all()
        
        # Formatear respuesta
        deudores = []
        for r in resultados:
            deudores.append({
                "cliente_id": r.cliente_id,
                "cliente_nombre": r.cliente_nombre,
                "concepto": r.concepto,
                "monto_total": float(r.monto_total or 0),
                "monto_pagado": float(r.monto_pagado or 0),
                "saldo_pendiente": float(r.saldo_pendiente or 0),
                "fecha": r.fecha.isoformat() if r.fecha else None,
                "origen": r.origen,
                "trabajo_id": r.trabajo_id,
                "deuda_manual_id": r.deuda_manual_id,
                "porcentaje_pagado": r.porcentaje_pagado,
                "estado_deuda": r.estado_deuda or ('pagado' if r.porcentaje_pagado == 100 else 'pendiente')
            })
        
        # Estadísticas
        total_trabajos = trabajos_query.count()
        total_deudas = deudas_manuales_query.count()
        
        return {
            "total_registros": total_trabajos + total_deudas,
            "total_clientes_unicos": len(set([d["cliente_id"] for d in deudores])),
            "total_adeudado": sum([d["saldo_pendiente"] for d in deudores]),
            "deudores": deudores,
            "resumen_por_origen": {
                "trabajos": total_trabajos,
                "manuales": total_deudas
            }
        }
        
    except Exception as e:
        print(f"Error en get_deudas_pendientes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v1/deudas/manual/")
async def crear_deuda_manual_endpoint(
    deuda_data: dict,
    db: Session = Depends(get_db)
):
    """Crear deuda manual"""
    try:
        from datetime import datetime
        
        cliente_id = deuda_data.get("cliente_id")
        nombre_trabajo = deuda_data.get("nombre_trabajo")
        monto_deuda = deuda_data.get("monto_deuda")
        observaciones = deuda_data.get("observaciones")
        
        # Validar cliente
        cliente = db.query(Cliente).filter(Cliente.id == cliente_id, Cliente.activo == True).first()
        if not cliente:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
        # Crear deuda
        nueva_deuda = DeudaManual(
            cliente_id=cliente_id,
            nombre_trabajo=nombre_trabajo,
            monto_deuda=monto_deuda,
            saldo_pendiente=monto_deuda,
            monto_pagado=0.0,
            estado="pendiente",
            observaciones=observaciones,
            fecha_registro=datetime.utcnow()
        )
        
        db.add(nueva_deuda)
        db.commit()
        db.refresh(nueva_deuda)
        
        return {
            "id": nueva_deuda.id,
            "cliente_id": nueva_deuda.cliente_id,
            "nombre_trabajo": nueva_deuda.nombre_trabajo,
            "monto_deuda": float(nueva_deuda.monto_deuda),
            "saldo_pendiente": float(nueva_deuda.saldo_pendiente),
            "monto_pagado": float(nueva_deuda.monto_pagado),
            "estado": nueva_deuda.estado,
            "fecha_registro": nueva_deuda.fecha_registro.isoformat(),
            "observaciones": nueva_deuda.observaciones,
            "es_convertida_a_trabajo": False,
            "trabajo_convertido_id": None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/v1/deudas/manual/{deuda_id}/pagar/")
async def pagar_deuda_manual(
    deuda_id: int,
    pago_data: dict,
    db: Session = Depends(get_db)
):
    """Registrar pago de deuda manual"""
    try:
        monto_pago = pago_data.get("monto_pago")
        
        if not monto_pago or monto_pago <= 0:
            raise HTTPException(status_code=400, detail="Monto de pago inválido")
        
        deuda = db.query(DeudaManual).filter(DeudaManual.id == deuda_id).first()
        if not deuda:
            raise HTTPException(status_code=404, detail="Deuda no encontrada")
        
        if deuda.estado == 'pagado':
            raise HTTPException(status_code=400, detail="Deuda ya pagada")
        
        if monto_pago > deuda.saldo_pendiente:
            raise HTTPException(status_code=400, detail="Monto excede el saldo pendiente")
        
        deuda.monto_pagado += monto_pago
        deuda.saldo_pendiente -= monto_pago
        
        if deuda.saldo_pendiente <= 0:
            deuda.estado = 'pagado'
        elif deuda.monto_pagado > 0:
            deuda.estado = 'parcial'
        
        deuda.fecha_actualizacion = datetime.utcnow()
        
        db.commit()
        db.refresh(deuda)
        
        return {
            "id": deuda.id,
            "cliente_id": deuda.cliente_id,
            "nombre_trabajo": deuda.nombre_trabajo,
            "monto_deuda": float(deuda.monto_deuda),
            "saldo_pendiente": float(deuda.saldo_pendiente),
            "monto_pagado": float(deuda.monto_pagado),
            "estado": deuda.estado,
            "fecha_registro": deuda.fecha_registro.isoformat(),
            "observaciones": deuda.observaciones,
            "es_convertida_a_trabajo": False,
            "trabajo_convertido_id": None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# main.py - Agregar después de tus otros endpoints

@app.post("/asistencia/marcar-falta", response_class=RedirectResponse)
async def marcar_falta(
    request: Request,
    empleado_id: int = Form(...),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Verificar que el empleado existe
    empleado = await db.get(Empleado, empleado_id)
    if not empleado:
        return RedirectResponse(
            url="/empleados?error=Empleado+no+encontrado", 
            status_code=303
        )
    
    hoy = date.today()
    
    # Verificar si ya existe registro para hoy
    asistencia = (await db.execute(
        select(Asistencia).where(
            Asistencia.empleado_id == empleado_id,
            Asistencia.fecha == hoy
        )
    )).scalar_one_or_none()
    
    if asistencia:
        # Si ya existe, actualizar a ausente
        asistencia.estado = "ausente"
    else:
        # Crear nuevo registro
        nueva_asistencia = Asistencia(
            empleado_id=empleado_id,
            fecha=hoy,
            estado="ausente"
        )
        db.add(nueva_asistencia)
    
    await db.commit()
    
    return RedirectResponse(
        url=f"/empleados?falta=1", 
        status_code=303
    )


@app.post("/asistencia/desmarcar-falta", response_class=RedirectResponse)
async def desmarcar_falta(
    request: Request,
    empleado_id: int = Form(...),
    user=Depends(get_current_user_from_session),
    db: AsyncSession = Depends(get_db)
):
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    hoy = date.today()
    
    # Buscar y eliminar el registro de ausencia
    asistencia = (await db.execute(
        select(Asistencia).where(
            Asistencia.empleado_id == empleado_id,
            Asistencia.fecha == hoy,
            Asistencia.estado == "ausente"
        )
    )).scalar_one_or_none()
    
    if asistencia:
        await db.delete(asistencia)
        await db.commit()
        mensaje = "desmarcar=1"
    else:
        mensaje = "error=No+hay+falta+registrada+para+hoy"
    
    return RedirectResponse(
        url=f"/empleados?{mensaje}", 
        status_code=303
    )

@app.middleware("http")
async def check_installation(request: Request, call_next):
    # Obtener la ruta
    path = request.url.path
    
    # Excluir rutas que no necesitan verificación
    if path.startswith("/static") or path.startswith("/install") or path.startswith("/login"):
        return await call_next(request)
    
    # Verificar si hay usuarios en la base de datos
    try:
        from app.database import async_session_maker
        from app.models import Usuario
        from sqlalchemy import select
        
        async with async_session_maker() as db:
            result = await db.execute(select(Usuario).limit(1))
            usuario = result.scalar_one_or_none()
            
            if not usuario:
                # Redirigir a instalación
                return RedirectResponse(url="/install", status_code=303)
    except Exception:
        # Si hay error de conexión, continuar
        pass
    
    return await call_next(request)

# main.py - Al final
from sqlalchemy import text

@app.get("/health")
async def health_check():
    try:
        # Verificar base de datos
        async with AsyncSessionLocal() as session:
            await session.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "ok", "timestamp": datetime.now().isoformat()}
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=503, detail="Service unavailable")

# ============================================
# HEALTH CHECK - PARA MANTENERLO DESPIERTO
# ============================================
@app.get("/health")
async def health_check():
    from datetime import datetime
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "database": "ok"
    }

# ============================================
# STATUS CHECK - PARA MONITOREO COMPLETO
# ============================================
@app.get("/status")
async def status_check():
    """Endpoint para verificar el estado completo del sistema"""
    from datetime import datetime
    import os
    
    status_data = {
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "environment": os.getenv("ENVIRONMENT", "development"),
        "database": "no verificada",
        "services": {
            "api": "running",
            "templates": "loaded"
        }
    }
    
    # Verificar base de datos
    try:
        from app.database import AsyncSessionLocal
        from sqlalchemy import text
        async with AsyncSessionLocal() as session:
            await session.execute(text("SELECT 1"))
        status_data["database"] = "conectada"
    except Exception as e:
        status_data["database"] = f"error: {str(e)}"
        status_data["status"] = "database_error"
    
    # Verificar que las rutas principales existen
    try:
        # Verificar que los routers están cargados
        routes = [route.path for route in app.routes]
        status_data["services"]["routes"] = len(routes)
    except Exception as e:
        status_data["services"]["routes_error"] = str(e)
    
    # Información del sistema
    status_data["system"] = {
        "python_version": os.sys.version[:50] if hasattr(os, 'sys') else "desconocido",
        "database_url": "configurada" if os.getenv("DATABASE_URL") else "no configurada",
        "secret_key": "configurada" if os.getenv("SECRET_KEY") else "no configurada"
    }
    
    return status_data


@app.get("/ping")
async def ping():
    return {"status": "pong", "timestamp": datetime.now().isoformat()}


# ============================================
# LOGS DE ERRORES - PARA DIAGNÓSTICO
# ============================================
import logging
from io import StringIO

# Configurar logging para capturar errores
log_stream = StringIO()
handler = logging.StreamHandler(log_stream)
handler.setLevel(logging.ERROR)
logging.getLogger().addHandler(handler)

@app.get("/logs")
async def ver_logs(user: Usuario = Depends(get_current_user_from_session)):
    if not user or user.rol != "administrador":
        return RedirectResponse(url="/login", status_code=303)
    
    # Obtener los últimos logs
    log_content = log_stream.getvalue()
    if not log_content:
        return HTMLResponse("<h1>📋 Logs del Sistema</h1><p>No hay errores registrados.</p>")
    
    # Mostrar logs en formato HTML
    return HTMLResponse(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Logs del Sistema</title>
        <style>
            body {{ font-family: monospace; padding: 20px; background: #1e1e1e; color: #d4d4d4; }}
            h1 {{ color: #4fc3f7; }}
            .log {{ background: #2d2d2d; padding: 10px; border-radius: 5px; white-space: pre-wrap; font-size: 12px; }}
            .error {{ color: #f44336; }}
            .warning {{ color: #ff9800; }}
            .info {{ color: #4fc3f7; }}
        </style>
    </head>
    <body>
        <h1>📋 Logs del Sistema</h1>
        <div class="log">{log_content}</div>
        <br>
        <a href="/dashboard" style="color: #4fc3f7;">← Volver al Dashboard</a>
    </body>
    </html>
    """)
 
  
   # En tu archivo de utilidades
from datetime import date
import requests


    
# Asegúrate de tener estos imports:
from sqlalchemy import delete
from collections import defaultdict

from app.api import router
app.include_router(router, prefix="/api")

@app.get("/", response_class=HTMLResponse)
async def root():
    return RedirectResponse(url="/login")

# Al final de tu main.py
@app.on_event("startup")
async def startup_event():
    async with async_engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


