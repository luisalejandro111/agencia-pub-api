# app/routers/deudas.py
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, literal, and_, or_
from typing import List, Optional
from decimal import Decimal
from datetime import datetime

from app.database import get_db
from app.models import Cliente, Trabajo, DeudaManual

router = APIRouter(prefix="/deudas", tags=["Cuentas por Cobrar"])


# ============================================
# ENDPOINT SIMPLIFICADO - GET /pendientes
# ============================================
@router.get("/pendientes")
async def listar_deudores_pendientes(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(100, description="Límite de registros"),
    search: Optional[str] = Query(None, description="Búsqueda por cliente"),
    origen: Optional[str] = Query(None, description="Filtrar por origen: 'trabajo' o 'manual'")
):
    """Obtener listado de deudores pendientes - VERSIÓN SIMPLIFICADA"""
    try:
        print("📋 GET /deudas/pendientes - Iniciando consulta...")
        
        # 1. Obtener trabajos con saldo pendiente (0% o 50%)
        trabajos_query = select(
            literal('trabajo').label('origen'),
            Trabajo.id.label('trabajo_id'),
            literal(None).label('deuda_manual_id'),
            Trabajo.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            Trabajo.nombre_trabajo.label('concepto'),
            Trabajo.monto_total.label('monto_total'),
            func.coalesce(Trabajo.monto_pagado_usd, 0).label('monto_pagado'),
            (Trabajo.monto_total - func.coalesce(Trabajo.monto_pagado_usd, 0)).label('saldo_pendiente'),
            Trabajo.fecha_creacion.label('fecha'),
            literal(None).label('estado_deuda')
        ).join(Cliente, Trabajo.cliente_id == Cliente.id).where(
            Trabajo.porcentaje_pagado.in_([0, 50]),
            Cliente.activo == True
        )
        
        # 2. Obtener deudas manuales pendientes
        deudas_manuales_query = select(
            literal('manual').label('origen'),
            literal(None).label('trabajo_id'),
            DeudaManual.id.label('deuda_manual_id'),
            DeudaManual.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            DeudaManual.nombre_trabajo.label('concepto'),
            DeudaManual.monto_deuda.label('monto_total'),
            DeudaManual.monto_pagado.label('monto_pagado'),
            DeudaManual.saldo_pendiente.label('saldo_pendiente'),
            DeudaManual.fecha_registro.label('fecha'),
            DeudaManual.estado.label('estado_deuda')
        ).join(Cliente, DeudaManual.cliente_id == Cliente.id).where(
            DeudaManual.estado.in_(['pendiente', 'parcial']),
            Cliente.activo == True
        )
        
        # Aplicar filtros
        if search:
            search_filter = f"%{search}%"
            trabajos_query = trabajos_query.where(Cliente.nombre_razon_social.ilike(search_filter))
            deudas_manuales_query = deudas_manuales_query.where(Cliente.nombre_razon_social.ilike(search_filter))
        
        if origen == 'trabajo':
            deudas_manuales_query = deudas_manuales_query.where(False)
        elif origen == 'manual':
            trabajos_query = trabajos_query.where(False)
        
        # Ejecutar consultas
        trabajos_result = await db.execute(trabajos_query)
        deudas_result = await db.execute(deudas_manuales_query)
        
        resultados = list(trabajos_result.all()) + list(deudas_result.all())
        
        print(f"✅ Encontrados {len(resultados)} registros")
        
        # Formatear respuesta
        deudores = []
        total_adeudado = 0.0
        
        for r in resultados:
            saldo = float(r.saldo_pendiente or 0)
            total_adeudado += saldo
            
            deudores.append({
                "cliente_id": r.cliente_id,
                "cliente_nombre": r.cliente_nombre or "Sin nombre",
                "concepto": r.concepto or "Sin concepto",
                "monto_total": float(r.monto_total or 0),
                "monto_pagado": float(r.monto_pagado or 0),
                "saldo_pendiente": saldo,
                "fecha": r.fecha.isoformat() if r.fecha else datetime.now().isoformat(),
                "origen": r.origen,
                "trabajo_id": r.trabajo_id,
                "deuda_manual_id": r.deuda_manual_id,
                "estado_deuda": r.estado_deuda or "pendiente"
            })
        
        # Contar por origen
        trabajos_count = len([d for d in deudores if d["origen"] == "trabajo"])
        manuales_count = len([d for d in deudores if d["origen"] == "manual"])
        
        return {
            "total_registros": len(deudores),
            "total_clientes_unicos": len(set([d["cliente_id"] for d in deudores])),
            "total_adeudado": round(total_adeudado, 2),
            "deudores": deudores[:limit],
            "resumen_por_origen": {
                "trabajos": trabajos_count,
                "manuales": manuales_count
            }
        }
        
    except Exception as e:
        print(f"❌ Error en listar_deudores_pendientes: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener listado de deudores: {str(e)}"
        )


# ============================================
# ENDPOINT - POST /manual
# ============================================
@router.post("/manual", status_code=status.HTTP_201_CREATED)
async def crear_deuda_manual(
    deuda_data: dict,
    db: AsyncSession = Depends(get_db)
):
    """Crear una deuda manual - VERSIÓN SIMPLIFICADA"""
    try:
        print("📝 POST /deudas/manual - Creando deuda manual...")
        print(f"Datos recibidos: {deuda_data}")
        
        cliente_id = deuda_data.get("cliente_id")
        nombre_trabajo = deuda_data.get("nombre_trabajo")
        monto_deuda = deuda_data.get("monto_deuda")
        observaciones = deuda_data.get("observaciones")
        
        # Validar campos obligatorios
        if not cliente_id:
            raise HTTPException(status_code=400, detail="cliente_id es requerido")
        if not nombre_trabajo:
            raise HTTPException(status_code=400, detail="nombre_trabajo es requerido")
        if not monto_deuda or monto_deuda <= 0:
            raise HTTPException(status_code=400, detail="monto_deuda debe ser mayor a 0")
        
        # Validar que el cliente existe
        stmt = select(Cliente).where(Cliente.id == cliente_id, Cliente.activo == True)
        result = await db.execute(stmt)
        cliente = result.scalar_one_or_none()
        
        if not cliente:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cliente con ID {cliente_id} no encontrado"
            )
        
        # Crear deuda
        nueva_deuda = DeudaManual(
            cliente_id=cliente_id,
            nombre_trabajo=nombre_trabajo,
            monto_deuda=Decimal(str(monto_deuda)),
            saldo_pendiente=Decimal(str(monto_deuda)),
            monto_pagado=Decimal('0.00'),
            estado="pendiente",
            observaciones=observaciones,
            fecha_registro=datetime.utcnow()
        )
        
        db.add(nueva_deuda)
        await db.commit()
        await db.refresh(nueva_deuda)
        
        print(f"✅ Deuda manual creada con ID: {nueva_deuda.id}")
        
        return {
            "id": nueva_deuda.id,
            "cliente_id": nueva_deuda.cliente_id,
            "nombre_trabajo": nueva_deuda.nombre_trabajo,
            "monto_deuda": float(nueva_deuda.monto_deuda),
            "saldo_pendiente": float(nueva_deuda.saldo_pendiente),
            "monto_pagado": float(nueva_deuda.monto_pagado),
            "estado": nueva_deuda.estado,
            "fecha_registro": nueva_deuda.fecha_registro.isoformat(),
            "observaciones": nueva_deuda.observaciones,
            "es_convertida_a_trabajo": False,
            "trabajo_convertido_id": None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        print(f"❌ Error en crear_deuda_manual: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear la deuda manual: {str(e)}"
        )


# ============================================
# ENDPOINT - PUT /manual/{deuda_id}/pagar
# ============================================
@router.put("/manual/{deuda_id}/pagar")
async def registrar_pago_deuda_manual(
    deuda_id: int,
    pago_data: dict,
    db: AsyncSession = Depends(get_db)
):
    """Registrar un pago en una deuda manual"""
    try:
        print(f"💰 PUT /deudas/manual/{deuda_id}/pagar - Registrando pago...")
        
        monto_pago = pago_data.get("monto_pago")
        
        if not monto_pago or monto_pago <= 0:
            raise HTTPException(status_code=400, detail="monto_pago debe ser mayor a 0")
        
        # Buscar la deuda
        stmt = select(DeudaManual).where(DeudaManual.id == deuda_id)
        result = await db.execute(stmt)
        deuda = result.scalar_one_or_none()
        
        if not deuda:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Deuda manual con ID {deuda_id} no encontrada"
            )
        
        if deuda.estado == 'pagado':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Esta deuda ya está completamente pagada"
            )
        
        if monto_pago > float(deuda.saldo_pendiente):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El monto del pago ({monto_pago}) excede el saldo pendiente ({float(deuda.saldo_pendiente)})"
            )
        
        # Actualizar
        deuda.monto_pagado += Decimal(str(monto_pago))
        deuda.saldo_pendiente -= Decimal(str(monto_pago))
        
        if deuda.saldo_pendiente <= 0:
            deuda.estado = 'pagado'
        elif deuda.monto_pagado > 0:
            deuda.estado = 'parcial'
        
        deuda.fecha_actualizacion = datetime.utcnow()
        
        await db.commit()
        await db.refresh(deuda)
        
        print(f"✅ Pago registrado para deuda {deuda_id}")
        
        return {
            "id": deuda.id,
            "cliente_id": deuda.cliente_id,
            "nombre_trabajo": deuda.nombre_trabajo,
            "monto_deuda": float(deuda.monto_deuda),
            "saldo_pendiente": float(deuda.saldo_pendiente),
            "monto_pagado": float(deuda.monto_pagado),
            "estado": deuda.estado,
            "fecha_registro": deuda.fecha_registro.isoformat(),
            "observaciones": deuda.observaciones,
            "es_convertida_a_trabajo": False,
            "trabajo_convertido_id": None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        print(f"❌ Error en registrar_pago_deuda_manual: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al registrar el pago: {str(e)}"
        )


# ============================================
# ENDPOINT - GET /manual
# ============================================
@router.get("/manual")
async def obtener_deudas_manuales(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(100, description="Límite de registros")
):
    """Obtener lista de deudas manuales"""
    try:
        stmt = select(DeudaManual).limit(limit)
        result = await db.execute(stmt)
        deudas = result.scalars().all()
        
        return [
            {
                "id": d.id,
                "cliente_id": d.cliente_id,
                "nombre_trabajo": d.nombre_trabajo,
                "monto_deuda": float(d.monto_deuda),
                "saldo_pendiente": float(d.saldo_pendiente),
                "monto_pagado": float(d.monto_pagado),
                "estado": d.estado,
                "fecha_registro": d.fecha_registro.isoformat(),
                "observaciones": d.observaciones
            }
            for d in deudas
        ]
        
    except Exception as e:
        print(f"❌ Error en obtener_deudas_manuales: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener deudas manuales: {str(e)}"
        )
        

# app/routers/deudas.py - Agrega al final del archivo

# ============================================
# EXPORTAR REPORTE EN EXCEL
# ============================================
@router.get("/exportar/excel")
async def exportar_deudas_excel(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = Query(None, description="Búsqueda por cliente"),
    origen: Optional[str] = Query(None, description="Filtrar por origen: 'trabajo' o 'manual'")
):
    """Exportar listado de deudas a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse
        import io
        
        print("📊 Exportando deudas a Excel...")
        
        # Obtener los mismos datos que en el listado
        # 1. Trabajos con saldo pendiente
        trabajos_query = select(
            literal('trabajo').label('origen'),
            Trabajo.id.label('trabajo_id'),
            literal(None).label('deuda_manual_id'),
            Trabajo.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            Trabajo.nombre_trabajo.label('concepto'),
            Trabajo.monto_total.label('monto_total'),
            func.coalesce(Trabajo.monto_pagado_usd, 0).label('monto_pagado'),
            (Trabajo.monto_total - func.coalesce(Trabajo.monto_pagado_usd, 0)).label('saldo_pendiente'),
            Trabajo.fecha_creacion.label('fecha'),
            literal(None).label('estado_deuda')
        ).join(Cliente, Trabajo.cliente_id == Cliente.id).where(
            Trabajo.porcentaje_pagado.in_([0, 50]),
            Cliente.activo == True
        )
        
        # 2. Deudas manuales
        deudas_manuales_query = select(
            literal('manual').label('origen'),
            literal(None).label('trabajo_id'),
            DeudaManual.id.label('deuda_manual_id'),
            DeudaManual.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            DeudaManual.nombre_trabajo.label('concepto'),
            DeudaManual.monto_deuda.label('monto_total'),
            DeudaManual.monto_pagado.label('monto_pagado'),
            DeudaManual.saldo_pendiente.label('saldo_pendiente'),
            DeudaManual.fecha_registro.label('fecha'),
            DeudaManual.estado.label('estado_deuda')
        ).join(Cliente, DeudaManual.cliente_id == Cliente.id).where(
            DeudaManual.estado.in_(['pendiente', 'parcial']),
            Cliente.activo == True
        )
        
        # Aplicar filtros
        if search:
            search_filter = f"%{search}%"
            trabajos_query = trabajos_query.where(Cliente.nombre_razon_social.ilike(search_filter))
            deudas_manuales_query = deudas_manuales_query.where(Cliente.nombre_razon_social.ilike(search_filter))
        
        if origen == 'trabajo':
            deudas_manuales_query = deudas_manuales_query.where(False)
        elif origen == 'manual':
            trabajos_query = trabajos_query.where(False)
        
        # Ejecutar consultas
        trabajos_result = await db.execute(trabajos_query)
        deudas_result = await db.execute(deudas_manuales_query)
        
        resultados = list(trabajos_result.all()) + list(deudas_result.all())
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deudas Pendientes"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["#", "Cliente", "Concepto", "Origen", "Monto Total (USD)", "Monto Pagado (USD)", "Saldo Pendiente (USD)", "Estado", "Fecha"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_idx, r in enumerate(resultados, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=r.cliente_nombre or "Sin nombre").border = border
            ws.cell(row=row_idx, column=3, value=r.concepto or "Sin concepto").border = border
            ws.cell(row=row_idx, column=4, value=r.origen.upper()).border = border
            ws.cell(row=row_idx, column=5, value=float(r.monto_total or 0)).border = border
            ws.cell(row=row_idx, column=6, value=float(r.monto_pagado or 0)).border = border
            ws.cell(row=row_idx, column=7, value=float(r.saldo_pendiente or 0)).border = border
            ws.cell(row=row_idx, column=8, value=(r.estado_deuda or "pendiente").upper()).border = border
            ws.cell(row=row_idx, column=9, value=r.fecha.strftime("%d/%m/%Y %H:%M") if r.fecha else "").border = border
            
            # Colorear saldo pendiente
            saldo_cell = ws.cell(row=row_idx, column=7)
            if float(r.saldo_pendiente or 0) > 0:
                saldo_cell.font = Font(color="DC2626", bold=True)
            else:
                saldo_cell.font = Font(color="16A34A", bold=True)
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        filename = f"deudas_pendientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f"✅ Excel exportado: {filename}")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error al exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al exportar Excel: {str(e)}"
        )


# ============================================
# EXPORTAR REPORTE EN PDF
# ============================================
@router.get("/exportar/pdf")
async def exportar_deudas_pdf(
    db: AsyncSession = Depends(get_db),
    search: Optional[str] = Query(None, description="Búsqueda por cliente"),
    origen: Optional[str] = Query(None, description="Filtrar por origen: 'trabajo' o 'manual'")
):
    """Exportar listado de deudas a PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas
        from fastapi.responses import StreamingResponse
        import io
        
        print("📊 Exportando deudas a PDF...")
        
        # Obtener los mismos datos
        trabajos_query = select(
            literal('trabajo').label('origen'),
            Trabajo.id.label('trabajo_id'),
            literal(None).label('deuda_manual_id'),
            Trabajo.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            Trabajo.nombre_trabajo.label('concepto'),
            Trabajo.monto_total.label('monto_total'),
            func.coalesce(Trabajo.monto_pagado_usd, 0).label('monto_pagado'),
            (Trabajo.monto_total - func.coalesce(Trabajo.monto_pagado_usd, 0)).label('saldo_pendiente'),
            Trabajo.fecha_creacion.label('fecha'),
            literal(None).label('estado_deuda')
        ).join(Cliente, Trabajo.cliente_id == Cliente.id).where(
            Trabajo.porcentaje_pagado.in_([0, 50]),
            Cliente.activo == True
        )
        
        deudas_manuales_query = select(
            literal('manual').label('origen'),
            literal(None).label('trabajo_id'),
            DeudaManual.id.label('deuda_manual_id'),
            DeudaManual.cliente_id,
            Cliente.nombre_razon_social.label('cliente_nombre'),
            DeudaManual.nombre_trabajo.label('concepto'),
            DeudaManual.monto_deuda.label('monto_total'),
            DeudaManual.monto_pagado.label('monto_pagado'),
            DeudaManual.saldo_pendiente.label('saldo_pendiente'),
            DeudaManual.fecha_registro.label('fecha'),
            DeudaManual.estado.label('estado_deuda')
        ).join(Cliente, DeudaManual.cliente_id == Cliente.id).where(
            DeudaManual.estado.in_(['pendiente', 'parcial']),
            Cliente.activo == True
        )
        
        if search:
            search_filter = f"%{search}%"
            trabajos_query = trabajos_query.where(Cliente.nombre_razon_social.ilike(search_filter))
            deudas_manuales_query = deudas_manuales_query.where(Cliente.nombre_razon_social.ilike(search_filter))
        
        if origen == 'trabajo':
            deudas_manuales_query = deudas_manuales_query.where(False)
        elif origen == 'manual':
            trabajos_query = trabajos_query.where(False)
        
        trabajos_result = await db.execute(trabajos_query)
        deudas_result = await db.execute(deudas_manuales_query)
        resultados = list(trabajos_result.all()) + list(deudas_result.all())
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=30
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=20
        )
        
        # Elementos del PDF
        elements = []
        
        # Título
        elements.append(Paragraph("📊 Reporte de Deudas Pendientes", title_style))
        elements.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 20))
        
        # Totales
        total_deudas = len(resultados)
        total_adeudado = sum(float(r.saldo_pendiente or 0) for r in resultados)
        
        resumen_data = [
            ['Total Deudas', 'Total Adeudado', 'Clientes Deudores'],
            [str(total_deudas), f'${total_adeudado:,.2f}', str(len(set(r.cliente_id for r in resultados)))]
        ]
        
        resumen_table = Table(resumen_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(resumen_table)
        elements.append(Spacer(1, 30))
        
        # Tabla de deudas
        table_data = [
            ['#', 'Cliente', 'Concepto', 'Origen', 'Total', 'Pagado', 'Saldo', 'Estado', 'Fecha']
        ]
        
        for idx, r in enumerate(resultados, 1):
            estado = (r.estado_deuda or "pendiente").upper()
            if estado == "PAGADO":
                estado_color = "🟢 " + estado
            elif estado == "PARCIAL":
                estado_color = "🟡 " + estado
            else:
                estado_color = "🔴 " + estado
            
            table_data.append([
                str(idx),
                (r.cliente_nombre or "Sin nombre")[:30],
                (r.concepto or "Sin concepto")[:40],
                (r.origen or "").upper(),
                f"${float(r.monto_total or 0):,.2f}",
                f"${float(r.monto_pagado or 0):,.2f}",
                f"${float(r.saldo_pendiente or 0):,.2f}",
                estado_color,
                r.fecha.strftime("%d/%m/%Y") if r.fecha else ""
            ])
        
        # Crear tabla
        table = Table(table_data, colWidths=[0.4*inch, 1.8*inch, 2*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Pie de página
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94A3B8'),
            alignment=1
        )
        elements.append(Paragraph("SGOAP - Sistema de Gestión de Operaciones y Administración de Proyectos", footer_style))
        elements.append(Paragraph("Reporte generado automáticamente", footer_style))
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"deudas_pendientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        print(f"✅ PDF exportado: {filename}")
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Error al exportar PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al exportar PDF: {str(e)}"
        )